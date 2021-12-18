import os.path
import sys
import matplotlib
import pandas as pd
import numpy as np
import openpyxl as op
import matplotlib.pyplot as plt
import statsmodels.api as sm
import sklearn

from PySide2.QtCore import (
    Qt
)

from PySide2.QtWidgets import (
    QWidget,
    QApplication,
    QPushButton,
    QHBoxLayout,
    QVBoxLayout,
    QGridLayout,
    QSpacerItem,
    QListWidget,
    QListWidgetItem,
    # QFileDialog,
    QLabel,
    QTabWidget,
    # QMessageBox,
    QSpinBox,
    QDoubleSpinBox,
    QLineEdit,
    QComboBox,
    QCheckBox,
    QScrollArea
)

from PySide2.QtGui import (
    QIcon,
    QPixmap
)

from lib.core.project_flags import *
import lib.core.machineLearningRegression as mlr
import lib.core.signalCompare as signComp

from lib.gui.guiStyle import setStyle_
import lib.gui.commonFunctions as coFunc

# *************************************************************************************************** #

matplotlib.use("Agg")  # Set matplotlib to use non-interface (don't plot figures)


class WidgetMachineLearningRegressionWidget(QWidget):
    def __init__(self, w=512, h=512, minW=256, minH=256, maxW=None, maxH=None,
                 winTitle='My Window', iconPath=None):
        super().__init__()
        self.setStyleSheet(setStyle_())  # Set the styleSheet
        self.iconPath = iconPath

        # Set this flag to True to show debugging messages to console
        self.debugMessageFlag = False

        # -------------------------------- #
        # ----- Private QTabWidget ------- #
        # -------------------------------- #

        # DICTIONARY FILE PARAMETERS
        self._DKEY_FILE_NAME: str = 'name'
        self._DKEY_FULLPATH: str = 'full-path'
        self._DKEY_COLUMNS: str = 'columns'
        self._DKEY_INPUT_LIST: str = 'input-list'
        self._DKEY_OUTPUT_LIST: str = 'output-list'
        self._DKEY_PRIMARY_EVENT_COLUMN: str = 'primary-event'

        # DICTIONARY MACHINE LEARNING PARAMETERS
        self._DKEY_MLP_TEST_PERCENTAGE: str = 'test-percentage'
        self._DKEY_MLP_TEST_PERCENTAGE_DISTRIBUTION: str = 'test-percentage-distribution'
        self._DKEY_MLP_HOLDOUT_PERCENTAGE: str = 'holdout-percentage'
        self._DKEY_MLP_HOLDOUT_PERCENTAGE_DISTRIBUTION: str = 'holdout-percentage-distribution'
        self._DKEY_MLP_EXPORT_FOLDER: str = 'export-folder'
        self._DKEY_MLP_EXPER_NUMBER: str = 'experiment-number'
        self._DKEY_MLP_ML_METHOD: str = 'ml-method'
        self._DKEY_MLP_METHOD_INDEX: str = 'ml-type-index'
        self._DKEY_MLP_MULTIFILE_TRAINING_PROCESSING: str = 'multifile-training-processing'

        # -------------------------------- #
        # ----- Set QTabWidget ----------- #
        # -------------------------------- #

        self.mainTabWidget = QTabWidget()  # Create a Tab Widget
        self.widgetTabInputOutput = WidgetTabInputOutput()  # create a tab for input output columns
        self.widgetTabMachineLearningSettings = WidgetTabMachineLearningSettings()

        # ------------------------------- #
        # ----- Set WidgetMethods ------- #
        # ------------------------------- #
        _WIDGET_OPTIONS_SIZE = 480

        # ----- RIDGE ----- #
        self.widgetOptions_Ridge = WidgetRidgeML(w=_WIDGET_OPTIONS_SIZE, h=_WIDGET_OPTIONS_SIZE,
                                                 minW=_WIDGET_OPTIONS_SIZE, minH=_WIDGET_OPTIONS_SIZE,
                                                 maxW=_WIDGET_OPTIONS_SIZE, maxH=_WIDGET_OPTIONS_SIZE,
                                                 winTitle='Ridge Options', iconPath=self.iconPath)

        # ----- SVR ----- #
        self.widgetOptions_SVR = WidgetSVRML(w=_WIDGET_OPTIONS_SIZE, h=_WIDGET_OPTIONS_SIZE,
                                             minW=_WIDGET_OPTIONS_SIZE, minH=_WIDGET_OPTIONS_SIZE,
                                             maxW=_WIDGET_OPTIONS_SIZE, maxH=_WIDGET_OPTIONS_SIZE,
                                             winTitle='SVR Options', iconPath=self.iconPath)

        # ---------------------- #
        # ----- Set Window ----- #
        # ---------------------- #
        self.setWindowTitle(winTitle)  # Set Window Title
        self.setWindowIcon(QIcon(self.iconPath))  # Set Window Icon
        self.setGeometry(INT_SCREEN_WIDTH / 4, INT_SCREEN_HEIGHT / 4, w, h)  # Set Window Geometry
        self.setMinimumWidth(minW)  # Set Window Minimum Width
        self.setMinimumHeight(minH)  # Set Window Minimum Height
        if maxW is not None:
            self.setMaximumWidth(maxW)  # Set Window Maximum Width
        if maxH is not None:
            self.setMaximumHeight(maxH)  # Set Window Maximum Width

        self.vbox_main_layout = QVBoxLayout(self)  # Create the main vbox

        # -------------------------- #
        # ----- Set PushButton ----- #
        # -------------------------- #
        self.buttonAdd = QPushButton()  # Create button for Add
        self.buttonAdd.setMinimumWidth(INT_ADD_REMOVE_BUTTON_SIZE)  # Set Minimum Width
        self.buttonAdd.setMaximumWidth(INT_ADD_REMOVE_BUTTON_SIZE)  # Set Maximum Width
        self.buttonAdd.setMinimumHeight(INT_ADD_REMOVE_BUTTON_SIZE)  # Set Minimum Height
        self.buttonAdd.setMaximumHeight(INT_ADD_REMOVE_BUTTON_SIZE)  # Set Maximum Height
        self.buttonAdd.setIcon(QIcon(QPixmap(ICON_ADD)))  # Add Icon
        self.buttonAdd.setToolTip('Add table files.')  # Add Description

        self.buttonRemove = QPushButton()  # Create button for Remove
        self.buttonRemove.setMinimumWidth(INT_ADD_REMOVE_BUTTON_SIZE)  # Set Minimum Width
        self.buttonRemove.setMaximumWidth(INT_ADD_REMOVE_BUTTON_SIZE)  # Set Maximum Width
        self.buttonRemove.setMinimumHeight(INT_ADD_REMOVE_BUTTON_SIZE)  # Set Minimum Height
        self.buttonRemove.setMaximumHeight(INT_ADD_REMOVE_BUTTON_SIZE)  # Set Maximum Height
        self.buttonRemove.setIcon(QIcon(QPixmap(ICON_REMOVE)))  # Add Icon
        self.buttonRemove.setToolTip('Remove table files.')  # Add Description

        self.buttonExecute = QPushButton("Execute")
        self.buttonExecute.setMinimumWidth(0)  # Set Minimum Width
        self.buttonExecute.setMinimumHeight(INT_ADD_REMOVE_BUTTON_SIZE)  # Set Minimum Height
        self.buttonExecute.setToolTip('Run the machine learning process.')  # Add Description

        # -------------------------------- #
        # ----- Set QListWidgetItems ----- #
        # -------------------------------- #
        self.listWidget_FileList = QListWidget()  # Create a ListWidget
        self.listWidget_FileList.setMinimumWidth(300)
        self.listWidget_FileList.setMaximumWidth(500)
        self.listWidget_ColumnList = QListWidget()  # Create a ListWidget
        self.listWidget_ColumnList.setMinimumWidth(300)
        self.listWidget_ColumnList.setMaximumWidth(500)
        self.listWidget_ColumnList.setSelectionMode(QListWidget.ExtendedSelection)  # Set Extended Selection
        self.fileName = None

        # --------------------- #
        # ----- Variables ----- #
        # --------------------- #
        self.str_pathToTheProject = NEW_PROJECT_DEFAULT_FOLDER  # var to store the projectPath
        self.dict_tableFilesPaths = {}  # a dictionary to store the table files

        self.dict_machineLearningParameters = {
            self.dkey_mlpExperimentNumber(): self.widgetTabMachineLearningSettings.tabGeneral.getDefaultExperimentNumber(),
            self.dkey_mlpTestPercentage(): self.widgetTabMachineLearningSettings.tabGeneral.getDefaultTestPercentage(),
            self.dkey_mlpTestPercentageDistribution(): self.widgetTabMachineLearningSettings.tabGeneral.getDefaultTestPercentageDistribution(),
            self.dkey_mlpHoldoutPercentage(): self.widgetTabMachineLearningSettings.tabGeneral.getDefaultHoldoutPercentage(),
            self.dkey_mlpHoldoutPercentageDistribution(): self.widgetTabMachineLearningSettings.tabGeneral.getDefaultHoldoutPercentageDistribution(),
            self.dkey_mlpExportFolder(): self.widgetTabMachineLearningSettings.tabGeneral.getDefaultExportPath(),
            self.dkey_mlpMethod(): self.widgetTabMachineLearningSettings.tabGeneral.getDefaultMethod(),
            self.dkey_mlpMethodIndex(): self.widgetTabMachineLearningSettings.tabGeneral.getDefaultMethodIndex(),
            self.dkey_multifileTrainingProcessing(): self.widgetTabMachineLearningSettings.tabGeneral.getDefaultMultifileTrainingProcessing()
        }

        self.mlr_Regression = mlr.MachineLearningRegression()
        self.mlr_Regression.setMLR_dict()

        self.signComp_Methods = signComp.SignalCompare()
        self.signComp_Methods.setSC_dict()

    # DICTIONARY FILE PARAMETERS
    def dkeyFileName(self):
        return self._DKEY_FILE_NAME

    def dkeyFullPath(self):
        return self._DKEY_FULLPATH

    def dkeyColumnsList(self):
        return self._DKEY_COLUMNS

    def dkeyInputList(self):
        return self._DKEY_INPUT_LIST

    def dkeyOutputList(self):
        return self._DKEY_OUTPUT_LIST

    def dkeyPrimaryEventColumn(self):
        return self._DKEY_PRIMARY_EVENT_COLUMN

    # DICTIONARY MACHINE LEARNING PARAMETERS
    def dkey_mlpTestPercentage(self):
        return self._DKEY_MLP_TEST_PERCENTAGE

    def dkey_mlpTestPercentageDistribution(self):
        return self._DKEY_MLP_TEST_PERCENTAGE_DISTRIBUTION

    def dkey_mlpExportFolder(self):
        return self._DKEY_MLP_EXPORT_FOLDER

    def dkey_mlpHoldoutPercentage(self):
        return self._DKEY_MLP_HOLDOUT_PERCENTAGE

    def dkey_mlpHoldoutPercentageDistribution(self):
        return self._DKEY_MLP_HOLDOUT_PERCENTAGE_DISTRIBUTION

    def dkey_mlpExperimentNumber(self):
        return self._DKEY_MLP_EXPER_NUMBER

    def dkey_mlpMethod(self):
        return self._DKEY_MLP_ML_METHOD

    def dkey_mlpMethodIndex(self):
        return self._DKEY_MLP_METHOD_INDEX

    def dkey_multifileTrainingProcessing(self):
        return self._DKEY_MLP_MULTIFILE_TRAINING_PROCESSING

    # --------------------------- #
    # ----- Reuse Functions ----- #
    # --------------------------- #
    def set_IO_Tab(self):
        # Set main Tab Widget
        self.widgetTabInputOutput.setWidget()  # Set the Tab File Management Widget
        self.mainTabWidget.addTab(self.widgetTabInputOutput, "Input/Output Management")  # Add it to mainTanWidget

    def set_MLR_Defaults(self):
        self.widgetOptions_Ridge.setWidget()
        self.setWidgetRidgeDefaultValues()

        self.widgetOptions_SVR.setWidget()
        self.setWidgetSVRDefaultValues()

    def setWidget(self):
        """
        A function to create the widget components into the main QWidget
        :return: Nothing
        """

        self.set_IO_Tab()
        self.set_MLR_Defaults()
        self.widgetTabMachineLearningSettings.setWidget()
        self.mainTabWidget.addTab(self.widgetTabMachineLearningSettings, "Machine Learning Settings")

        # Disable Generate Button
        self.buttonExecute.setEnabled(False)

        # Set Column vbox
        labelColumnList = QLabel("Column List:")
        vbox_listColumns = QVBoxLayout()  # Create a Horizontal Box Layout
        vbox_listColumns.addWidget(labelColumnList)  # Add Label
        vbox_listColumns.addWidget(self.listWidget_ColumnList)  # Add Column List

        # Set add/remove button in vbox
        hbox_listFileButtons = QHBoxLayout()  # Create a Horizontal Box Layout
        hbox_listFileButtons.addWidget(self.buttonAdd)  # Add buttonAdd
        hbox_listFileButtons.addWidget(self.buttonRemove)  # Add buttonRemove
        hbox_listFileButtons.addWidget(self.buttonExecute)  # Add buttonGenerate

        # Set FileList in hbox
        labelFileList = QLabel("Opened File List:")
        vbox_listFile = QVBoxLayout()  # Create a Vertical Box Layout
        vbox_listFile.addWidget(labelFileList)  # Add Label
        vbox_listFile.addWidget(self.listWidget_FileList)  # Add FileList
        vbox_listFile.addLayout(vbox_listColumns)  # Add listColumns
        vbox_listFile.addLayout(hbox_listFileButtons)  # Add vbox_listFileButtons layout

        # Set List and Tab Widget Layout
        hbox_final_layout = QHBoxLayout()  # Create a Horizontal Box Layout
        hbox_final_layout.addLayout(vbox_listFile)  # Add the listFile layout to finalLayout
        hbox_final_layout.addWidget(self.mainTabWidget)  # Add the mainTabWidget to finalLayout

        # Set Main Layout
        self.vbox_main_layout.addLayout(hbox_final_layout)  # Add the final layout to mainLayout

        # ----------------------- #
        # ----- Set Actions ----- #
        # ----------------------- #
        self.setMainEvents_()  # Set the events/actions of buttons, listWidgets, etc., components

    def setWidgetRidgeDefaultValues(self):
        self.widgetOptions_Ridge.setOptionList_Tol(listValue=mlr.MLR_TOL_LIST)
        self.widgetOptions_Ridge.setOptionList_Solver(listValue=mlr.MLR_SOLVER_OPTIONS)
        self.widgetOptions_Ridge.setAlphaMin(value=self.mlr_Regression.getRidge_alphaMin_Default())
        self.widgetOptions_Ridge.setAlphaMax(value=self.mlr_Regression.getRidge_alphaMax_Default())
        self.widgetOptions_Ridge.setAlphaStep(value=self.mlr_Regression.getRidge_alphaStep_Default())
        self.widgetOptions_Ridge.setSelectedList_Tol(listValue=self.mlr_Regression.getRidge_Tol_Default())
        self.widgetOptions_Ridge.setSelectedList_Solver(listValue=self.mlr_Regression.getRidge_Solver_Default())

    def setWidgetSVRDefaultValues(self):
        self.widgetOptions_SVR.setOptionList_Kernel(listValue=mlr.MLR_KERNEL_OPTIONS)
        self.widgetOptions_SVR.setOptionList_Gamma(listValue=mlr.MLR_GAMMA_OPTIONS)
        self.widgetOptions_SVR.setOptionList_Tol(listValue=mlr.MLR_TOL_LIST)

        self.widgetOptions_SVR.setSelectedList_Kernel(listValue=self.mlr_Regression.getSVR_Kernel_Default())
        self.widgetOptions_SVR.setSelectedList_Gamma(listValue=self.mlr_Regression.getSVR_Gamma_Default())
        self.widgetOptions_SVR.setSelectedList_Tol(listValue=self.mlr_Regression.getSVR_Tol_Default())

    # ---------------------------------- #
    # ----- Reuse Action Functions ----- #
    # ---------------------------------- #

    def addItemsToList(self, fullPath):
        fileName = fullPath.split('/')[-1:][0]  # find the name of the file
        # Create the dictionary
        self.dict_tableFilesPaths[fileName] = {self.dkeyFileName(): fileName,
                                               self.dkeyFullPath(): fullPath,
                                               self.dkeyColumnsList(): file_manip.getColumnNames(fullPath),
                                               self.dkeyInputList(): [],
                                               self.dkeyOutputList(): [],
                                               self.dkeyPrimaryEventColumn(): None
                                               }
        self.listWidget_FileList.addItem(QListWidgetItem(fileName))  # Add Item to List

    # -------------------------------- #
    # ----- Print/Show Functions ----- #
    # -------------------------------- #
    def prt_dict_tableFilePaths(self):
        """
        Print the dictionary tableFilesPaths values
        :return: Nothing
        """
        for key in self.dict_tableFilesPaths.keys():
            print("file-key: ", key)
            for sec_key in self.dict_tableFilesPaths[key].keys():
                print(str(sec_key) + ': ', self.dict_tableFilesPaths[key][sec_key])
            print()

    # ------------------ #
    # ----- Events ----- #
    # ------------------ #
    # ***** SET EVENTS FUNCTIONS ***** #
    def setEvents_(self):
        # buttonInputColumn
        self.widgetTabInputOutput.buttonInputColumn.clicked.connect(self.actionButtonInput)
        # buttonRemInputColumn
        self.widgetTabInputOutput.buttonRemInputColumn.clicked.connect(self.actionButtonRemInput)
        # buttonOutputColumn
        self.widgetTabInputOutput.buttonOutputColumn.clicked.connect(self.actionButtonOutput)
        # buttonRemOutputColumn
        self.widgetTabInputOutput.buttonRemOutputColumn.clicked.connect(self.actionButtonRemOutput)
        # buttonPrimaryEvent
        self.widgetTabInputOutput.buttonPrimaryEvent.clicked.connect(self.actionButtonPrimaryEvent)
        # buttonRemPrimaryEvent
        self.widgetTabInputOutput.buttonRemPrimaryEvent.clicked.connect(self.actionButtonRemPrimaryEvent)

    def setTabSettingsGeneralEvents_(self):
        # Spin Boxes Events
        self.widgetTabMachineLearningSettings.tabGeneral.spinBox_ExperimentNumber.valueChanged.connect(
            self.actionExperimentResultChange)

        self.widgetTabMachineLearningSettings.tabGeneral.spinBox_MachineLearningMethodsIndex.valueChanged.connect(
            self.actionMachineLearningMethodIndexChange)

        # Double Spin Boxes  Events
        self.widgetTabMachineLearningSettings.tabGeneral.doubleSpinBox_TestPercentage.valueChanged.connect(
            self.actionTestPercentageChange)

        self.widgetTabMachineLearningSettings.tabGeneral.doubleSpinBox_HoldoutPercentage.valueChanged.connect(
            self.actionHoldoutPercentageChange)

        # Line Edit Events
        self.widgetTabMachineLearningSettings.tabGeneral.lineEdit_SetOutPath.textChanged.connect(
            self.actionLineEditChange)

        # Button Events
        self.widgetTabMachineLearningSettings.tabGeneral.buttonSetOutPath.clicked.connect(
            self.actionButtonSetOutPathClicked)

        # Combo Box Events
        self.widgetTabMachineLearningSettings.tabGeneral.comboBox_MachineLearningMethods.currentTextChanged.connect(
            self.actionMachineLearningMethodChange)

        self.widgetTabMachineLearningSettings.tabGeneral.comboBox_MultifileTrainingProcessing.currentTextChanged.connect(
            self.actionMultifileTrainingProcessingChange)

        self.widgetTabMachineLearningSettings.tabGeneral.comboBox_TestPercentageDistribution.currentTextChanged.connect(
            self.actionTestPercentageDistributionChange)

        self.widgetTabMachineLearningSettings.tabGeneral.comboBox_HoldoutPercentageDistribution.currentTextChanged.connect(
            self.actionHoldoutPercentageDistributionChange)

    def setTabSettingsRegressionMethodsEvents_(self):
        # ChangeState -> LinearRegression
        self.widgetTabMachineLearningSettings.tabRegressionMethods.checkbox_LinearRegression.stateChanged.connect(
            self.actionStateChangeLinearRegression)
        # ChangeState -> Ridge
        self.widgetTabMachineLearningSettings.tabRegressionMethods.checkbox_Ridge.stateChanged.connect(
            self.actionStateChangeRidge)
        # ChangeState -> BayesianRidge
        self.widgetTabMachineLearningSettings.tabRegressionMethods.checkbox_BayesianRidge.stateChanged.connect(
            self.actionStateChangeBayesianRidge)
        # ChangeState -> Lasso
        self.widgetTabMachineLearningSettings.tabRegressionMethods.checkbox_Lasso.stateChanged.connect(
            self.actionStateChangeLasso)
        # ChangeState -> LassoLars
        self.widgetTabMachineLearningSettings.tabRegressionMethods.checkbox_LassoLars.stateChanged.connect(
            self.actionStateChangeLassoLars)
        # ChangeState -> TweedieRegressor
        self.widgetTabMachineLearningSettings.tabRegressionMethods.checkbox_TweedieRegressor.stateChanged.connect(
            self.actionStateChangeTweedieRegressor)
        # ChangeState -> SGDRegressor
        self.widgetTabMachineLearningSettings.tabRegressionMethods.checkbox_SGDRegressor.stateChanged.connect(
            self.actionStateChangeSGDRegressor)
        # ChangeState -> SVR
        self.widgetTabMachineLearningSettings.tabRegressionMethods.checkbox_SVR.stateChanged.connect(
            self.actionStateChangeSVR)
        # ChangeState -> LinearSVR
        self.widgetTabMachineLearningSettings.tabRegressionMethods.checkbox_LinearSVR.stateChanged.connect(
            self.actionStateChangeLinearSVR)
        # ChangeState -> NearestNeighbor
        self.widgetTabMachineLearningSettings.tabRegressionMethods.checkbox_NearestNeighbor.stateChanged.connect(
            self.actionStateChangeNearestNeighbor)
        # ChangeState -> KNearestRegressor
        self.widgetTabMachineLearningSettings.tabRegressionMethods.checkbox_KNeighborsRegressor.stateChanged.connect(
            self.actionStateChangeKNeighborsRegressor)
        # ChangeState -> DecisionTreeRegressor
        self.widgetTabMachineLearningSettings.tabRegressionMethods.checkbox_DecisionTreeRegressor.stateChanged.connect(
            self.actionStateChangeDecisionTreeRegressor)
        # ChangeState -> RandomForestRegressor
        self.widgetTabMachineLearningSettings.tabRegressionMethods.checkbox_RandomForestRegressor.stateChanged.connect(
            self.actionStateChangeRandomForestRegressor)
        # ChangeState -> AdaBoostRegressor
        self.widgetTabMachineLearningSettings.tabRegressionMethods.checkbox_AdaBoostRegressor.stateChanged.connect(
            self.actionStateChangeAdaBoostRegressor)
        # ChangeState -> GradientBoostingRegressor
        self.widgetTabMachineLearningSettings.tabRegressionMethods.checkbox_GradientBoostingRegressor.stateChanged.connect(
            self.actionStateChangeGradientBoostingRegressor)

        # ButtonClicked (Options) -> Ridge
        self.widgetTabMachineLearningSettings.tabRegressionMethods.button_Ridge.clicked.connect(
            self.actionButtonClickedRidge)
        # ButtonClicked (Options) -> BayesianRidge
        self.widgetTabMachineLearningSettings.tabRegressionMethods.button_BayesianRidge.clicked.connect(
            self.actionButtonClickedBayesianRidge)
        # ButtonClicked (Options) -> Lasso
        self.widgetTabMachineLearningSettings.tabRegressionMethods.button_Lasso.clicked.connect(
            self.actionButtonClickedLasso)
        # ButtonClicked (Options) -> LassoLars
        self.widgetTabMachineLearningSettings.tabRegressionMethods.button_LassoLars.clicked.connect(
            self.actionButtonClickedLassoLars)
        # ButtonClicked (Options) -> TweedieRegressor
        self.widgetTabMachineLearningSettings.tabRegressionMethods.button_TweedieRegressor.clicked.connect(
            self.actionButtonClickedTweedieRegressor)
        # ButtonClicked (Options) -> SGDRegressor
        self.widgetTabMachineLearningSettings.tabRegressionMethods.button_SGDRegressor.clicked.connect(
            self.actionButtonClickedSGDRegressor)
        # ButtonClicked (Options) -> SVR
        self.widgetTabMachineLearningSettings.tabRegressionMethods.button_SVR.clicked.connect(
            self.actionButtonClickedSVR)
        # ButtonClicked (Options) -> LinearSVR
        self.widgetTabMachineLearningSettings.tabRegressionMethods.button_LinearSVR.clicked.connect(
            self.actionButtonClickedLinearSVR)
        # ButtonClicked (Options) -> NearestNeighbor
        self.widgetTabMachineLearningSettings.tabRegressionMethods.button_NearestNeighbor.clicked.connect(
            self.actionButtonClickedNearestNeighbor)
        # ButtonClicked (Options) -> KNearestRegressor
        self.widgetTabMachineLearningSettings.tabRegressionMethods.button_KNeighborsRegressor.clicked.connect(
            self.actionButtonClickedKNeighborsRegressor)
        # ButtonClicked (Options) -> DecisionTreeRegressor
        self.widgetTabMachineLearningSettings.tabRegressionMethods.button_DecisionTreeRegressor.clicked.connect(
            self.actionButtonClickedDecisionTreeRegressor)
        # ButtonClicked (Options) -> RandomForestRegressor
        self.widgetTabMachineLearningSettings.tabRegressionMethods.button_RandomForestRegressor.clicked.connect(
            self.actionButtonClickedRandomForestRegressor)
        # ButtonClicked (Options) -> AdaBoostRegressor
        self.widgetTabMachineLearningSettings.tabRegressionMethods.button_AdaBoostRegressor.clicked.connect(
            self.actionButtonClickedAdaBoostRegressor)
        # ButtonClicked (Options) -> GradientBoostingRegressor
        self.widgetTabMachineLearningSettings.tabRegressionMethods.button_GradientBoostingRegressor.clicked.connect(
            self.actionButtonClickedGradientBoostingRegressor)

    def setTabSettingsDeepRegressionMethodsEvents_(self):
        # ChangeState -> Covid_Convolutional_1D_LongShortTermMemory
        self.widgetTabMachineLearningSettings.tabDeepRegressionMethods.checkbox_Covid_Convolutional_1D_LongShortTermMemory.stateChanged.connect(
            self.actionStateChange_Covid_DeepNeuralNetwork)
        # ChangeState -> Covid_LongShortTermMemoryNeuralNetwork
        self.widgetTabMachineLearningSettings.tabDeepRegressionMethods.checkbox_Covid_LongShortTermMemoryNeuralNetwork.stateChanged.connect(
            self.actionStateChange_Covid_LongShortTermMemoryNeuralNetwork)
        # ChangeState -> Covid_LongShortTermMemoryNeuralNetwork_Simple
        self.widgetTabMachineLearningSettings.tabDeepRegressionMethods.checkbox_Covid_LongShortTermMemoryNeuralNetwork_Simple.stateChanged.connect(
            self.actionStateChange_Covid_LongShortTermMemoryNeuralNetwork_Simple)
        # ChangeState -> Covid_SimpleRecurrentNeuralNetwork
        self.widgetTabMachineLearningSettings.tabDeepRegressionMethods.checkbox_Covid_SimpleRecurrentNeuralNetwork.stateChanged.connect(
            self.actionStateChange_Covid_SimpleRecurrentNeuralNetwork)

    def setTabSettingsSignalCompareEvents_(self):
        self.widgetTabMachineLearningSettings.tabSignalCompare.checkbox_PearsonCorr.stateChanged.connect(
            self.actionStateChangePearsonCorr)
        self.widgetTabMachineLearningSettings.tabSignalCompare.checkbox_TimeLaggedCrossCorrelation.stateChanged.connect(
            self.actionStateChangeTimeLagCrossCorrelation)
        self.widgetTabMachineLearningSettings.tabSignalCompare.checkbox_TimeLaggedCrossCorrelationNoSplits.stateChanged.connect(
            self.actionStateChangeTimeLagCrossCorrelationNoSplits)
        self.widgetTabMachineLearningSettings.tabSignalCompare.checkbox_RollingWindowTimeLaggedCrossCorrelation.stateChanged.connect(
            self.actionStateChangeRollingWindowTimeLagCrossCorrelation)
        self.widgetTabMachineLearningSettings.tabSignalCompare.checkbox_DynamicTimeWarping.stateChanged.connect(
            self.actionStateChangeDynamicTimeWarping)

    # Set event for RidgeOptions
    def setWidgetRidgeEvents_(self):
        # Set Button Events
        self.widgetOptions_Ridge.button_TolAdd.clicked.connect(self.actionButtonClicked_Ridge_TolAdd)
        self.widgetOptions_Ridge.button_SolverAdd.clicked.connect(self.actionButtonClicked_Ridge_SolverAdd)
        self.widgetOptions_Ridge.button_TolRemove.clicked.connect(
            self.widgetOptions_Ridge.removeItemFromList_TolSelected)
        self.widgetOptions_Ridge.button_SolverRemove.clicked.connect(
            self.widgetOptions_Ridge.removeItemFromList_SolverSelected)
        self.widgetOptions_Ridge.button_RestoreDefault.clicked.connect(self.setWidgetRidgeDefaultValues)

        # Set ChangeValue Events
        self.widgetOptions_Ridge.doubleSpinBox_AlphaMin.valueChanged.connect(self.actionChange_Ridge_AlphaMin)
        self.widgetOptions_Ridge.doubleSpinBox_AlphaMax.valueChanged.connect(self.actionChange_Ridge_AlphaMax)
        self.widgetOptions_Ridge.doubleSpinBox_AlphaStep.valueChanged.connect(self.actionChange_Ridge_AlphaStep)
        self.widgetOptions_Ridge.listWidget_TolSelectedList.model().rowsInserted.connect(
            self.actionChange_Ridge_SelectedTol)
        self.widgetOptions_Ridge.listWidget_TolSelectedList.model().rowsRemoved.connect(
            self.actionChange_Ridge_SelectedTol)
        self.widgetOptions_Ridge.listWidget_SolverSelectedList.model().rowsInserted.connect(
            self.actionChange_Ridge_SelectedSolver)
        self.widgetOptions_Ridge.listWidget_SolverSelectedList.model().rowsRemoved.connect(
            self.actionChange_Ridge_SelectedSolver)

    # Set event for SVROptions
    def setWidgetSVREvents_(self):
        # Set Button Events
        self.widgetOptions_SVR.button_KernelAdd.clicked.connect(self.actionButtonClicked_SVR_KernelAdd)
        self.widgetOptions_SVR.button_GammaAdd.clicked.connect(self.actionButtonClicked_SVR_GammaAdd)
        self.widgetOptions_SVR.button_TolAdd.clicked.connect(self.actionButtonClicked_SVR_TolAdd)

        self.widgetOptions_SVR.button_KernelRemove.clicked.connect(
            self.widgetOptions_SVR.removeItemFromList_KernelSelected)
        self.widgetOptions_SVR.button_GammaRemove.clicked.connect(
            self.widgetOptions_SVR.removeItemFromList_GammaSelected)
        self.widgetOptions_SVR.button_TolRemove.clicked.connect(
            self.widgetOptions_SVR.removeItemFromList_TolSelected)

        self.widgetOptions_SVR.button_RestoreDefault.clicked.connect(self.setWidgetSVRDefaultValues)

        # Set ChangeValue Events
        self.widgetOptions_SVR.listWidget_KernelSelectedList.model().rowsInserted.connect(
            self.actionChange_SVR_SelectedKernel)
        self.widgetOptions_SVR.listWidget_KernelSelectedList.model().rowsRemoved.connect(
            self.actionChange_SVR_SelectedKernel)
        self.widgetOptions_SVR.listWidget_GammaSelectedList.model().rowsInserted.connect(
            self.actionChange_SVR_SelectedGamma)
        self.widgetOptions_SVR.listWidget_GammaSelectedList.model().rowsRemoved.connect(
            self.actionChange_SVR_SelectedGamma)
        self.widgetOptions_SVR.listWidget_TolSelectedList.model().rowsInserted.connect(
            self.actionChange_SVR_SelectedTol)
        self.widgetOptions_SVR.listWidget_TolSelectedList.model().rowsRemoved.connect(
            self.actionChange_SVR_SelectedTol)

    # ***** MAIN EVENTS FUNCTION ***** #
    def setMainEvents_(self):
        # Button Events
        self.buttonAdd.clicked.connect(self.actionButtonAdd)  # buttonAdd -> clicked
        self.buttonRemove.clicked.connect(self.actionButtonRemove)  # buttonRemove -> clicked
        self.buttonExecute.clicked.connect(self.actionButtonExecute)  # buttonGenerate -> clicked
        # ListWidget Events
        self.listWidget_FileList.currentRowChanged.connect(self.actionFileListRowChanged_event)

        self.setEvents_()  # set the user specified event (inherited)
        self.setTabSettingsGeneralEvents_()  # set the tab settings GENERAL events
        self.setTabSettingsRegressionMethodsEvents_()  # set the tab settings REGRESSION METHODS events
        self.setTabSettingsDeepRegressionMethodsEvents_()  # set the tab settings DEEP REGRESSION METHODS events
        self.setTabSettingsSignalCompareEvents_()  # set the tab settings SIGNAL COMPARE events

        self.setWidgetRidgeEvents_()  # set the events of Ridge Options
        self.setWidgetSVREvents_()  # set the events of SVR Options

    # -------------------------- #
    # ----- Events Actions ----- #
    # -------------------------- #
    # ***** SET MAIN EVENTS ACTIONS *** #
    def actionButtonAdd(self):
        # Open dialog
        success, dialog = coFunc.openFileDialog(
            classRef=self,
            dialogName='Open Table File',
            dialogOpenAt=self.str_pathToTheProject,
            dialogFilters=["CSV File Format (*.csv)",
                           "Microsoft Office Excel File (*.xlsx)"],
            dialogMultipleSelection=True)

        if success:  # if True
            for filePath in dialog.selectedFiles():  # for each file in all selected files
                fileName = filePath.split('/')[-1:][0]  # get the filename
                # print(fullPath)
                # print(fileName)
                # print()
                if fileName not in self.dict_tableFilesPaths.keys():  # if file haven't added before
                    self.addItemsToList(filePath)  # add file to the table list

            if self.listWidget_FileList.currentItem() is None:  # Set row 0 as current row
                self.listWidget_FileList.setCurrentRow(0)  # Set current row

            if self.dict_tableFilesPaths.keys().__len__() >= 1:
                self.buttonExecute.setEnabled(True)
            # self.prt_dict_tableFilePaths()

    def updateButtonRemove(self):
        self.updateInputList()
        self.updateOutputList()
        self.updatePrimaryEventList()

    def actionButtonRemove(self):
        if self.listWidget_FileList.currentItem() is not None:  # if some item is selected
            self.dict_tableFilesPaths.pop(self.fileName, None)  # Delete item from dict
            self.listWidget_FileList.takeItem(self.listWidget_FileList.currentRow())  # Delete item from widget
            self.actionFileListRowChanged_event()  # run the row changed event
            self.updateButtonRemove()

            # if there are not enough files loaded
            if self.dict_tableFilesPaths.keys().__len__() < 1:
                self.buttonExecute.setEnabled(False)  # disable the Execute Button

    # *********************************************************** #
    # *********** Helping Functions for ButtonExecute *********** #
    # *********************************************************** #
    # *                                                         * #

    def BE_errorExist(self, fName):
        errorType404 = "Error 404: "  # set error type
        # Check if Input Columns don't exist for the specified file
        if not self.dict_tableFilesPaths[fName][self.dkeyInputList()]:  # if True
            msg = "In < " + fName + " > no INPUT columns found!"  # set message
            coFunc.consoleMessage(errorType404 + msg)  # print message to Console as Warning
            coFunc.errorMessageDialog(classRef=self,
                                      errorType=errorType404,
                                      textMessageInfo=msg,
                                      iconPath=self.iconPath)  # print message to Error Dialog
            return True  # return True
        # Check if Output Columns don't exist for the specified file
        elif not self.dict_tableFilesPaths[fName][self.dkeyOutputList()]:  # if True
            msg = "In < " + fName + " > no OUTPUT columns found!"  # set message
            coFunc.consoleMessage(errorType404 + msg)  # print message to Console as Warning
            coFunc.errorMessageDialog(classRef=self,
                                      errorType=errorType404,
                                      textMessageInfo=msg,
                                      iconPath=self.iconPath)  # print message to Error Dialog
            return True  # return True
        return False  # return False

    def BE_readFileData(self, fName):
        # find the file suffix (extension) and take is as lowercase without the comma
        suffix = os.path.splitext(self.dict_tableFilesPaths[fName][self.dkeyFullPath()])[1].lower().split('.')[1]
        # create a variable named fileData and set it to None (it will store panda array if all goes well)
        fileData = None
        # create a list to store all the needed columns (we don't need columns we will not use - RAM saving)
        fileDataUseColumns = []
        # Check if user selected primary event column
        if self.dict_tableFilesPaths[fName][self.dkeyPrimaryEventColumn()] is not None:
            fileDataUseColumns.append(self.dict_tableFilesPaths[fName][self.dkeyPrimaryEventColumn()])  # add it to list
            # print(self.dict_tableFilesPaths[fName][self.dkeyPrimaryEventColumn()])
        # append to list (if not already) the values from input column list
        [fileDataUseColumns.append(col) for col in self.dict_tableFilesPaths[fName][self.dkeyInputList()]
         if col not in fileDataUseColumns]
        # append to list (if not already) the values from output column list
        [fileDataUseColumns.append(col) for col in self.dict_tableFilesPaths[fName][self.dkeyOutputList()]
         if col not in fileDataUseColumns]
        # print(fileDataUseColumns)

        print("Read " + suffix + " file.")  # info message

        # Read the file Data - CSV
        if suffix == 'csv':
            # read only the needed columns
            fileData = pd.read_csv(self.dict_tableFilesPaths[fName][self.dkeyFullPath()])[fileDataUseColumns]
            # print(fileData.keys())
        elif suffix == 'xlsx':
            # read only the needed columns
            fileData = pd.read_excel(self.dict_tableFilesPaths[fName][self.dkeyFullPath()])[fileDataUseColumns]

        # print(fileData)

        if fileData is not None:  # if file has been read successfully
            fileData.fillna(method='ffill', inplace=True)  # fill na values using ffill
            fileData.fillna(method='bfill', inplace=True)  # fill na values using bfill
        return fileData  # return the fileData

    def BE_setTrainValTestArrays(self, dictDataInput: {}, dictDataOutput: {}, inputHeaders: [], outputHeaders: []):
        def getTrainValTest(inputArr, outputArr):
            datasetSize = inputArr.__len__()
            testPercentage = self.dict_machineLearningParameters[self.dkey_mlpTestPercentage()]
            X_train_val_ = []
            y_train_val_ = []
            X_test_ = []
            y_test_ = []
            testTrainIndex = None

            if self.dict_machineLearningParameters[
                self.dkey_mlpTestPercentageDistribution()] == MLPF_PERCENTAGE_DISTRIBUTION_RANDOM_FROM_START:
                permIndexes = np.random.permutation(datasetSize).tolist()
                trainIdxs = permIndexes[int(testPercentage * datasetSize):]
                testIdxs = permIndexes[:int(testPercentage * datasetSize)]

                X_train_val_ = np.array(inputArr)[trainIdxs]
                y_train_val_ = np.array(outputArr)[trainIdxs]
                X_test_ = np.array(inputArr)[testIdxs]
                y_test_ = np.array(outputArr)[testIdxs]

            elif self.dict_machineLearningParameters[
                self.dkey_mlpTestPercentageDistribution()] == MLPF_PERCENTAGE_DISTRIBUTION_RANDOM_FROM_MIDDLE:
                permIndexes = np.random.permutation(datasetSize).tolist()
                midIndex = datasetSize / 2
                sliceSize = testPercentage * datasetSize
                startIndex = int(midIndex - sliceSize / 2)
                endIndex = int(midIndex - sliceSize / 2)

                trainIdxs = permIndexes[:startIndex] + permIndexes[-endIndex:]
                testIdxs = permIndexes[startIndex:-endIndex]

                X_train_val_ = np.array(inputArr)[trainIdxs]
                y_train_val_ = np.array(outputArr)[trainIdxs]
                X_test_ = np.array(inputArr)[testIdxs]
                y_test_ = np.array(outputArr)[testIdxs]

            elif self.dict_machineLearningParameters[
                self.dkey_mlpTestPercentageDistribution()] == MLPF_PERCENTAGE_DISTRIBUTION_RANDOM_FROM_END:
                permIndexes = np.random.permutation(datasetSize).tolist()
                trainIdxs = permIndexes[:-int(testPercentage * datasetSize)]
                testIdxs = permIndexes[-int(testPercentage * datasetSize):]

                X_train_val_ = np.array(inputArr)[trainIdxs]
                y_train_val_ = np.array(outputArr)[trainIdxs]
                X_test_ = np.array(inputArr)[testIdxs]
                y_test_ = np.array(outputArr)[testIdxs]

            elif self.dict_machineLearningParameters[
                self.dkey_mlpTestPercentageDistribution()] == MLPF_PERCENTAGE_DISTRIBUTION_FROM_START:
                datasetIndexes = list(range(datasetSize))
                trainIdxs = datasetIndexes[int(testPercentage * datasetSize):]
                testIdxs = datasetIndexes[:int(testPercentage * datasetSize)]
                testTrainIndex = [testPercentage * datasetSize]

                X_train_val_ = np.array(inputArr)[trainIdxs]
                y_train_val_ = np.array(outputArr)[trainIdxs]
                X_test_ = np.array(inputArr)[testIdxs]
                y_test_ = np.array(outputArr)[testIdxs]

            elif self.dict_machineLearningParameters[
                self.dkey_mlpTestPercentageDistribution()] == MLPF_PERCENTAGE_DISTRIBUTION_FROM_MIDDLE:
                datasetIndexes = list(range(datasetSize))
                midIndex = datasetSize / 2
                sliceSize = testPercentage * datasetSize
                startIndex = int(midIndex - sliceSize / 2)
                endIndex = int(midIndex - sliceSize / 2)
                testTrainIndex = [startIndex, endIndex]

                trainIdxs = datasetIndexes[:startIndex] + datasetIndexes[-endIndex:]
                testIdxs = datasetIndexes[startIndex:-endIndex]

                X_train_val_ = np.array(inputArr)[trainIdxs]
                y_train_val_ = np.array(outputArr)[trainIdxs]
                X_test_ = np.array(inputArr)[testIdxs]
                y_test_ = np.array(outputArr)[testIdxs]

            elif self.dict_machineLearningParameters[
                self.dkey_mlpTestPercentageDistribution()] == MLPF_PERCENTAGE_DISTRIBUTION_FROM_END:
                datasetIndexes = np.array(list(range(datasetSize)))
                trainIdxs = datasetIndexes[:-int(testPercentage * datasetSize)]
                testIdxs = datasetIndexes[-int(testPercentage * datasetSize):]
                testTrainIndex = [datasetSize - (testPercentage * datasetSize)]

                X_train_val_ = np.array(inputArr)[trainIdxs]
                y_train_val_ = np.array(outputArr)[trainIdxs]
                X_test_ = np.array(inputArr)[testIdxs]
                y_test_ = np.array(outputArr)[testIdxs]

            return X_train_val_, y_train_val_, X_test_, y_test_, testTrainIndex

        X_full = {}  # Create an input dict to store the values of train_val + test lists
        y_full = {}  # Create an output dict to store the values of train_val + test lists
        X_train_val = {}  # Create an input dict to store the values of train_val lists
        y_train_val = {}  # Create an output dict to store the values of train_val lists
        X_test = {}  # Create an input dict to store the values of test lists
        y_test = {}  # Create an output dict to store the values of test lists
        # Shorten the name of methodIndex
        methodIndex = self.dict_machineLearningParameters[self.dkey_mlpMethodIndex()]
        dict_test_train_index = {}

        inputHeaderColumnsForML = []
        outputHeaderColumnsForML = []

        # Check if the user selected Sequential method
        if self.dict_machineLearningParameters[self.dkey_mlpMethod()] == MLPF_METHOD_SEQUENTIAL_REGRESSION:
            for _event_ in dictDataInput.keys():
                X_full[_event_] = []
                y_full[_event_] = []
                X_train_val[_event_] = []
                y_train_val[_event_] = []
                X_test[_event_] = []
                y_test[_event_] = []
                dict_test_train_index[_event_] = []

                tmp_event_arr_input = []
                tmp_event_arr_output = []
                for _index_ in range(0, dictDataInput[_event_].__len__() - (2 * methodIndex + 1)):
                    tmp_arr_input = []
                    tmp_arr_output = []
                    for _index2_ in range(_index_, _index_ + methodIndex):
                        tmp_arr_input.extend(dictDataInput[_event_][_index2_])
                        tmp_arr_output.extend(dictDataOutput[_event_][_index2_ + methodIndex])
                    tmp_event_arr_input.append(tmp_arr_input)
                    tmp_event_arr_output.append(tmp_arr_output)
                    X_full[_event_].append(tmp_arr_input)
                    y_full[_event_].append(tmp_arr_output)

                # print("Event: ", _event_)
                # print("InputShape: ", np.array(tmp_event_arr_input).shape)
                # print("OutputShape: ", np.array(tmp_event_arr_output).shape)
                # print()

                X_train_val[_event_], y_train_val[_event_], X_test[_event_], y_test[
                    _event_], dict_test_train_index[_event_] = getTrainValTest(
                    tmp_event_arr_input,
                    tmp_event_arr_output)
                # np.convolve(np.array(X_full[_event_]), np.ones(methodIndex), 'valid') / methodIndex
                X_full[_event_] = np.array(X_full[_event_]).T
                y_full[_event_] = np.array(y_full[_event_]).T

                for _index_ in range(X_full[_event_].shape[0]):
                    X_full[_event_][_index_] = np.convolve(X_full[_event_][_index_], np.ones(methodIndex) / methodIndex,
                                                           mode='same')
                for _index_ in range(y_full[_event_].shape[0]):
                    y_full[_event_][_index_] = np.convolve(y_full[_event_][_index_], np.ones(methodIndex) / methodIndex,
                                                           mode='same')

                X_full[_event_] = np.array(X_full[_event_]).T
                y_full[_event_] = np.array(y_full[_event_]).T

            for _index_ in range(0, methodIndex):
                for columnName in inputHeaders:
                    inputHeaderColumnsForML.append(columnName + '_SEQ_' + str(_index_))
            for _index_ in range(0, methodIndex):
                for columnName in outputHeaders:
                    outputHeaderColumnsForML.append(columnName + '_SEQ_' + str(_index_))

        # Else if the user selected Average method
        elif self.dict_machineLearningParameters[self.dkey_mlpMethod()] == MLPF_METHOD_AVERAGE_REGRESSION:
            for _event_ in dictDataInput.keys():
                X_full[_event_] = []
                y_full[_event_] = []
                X_train_val[_event_] = []
                y_train_val[_event_] = []
                X_test[_event_] = []
                y_test[_event_] = []
                dict_test_train_index[_event_] = []

                tmp_event_arr_input = []
                tmp_event_arr_output = []
                for _index_ in range(0, dictDataInput[_event_].__len__() - (methodIndex + 1)):
                    tmp_arr_input = np.zeros(np.array(dictDataInput[_event_][_index_]).shape)
                    tmp_arr_output = np.zeros(np.array(dictDataOutput[_event_][_index_]).shape)
                    for _index2_ in range(_index_, _index_ + methodIndex):
                        tmp_arr_input += np.array(np.array(dictDataInput[_event_][_index2_]))
                        tmp_arr_output += np.array(np.array(dictDataOutput[_event_][_index2_]))
                    tmp_arr_input /= methodIndex
                    tmp_arr_output /= methodIndex

                    tmp_event_arr_input.append(tmp_arr_input.tolist())
                    tmp_event_arr_output.append(tmp_arr_output.tolist())
                    X_full[_event_].append(tmp_arr_input.tolist())
                    y_full[_event_].append(tmp_arr_output.tolist())

                # print("Event: ", _event_)
                # print("InputShape: ", np.array(tmp_event_arr_input).shape)
                # print("OutputShape: ", np.array(tmp_event_arr_output).shape)
                # print()

                X_train_val[_event_], y_train_val[_event_], X_test[_event_], y_test[
                    _event_], dict_test_train_index[_event_] = getTrainValTest(
                    tmp_event_arr_input,
                    tmp_event_arr_output)
                X_full[_event_] = np.array(X_full[_event_]).T
                y_full[_event_] = np.array(y_full[_event_]).T

                for _index_ in range(X_full[_event_].shape[0]):
                    X_full[_event_][_index_] = np.convolve(X_full[_event_][_index_], np.ones(methodIndex) / methodIndex,
                                                           mode='same')
                for _index_ in range(y_full[_event_].shape[0]):
                    y_full[_event_][_index_] = np.convolve(y_full[_event_][_index_], np.ones(methodIndex) / methodIndex,
                                                           mode='same')

                X_full[_event_] = np.array(X_full[_event_]).T
                y_full[_event_] = np.array(y_full[_event_]).T

            inputHeaderColumnsForML = inputHeaders
            outputHeaderColumnsForML = outputHeaders

        # Check if the user selected Sequential Average method
        elif self.dict_machineLearningParameters[self.dkey_mlpMethod()] == MLPF_METHOD_SEQUENTIAL_AVERAGE_REGRESSION:
            for _event_ in dictDataInput.keys():
                X_full[_event_] = []
                y_full[_event_] = []
                X_train_val[_event_] = []
                y_train_val[_event_] = []
                X_test[_event_] = []
                y_test[_event_] = []
                dict_test_train_index[_event_] = []

                tmp_event_arr_input = []
                tmp_event_arr_output = []
                for _index_ in range(0, dictDataInput[_event_].__len__() - (2 * methodIndex + 1)):
                    tmp_arr_input = []
                    tmp_arr_output = []

                    tmp_arr_input_mean = np.zeros(np.array(dictDataInput[_event_][_index_]).shape)
                    for _index2_ in range(_index_, _index_ + methodIndex):
                        tmp_arr_input_mean += dictDataInput[_event_][_index2_]
                    tmp_arr_input_mean /= methodIndex

                    for _index2_ in range(_index_, _index_ + methodIndex):
                        # Non absolute values - Uncomment this line for setting the input as non absolute values
                        tmp_input_ext = np.array(dictDataInput[_event_][_index2_]) - tmp_arr_input_mean
                        # Absolute values  - Uncomment this line for setting the input as absolute values
                        # tmp_input_ext = np.absolute(np.array(dictDataInput[_event_][_index2_]) - tmp_arr_input_mean)
                        tmp_arr_input.extend(tmp_input_ext.tolist())
                        tmp_arr_output.extend(dictDataOutput[_event_][_index2_ + methodIndex])

                    tmp_event_arr_input.append(tmp_arr_input)
                    tmp_event_arr_output.append(tmp_arr_output)
                    X_full[_event_].append(tmp_arr_input)
                    y_full[_event_].append(tmp_arr_output)

                # print("Event: ", _event_)
                # print("InputShape: ", np.array(tmp_event_arr_input).shape)
                # print("OutputShape: ", np.array(tmp_event_arr_output).shape)
                # print()

                X_train_val[_event_], y_train_val[_event_], X_test[_event_], y_test[
                    _event_], dict_test_train_index[_event_] = getTrainValTest(
                    tmp_event_arr_input,
                    tmp_event_arr_output)
                X_full[_event_] = np.array(X_full[_event_]).T
                y_full[_event_] = np.array(y_full[_event_]).T

                for _index_ in range(X_full[_event_].shape[0]):
                    X_full[_event_][_index_] = np.convolve(X_full[_event_][_index_], np.ones(methodIndex) / methodIndex,
                                                           mode='same')
                for _index_ in range(y_full[_event_].shape[0]):
                    y_full[_event_][_index_] = np.convolve(y_full[_event_][_index_], np.ones(methodIndex) / methodIndex,
                                                           mode='same')

                X_full[_event_] = np.array(X_full[_event_]).T
                y_full[_event_] = np.array(y_full[_event_]).T

            for _index_ in range(0, methodIndex):
                for columnName in inputHeaders:
                    inputHeaderColumnsForML.append(columnName + '_SEQ_' + str(_index_))
            for _index_ in range(0, methodIndex):
                for columnName in outputHeaders:
                    outputHeaderColumnsForML.append(columnName + '_SEQ_' + str(_index_))

        else:
            pass

        return X_train_val, y_train_val, X_test, y_test, X_full, y_full, \
               inputHeaderColumnsForML, outputHeaderColumnsForML, dict_test_train_index

    @staticmethod
    def BE_getArrayFromDictList(inputDictList: dict):
        tmpOutput = []
        for key in inputDictList.keys():
            for value in inputDictList[key]:
                tmpOutput.append(value)
        return np.array(tmpOutput)

    @staticmethod
    def BE_saveRealPredictedOutputFigure(y_Real: pd.DataFrame, y_Pred: pd.DataFrame, exportPath: str, y_max=None,
                                         trainTestSplit=None, title='Figure: Real-Predicted Values',
                                         yLabel='', xLabel=''):
        y_Real.plot(style=['bs-'])
        y_Pred.plot(style=['go-'])
        plt.gcf().set_size_inches(24.8, 12.4)
        plt.gcf().subplots_adjust(bottom=0.25)
        plt.title(title, fontsize=25)
        plt.legend(fontsize=20, loc='upper center', bbox_to_anchor=(0.5, -0.05), fancybox=True, shadow=False, ncol=2)
        plt.yticks(fontsize=20)
        plt.ylabel(yLabel,
                   fontsize=22.5)
        plt.xticks(fontsize=20)
        plt.xlabel(xLabel,
                   fontsize=22.5)

        if y_max is not None:
            y_bottom, y_top = plt.ylim(0, y_max)
        else:
            y_bottom, y_top = plt.ylim(0)
        if trainTestSplit is not None:
            for _vxLine_ in trainTestSplit:
                plt.vlines(_vxLine_, y_bottom, y_top, colors='r', linestyles='dashed')
        plt.savefig(exportPath, bbox_inches='tight', dpi=100)
        plt.close()

    @staticmethod
    def BE_saveQQPlotFigure(y_Real: np.ndarray, y_Pred: np.ndarray, exportPath: str, y_max=None,
                            trainTestSplit=None, title='Figure: QQ-PLot',
                            yLabel='', xLabel=''):

        y_QQ = np.array([[x, y] for (x, y) in (y_Real, y_Pred)])
        sm.qqplot(y_QQ, line='45')
        plt.gcf().set_size_inches(24.8, 12.4)
        plt.gcf().subplots_adjust(bottom=0.25)
        plt.title(title, fontsize=25)
        plt.legend(fontsize=20, loc='upper center', bbox_to_anchor=(0.5, -0.05), fancybox=True, shadow=False, ncol=2)
        plt.yticks(fontsize=20)
        plt.ylabel(yLabel,
                   fontsize=22.5)
        plt.xticks(fontsize=20)
        plt.xlabel(xLabel,
                   fontsize=22.5)

        if y_max is not None:
            y_bottom, y_top = plt.ylim(0, y_max)
        else:
            y_bottom, y_top = plt.ylim(0)
        if trainTestSplit is not None:
            for _vxLine_ in trainTestSplit:
                plt.vlines(_vxLine_, y_bottom, y_top, colors='r', linestyles='dashed')
        plt.savefig(exportPath, bbox_inches='tight', dpi=100)
        plt.close()

    @staticmethod
    def BE_saveAbsoluteErrorsFigure(absErrors: pd.DataFrame, exportPath: str, y_max=None,
                                    trainTestSplit=None, title='Figure: Absolute Errors',
                                    yLabel='', xLabel=''):
        absErrors.plot()
        plt.gcf().set_size_inches(24.8, 12.4)
        plt.gcf().subplots_adjust(bottom=0.25)
        plt.title(title, fontsize=25)
        plt.legend(fontsize=20, loc='upper center', bbox_to_anchor=(0.5, -0.05), fancybox=True, shadow=False, ncol=2)
        plt.yticks(fontsize=20)
        plt.ylabel(yLabel,
                   fontsize=22.5)
        plt.xticks(fontsize=20)
        plt.xlabel(xLabel,
                   fontsize=22.5)

        if y_max is not None:
            y_bottom, y_top = plt.ylim(0, y_max)
        else:
            y_bottom, y_top = plt.ylim(0)
        if trainTestSplit is not None:
            for _vxLine_ in trainTestSplit:
                plt.vlines(_vxLine_, y_bottom, y_top, colors='r', linestyles='dashed')
        plt.savefig(exportPath, bbox_inches='tight', dpi=100)
        plt.close()

    def BE_calculateErrors(self, y_Real_norm: np.ndarray, y_Pred_norm: np.ndarray,
                           y_Real_denorm: np.ndarray, y_Pred_denorm: np.ndarray,
                           path_toSavePlot=None, trainTestSplit=None,
                           title='Figure: Absolute Errors'):
        tmpAppendRow = []

        DY_RealPred_abs_norm = np.abs(y_Real_norm - y_Pred_norm)
        DY_RealPred_abs_denorm = np.abs(y_Real_denorm - y_Pred_denorm)

        # max
        norm_err_max_TrainValTest = DY_RealPred_abs_norm.max()
        denorm_err_max_TrainValTest = DY_RealPred_abs_denorm.max()
        # min
        norm_err_min_TrainValTest = DY_RealPred_abs_norm.min()
        denorm_err_min_TrainValTest = DY_RealPred_abs_denorm.min()
        # mse
        norm_err_mse_TrainValTest = sklearn.metrics.mean_squared_error(y_Real_norm,
                                                                       y_Pred_norm)
        denorm_err_mse_TrainValTest = sklearn.metrics.mean_squared_error(y_Real_denorm,
                                                                         y_Pred_denorm)
        # rmse
        norm_err_rmse_TrainValTest = np.sqrt(norm_err_mse_TrainValTest)
        denorm_err_rmse_TrainValTest = np.sqrt(denorm_err_mse_TrainValTest)
        # mae
        norm_err_mae_TrainValTest = sklearn.metrics.mean_absolute_error(y_Real_norm,
                                                                        y_Pred_norm)
        denorm_err_mae_TrainValTest = sklearn.metrics.mean_absolute_error(y_Real_denorm,
                                                                          y_Pred_denorm)

        tmpAppendRow.append(norm_err_max_TrainValTest)
        tmpAppendRow.append(denorm_err_max_TrainValTest)
        tmpAppendRow.append(norm_err_min_TrainValTest)
        tmpAppendRow.append(denorm_err_min_TrainValTest)
        tmpAppendRow.append(norm_err_mse_TrainValTest)
        tmpAppendRow.append(denorm_err_mse_TrainValTest)
        tmpAppendRow.append(norm_err_rmse_TrainValTest)
        tmpAppendRow.append(denorm_err_rmse_TrainValTest)
        tmpAppendRow.append(norm_err_mae_TrainValTest)
        tmpAppendRow.append(denorm_err_mae_TrainValTest)

        if path_toSavePlot is not None:
            print(file_manip.getCurrentDatetimeForConsole() +
                  "::Plot Denormalized Absolute Error Figure")
            y_max = DY_RealPred_abs_denorm.max()
            yLabel = 'Absolute Error'
            xLabel = ''
            self.BE_saveAbsoluteErrorsFigure(absErrors=pd.DataFrame(DY_RealPred_abs_denorm, columns=['Absolute Errors']),
                                             exportPath=path_toSavePlot,
                                             y_max=y_max,
                                             trainTestSplit=trainTestSplit,
                                             title=title,
                                             yLabel=yLabel, xLabel=xLabel)

        return tmpAppendRow

    # *                                                         * #
    # *********************************************************** #

    def actionButtonExecute(self):
        # FUNCTION FLAGS
        _FF_KEY_DATA = 'Data'
        _FF_KEY_COLUMN_PRIMARY_EVENT_DATA = 'Column Primary Event Data'
        _FF_KEY_PRIMARY_EVENT_UNIQUE_VALUES = 'Primary Event Unique Values'
        _FF_KEY_INPUT_COLUMNS = 'Input Columns'
        _FF_KEY_OUTPUT_COLUMNS = 'Output Columns'
        _FF_KEY_INPUT_COLUMNS_FOR_ML = 'Input Columns Machine Learning'
        _FF_KEY_OUTPUT_COLUMNS_FOR_ML = 'Output Columns Machine Learning'
        _FF_KEY_OUT_COL_HEADER_REAL = 'Output Header Real'
        _FF_KEY_OUT_COL_HEADER_PRED = 'Output Header Predicted'
        _FF_KEY_INP_COL_DENORM_VAL = 'Denormalize Input Values'
        _FF_KEY_OUT_COL_DENORM_VAL = 'Denormalize Output Values'
        _FF_KEY_TRAIN_VAL_ARRAY = 'Training Validation Array'
        _FF_KEY_TEST_ARRAY = 'Test Array'
        _FF_KEY_FULL_ARRAY = 'Full Array'
        _FF_KEY_INPUT = 'Input Array'
        _FF_KEY_OUTPUT = 'Output Array'
        _FF_KEY_TRAIN_TEST_SPLIT_INDEX = 'Train-Test Split Index'
        dict_fileData = {}

        # If true run the main pipeline
        if self.dict_tableFilesPaths.keys().__len__() > 0:  # if there is at least a file (safety if)
            # 00 - Error Checking
            for fileName in self.dict_tableFilesPaths.keys():  # for each file in tableFilePaths
                if self.BE_errorExist(fileName):  # if errors exists
                    return  # exit the function

            # 01 - Run The Main Routine
            for fileName in self.dict_tableFilesPaths.keys():
                # Create a tmp variable primaryEvent for code simplicity
                primaryEvent = self.dict_tableFilesPaths[fileName][self.dkeyPrimaryEventColumn()]
                dict_fileData[fileName] = {}  # Create a dictionary for file Data
                # read the data and store them in dictionary
                dict_fileData[fileName][_FF_KEY_DATA] = self.BE_readFileData(fileName)

                # Set the input/output columns
                dict_fileData[fileName][_FF_KEY_INPUT_COLUMNS] = self.dict_tableFilesPaths[fileName][
                    self.dkeyInputList()]
                dict_fileData[fileName][_FF_KEY_OUTPUT_COLUMNS] = self.dict_tableFilesPaths[fileName][
                    self.dkeyOutputList()]

                # Set the normalize/denormalize values
                dict_fileData[fileName][_FF_KEY_INP_COL_DENORM_VAL] = {}
                for _value_ in dict_fileData[fileName][_FF_KEY_INPUT_COLUMNS]:
                    tmp_col_max_value = dict_fileData[fileName][_FF_KEY_DATA][_value_].max()
                    dict_fileData[fileName][_FF_KEY_INP_COL_DENORM_VAL][_value_] = tmp_col_max_value
                dict_fileData[fileName][_FF_KEY_OUT_COL_DENORM_VAL] = {}
                for _value_ in dict_fileData[fileName][_FF_KEY_OUTPUT_COLUMNS]:
                    tmp_col_max_value = dict_fileData[fileName][_FF_KEY_DATA][_value_].max()
                    dict_fileData[fileName][_FF_KEY_OUT_COL_DENORM_VAL][_value_] = tmp_col_max_value

                # Set a path for exporting the input-output data
                currentFileName = os.path.splitext(fileName)[0]
                currentDatetime = file_manip.getCurrentDatetimeForPath()  # Find Current Datetime
                exportPrimaryDir = self.dict_machineLearningParameters[self.dkey_mlpExportFolder()] + \
                                   '/' + currentFileName + '/' + currentDatetime
                exportDataFolder = os.path.normpath(exportPrimaryDir + '/Data')
                file_manip.checkAndCreateFolders(exportDataFolder)  # check if path exists and if not create it
                # Export the input values
                file_manip.exportDictionaryNonList(dictForExport=dict_fileData[fileName][_FF_KEY_INP_COL_DENORM_VAL],
                                                   exportPath=os.path.normpath(
                                                       exportDataFolder + '/InputColumnsDenormValues.csv'),
                                                   headerLine=['InputColumn, DenormValue'])
                # Export the output values
                file_manip.exportDictionaryNonList(dictForExport=dict_fileData[fileName][_FF_KEY_OUT_COL_DENORM_VAL],
                                                   exportPath=os.path.normpath(
                                                       exportDataFolder + '/OutputColumnsDenormValues.csv'),
                                                   headerLine=['OutputColumn, DenormValue'])

                # Set the PRIMARY_EVENT_DATA as a dict
                dict_fileData[fileName][_FF_KEY_COLUMN_PRIMARY_EVENT_DATA] = {_FF_KEY_INPUT: {}, _FF_KEY_OUTPUT: {}}
                # if primaryEvent value is not None (means the user picked a primary column)
                if primaryEvent is not None:
                    # add the unique event values in the dictionary with key PRIMARY_EVENT_UNIQUE_VALUES
                    dict_fileData[fileName][_FF_KEY_PRIMARY_EVENT_UNIQUE_VALUES] = []
                    [dict_fileData[fileName][_FF_KEY_PRIMARY_EVENT_UNIQUE_VALUES].append(value)
                     for value in dict_fileData[fileName][_FF_KEY_DATA][primaryEvent]
                     if value not in dict_fileData[fileName][_FF_KEY_PRIMARY_EVENT_UNIQUE_VALUES]]

                    # Create a dictionary to normalize the data
                    tmp_dict_normValues = {}
                    for key in dict_fileData[fileName][_FF_KEY_INP_COL_DENORM_VAL].keys():
                        tmp_dict_normValues[key] = dict_fileData[fileName][_FF_KEY_INP_COL_DENORM_VAL][key]
                    for key in dict_fileData[fileName][_FF_KEY_OUT_COL_DENORM_VAL].keys():
                        tmp_dict_normValues[key] = dict_fileData[fileName][_FF_KEY_OUT_COL_DENORM_VAL][key]

                    # Create the event keys and store the corresponding rows for each key
                    for _event_ in dict_fileData[fileName][_FF_KEY_PRIMARY_EVENT_UNIQUE_VALUES]:
                        tmp_df = dict_fileData[fileName][_FF_KEY_DATA].copy()  # create a tmp copy of the data
                        tmp_df = tmp_df.loc[tmp_df[primaryEvent] == _event_]  # select only the rows of the events
                        del tmp_df[primaryEvent]

                        for key in tmp_df.keys():
                            tmp_df[key] /= tmp_dict_normValues[key]

                        tmp_input_columns = dict_fileData[fileName][_FF_KEY_INPUT_COLUMNS]
                        tmp_output_columns = dict_fileData[fileName][_FF_KEY_OUTPUT_COLUMNS]

                        dict_fileData[fileName][_FF_KEY_COLUMN_PRIMARY_EVENT_DATA][_FF_KEY_INPUT][_event_] = tmp_df[
                            tmp_input_columns].values.tolist()
                        dict_fileData[fileName][_FF_KEY_COLUMN_PRIMARY_EVENT_DATA][_FF_KEY_OUTPUT][_event_] = tmp_df[
                            tmp_output_columns].values.tolist()

                else:
                    # Create a dictionary to normalize the date
                    tmp_dict_normValues = {}
                    for key in dict_fileData[fileName][_FF_KEY_INP_COL_DENORM_VAL].keys():
                        tmp_dict_normValues[key] = dict_fileData[fileName][_FF_KEY_INP_COL_DENORM_VAL][key]
                    for key in dict_fileData[fileName][_FF_KEY_OUT_COL_DENORM_VAL].keys():
                        tmp_dict_normValues[key] = dict_fileData[fileName][_FF_KEY_OUT_COL_DENORM_VAL][key]

                    tmp_df = dict_fileData[fileName][_FF_KEY_DATA].copy()

                    for key in tmp_df.keys():
                        tmp_df[key] /= tmp_dict_normValues[key]

                    tmp_input_columns = dict_fileData[fileName][_FF_KEY_INPUT_COLUMNS]
                    tmp_output_columns = dict_fileData[fileName][_FF_KEY_OUTPUT_COLUMNS]

                    dict_fileData[fileName][_FF_KEY_COLUMN_PRIMARY_EVENT_DATA][_FF_KEY_INPUT][_FF_KEY_DATA] = tmp_df[
                        tmp_input_columns].values.tolist()
                    dict_fileData[fileName][_FF_KEY_COLUMN_PRIMARY_EVENT_DATA][_FF_KEY_OUTPUT][_FF_KEY_DATA] = tmp_df[
                        tmp_output_columns].values.tolist()

                dict_fileData[fileName][_FF_KEY_TRAIN_VAL_ARRAY] = {}
                dict_fileData[fileName][_FF_KEY_TEST_ARRAY] = {}
                dict_fileData[fileName][_FF_KEY_FULL_ARRAY] = {}
                dict_fileData[fileName][_FF_KEY_INPUT_COLUMNS_FOR_ML] = []
                dict_fileData[fileName][_FF_KEY_OUTPUT_COLUMNS_FOR_ML] = []
                dict_fileData[fileName][_FF_KEY_TRAIN_TEST_SPLIT_INDEX] = {}

                (dict_fileData[fileName][_FF_KEY_TRAIN_VAL_ARRAY][_FF_KEY_INPUT],
                 dict_fileData[fileName][_FF_KEY_TRAIN_VAL_ARRAY][_FF_KEY_OUTPUT],
                 dict_fileData[fileName][_FF_KEY_TEST_ARRAY][_FF_KEY_INPUT],
                 dict_fileData[fileName][_FF_KEY_TEST_ARRAY][_FF_KEY_OUTPUT],
                 dict_fileData[fileName][_FF_KEY_FULL_ARRAY][_FF_KEY_INPUT],
                 dict_fileData[fileName][_FF_KEY_FULL_ARRAY][_FF_KEY_OUTPUT],
                 dict_fileData[fileName][_FF_KEY_INPUT_COLUMNS_FOR_ML],
                 dict_fileData[fileName][_FF_KEY_OUTPUT_COLUMNS_FOR_ML],
                 dict_fileData[fileName][_FF_KEY_TRAIN_TEST_SPLIT_INDEX]) = \
                    self.BE_setTrainValTestArrays(
                        dictDataInput=dict_fileData[fileName][_FF_KEY_COLUMN_PRIMARY_EVENT_DATA][_FF_KEY_INPUT],
                        dictDataOutput=dict_fileData[fileName][_FF_KEY_COLUMN_PRIMARY_EVENT_DATA][_FF_KEY_OUTPUT],
                        inputHeaders=dict_fileData[fileName][_FF_KEY_INPUT_COLUMNS],
                        outputHeaders=dict_fileData[fileName][_FF_KEY_OUTPUT_COLUMNS])

                tmp_input_header_arr = ['Event']
                tmp_input_header_arr.extend(dict_fileData[fileName][_FF_KEY_INPUT_COLUMNS_FOR_ML])
                tmp_output_header_arr = ['Event']
                tmp_output_header_arr.extend(dict_fileData[fileName][_FF_KEY_OUTPUT_COLUMNS_FOR_ML])

                print(file_manip.getCurrentDatetimeForConsole() + "::Export INPUT training array...")
                file_manip.exportDictionaryList(
                    dictForExport=dict_fileData[fileName][_FF_KEY_TRAIN_VAL_ARRAY][_FF_KEY_INPUT],
                    exportPath=os.path.normpath(
                        exportDataFolder + '/InputTrainingValidation.csv'),
                    headerLine=tmp_input_header_arr)

                print(file_manip.getCurrentDatetimeForConsole() + "::Export OUTPUT training array...")
                file_manip.exportDictionaryList(
                    dictForExport=dict_fileData[fileName][_FF_KEY_TRAIN_VAL_ARRAY][_FF_KEY_OUTPUT],
                    exportPath=os.path.normpath(
                        exportDataFolder + '/OutputTrainingValidation.csv'),
                    headerLine=tmp_output_header_arr)

                print(file_manip.getCurrentDatetimeForConsole() + "::Export INPUT testing array...")
                file_manip.exportDictionaryList(
                    dictForExport=dict_fileData[fileName][_FF_KEY_TEST_ARRAY][_FF_KEY_INPUT],
                    exportPath=os.path.normpath(
                        exportDataFolder + '/InputTest.csv'),
                    headerLine=tmp_input_header_arr)

                print(file_manip.getCurrentDatetimeForConsole() + "::Export OUTPUT testing array...")
                file_manip.exportDictionaryList(
                    dictForExport=dict_fileData[fileName][_FF_KEY_TEST_ARRAY][_FF_KEY_OUTPUT],
                    exportPath=os.path.normpath(
                        exportDataFolder + '/OutputTest.csv'),
                    headerLine=tmp_output_header_arr)

                # Set the output headers Real/Pred which will be used later in the exported results
                dict_fileData[fileName][_FF_KEY_OUT_COL_HEADER_REAL] = []
                [dict_fileData[fileName][_FF_KEY_OUT_COL_HEADER_REAL].append(col + "_Real")
                 for col in dict_fileData[fileName][_FF_KEY_OUTPUT_COLUMNS_FOR_ML]]
                dict_fileData[fileName][_FF_KEY_OUT_COL_HEADER_PRED] = []
                [dict_fileData[fileName][_FF_KEY_OUT_COL_HEADER_PRED].append(col + "_Pred")
                 for col in dict_fileData[fileName][_FF_KEY_OUTPUT_COLUMNS_FOR_ML]]

                # print(dict_fileData[fileName])

                headerCorrelation = ['Method']
                for column_name in dict_fileData[fileName][_FF_KEY_OUTPUT_COLUMNS_FOR_ML]:
                    headerCorrelation.append(column_name)
                dirExportPlot = 'ExportPlots'

                # 02 - Run Machine Learning Process
                print(file_manip.getCurrentDatetimeForConsole() + "::Start Machine Learning Process")
                # ******** THIS CODE MAY BE EDITED WITH THREADING ****
                X_TrainVal = self.BE_getArrayFromDictList(
                    dict_fileData[fileName][_FF_KEY_TRAIN_VAL_ARRAY][_FF_KEY_INPUT])
                y_TrainVal = self.BE_getArrayFromDictList(
                    dict_fileData[fileName][_FF_KEY_TRAIN_VAL_ARRAY][_FF_KEY_OUTPUT])
                X_Test = self.BE_getArrayFromDictList(
                    dict_fileData[fileName][_FF_KEY_TEST_ARRAY][_FF_KEY_INPUT])
                y_Test = self.BE_getArrayFromDictList(
                    dict_fileData[fileName][_FF_KEY_TEST_ARRAY][_FF_KEY_OUTPUT])
                listStr_ModelPaths, exportBaseDirPath, workbookDirPath = \
                    self.mlr_Regression.fit(X_TrainVal=X_TrainVal,
                                            y_TrainVal=y_TrainVal,
                                            X_Test=X_Test,
                                            y_Test=y_Test,
                                            exportFolder=exportPrimaryDir)
                # Create a workbook for storing the Errors for each unique event
                workbookPath_ErrorsForTrainValTest = workbookDirPath + '/' + currentFileName + '_ErrorsForTrainValTest.xlsx'
                workbookPath_ErrorsForTrainVal = workbookDirPath + '/' + currentFileName + '_ErrorsForTrainVal.xlsx'
                workbookPath_ErrorsForTest = workbookDirPath + '/' + currentFileName + '_ErrorsForTest.xlsx'

                # for each uniqueEvent
                uniqueEventCounter = 0
                uniqueEventSize = dict_fileData[fileName][_FF_KEY_PRIMARY_EVENT_UNIQUE_VALUES].__len__()
                for _uniqueEvent_ in dict_fileData[fileName][_FF_KEY_PRIMARY_EVENT_UNIQUE_VALUES]:
                    print('\n' + file_manip.getCurrentDatetimeForConsole() + '::(' + str(uniqueEventCounter + 1) +
                          ' / ' + str(uniqueEventSize) + ') Event: ', _uniqueEvent_)
                    uniqueEventCounter += 1
                    # Shorten the variable name of trainTestSplitIndex
                    trainTestSplitIndex = dict_fileData[fileName][_FF_KEY_TRAIN_TEST_SPLIT_INDEX][_uniqueEvent_]

                    # ---------------------------------------------------------------------------- #

                    # ***** workbookPath_ErrorsForTrainValTest ***** #
                    # try to load the workbook for storing the errors
                    try:
                        wb_ErrorsForTrainValTest = op.load_workbook(workbookPath_ErrorsForTrainValTest)  # load workbook
                        ws_ErrorsForTrainValTest = wb_ErrorsForTrainValTest.worksheets[0]  # select first worksheet
                    except FileNotFoundError:  # exception: Create the workbook
                        # Create the header row
                        headers_row = ['Event', 'Technique']
                        headers_errors = ['_MAX_NORM', '_MAX_DENORM',
                                          '_MIN_NORM', '_MIN_DENORM',
                                          '_MSE_NORM', '_MSE_DENORM',
                                          '_RMSE_NORM', '_RMSE_DENORM',
                                          '_MAE_NORM', '_MAE_DENORM']
                        for _header_ in dict_fileData[fileName][_FF_KEY_OUTPUT_COLUMNS_FOR_ML]:
                            for _headerError_ in headers_errors:
                                headers_row.append(_header_ + _headerError_)

                        wb_ErrorsForTrainValTest = op.Workbook()  # open a workbook
                        ws_ErrorsForTrainValTest = wb_ErrorsForTrainValTest.active  # set a worksheet active
                        ws_ErrorsForTrainValTest.append(headers_row)  # append to worksheet the header row
                        wb_ErrorsForTrainValTest.save(workbookPath_ErrorsForTrainValTest)  # save the workbook

                    # ***** workbookPath_ErrorsForTrainVal ***** #
                    try:
                        wb_ErrorsForTrainVal = op.load_workbook(workbookPath_ErrorsForTrainVal)  # load workbook
                        ws_ErrorsForTrainVal = wb_ErrorsForTrainVal.worksheets[0]  # select first worksheet
                    except FileNotFoundError:  # exception: Create the workbook
                        # Create the header row
                        headers_row = ['Event', 'Technique']
                        headers_errors = ['_MAX_NORM', '_MAX_DENORM',
                                          '_MIN_NORM', '_MIN_DENORM',
                                          '_MSE_NORM', '_MSE_DENORM',
                                          '_RMSE_NORM', '_RMSE_DENORM',
                                          '_MAE_NORM', '_MAE_DENORM']
                        for _header_ in dict_fileData[fileName][_FF_KEY_OUTPUT_COLUMNS_FOR_ML]:
                            for _headerError_ in headers_errors:
                                headers_row.append(_header_ + _headerError_)

                        wb_ErrorsForTrainVal = op.Workbook()  # open a workbook
                        ws_ErrorsForTrainVal = wb_ErrorsForTrainVal.active  # set a worksheet active
                        ws_ErrorsForTrainVal.append(headers_row)  # append to worksheet the header row
                        wb_ErrorsForTrainVal.save(workbookPath_ErrorsForTrainVal)  # save the workbook

                    # ***** workbookPath_ErrorsForTest ***** #
                    try:
                        wb_ErrorsForTest = op.load_workbook(workbookPath_ErrorsForTest)  # load workbook
                        ws_ErrorsForTest = wb_ErrorsForTest.worksheets[0]  # select first worksheet
                    except FileNotFoundError:  # exception: Create the workbook
                        # Create the header row
                        headers_row = ['Event', 'Technique']
                        headers_errors = ['_MAX_NORM', '_MAX_DENORM',
                                          '_MIN_NORM', '_MIN_DENORM',
                                          '_MSE_NORM', '_MSE_DENORM',
                                          '_RMSE_NORM', '_RMSE_DENORM',
                                          '_MAE_NORM', '_MAE_DENORM']
                        for _header_ in dict_fileData[fileName][_FF_KEY_OUTPUT_COLUMNS_FOR_ML]:
                            for _headerError_ in headers_errors:
                                headers_row.append(_header_ + _headerError_)

                        wb_ErrorsForTest = op.Workbook()  # open a workbook
                        ws_ErrorsForTest = wb_ErrorsForTest.active  # set a worksheet active
                        ws_ErrorsForTest.append(headers_row)  # append to worksheet the header row
                        wb_ErrorsForTest.save(workbookPath_ErrorsForTest)  # save the workbook

                    # ---------------------------------------------------------------------------- #

                    # Store the INPUT and OUTPUT of current event to shortened named variables
                    df_x_TrainValTest = dict_fileData[fileName][_FF_KEY_FULL_ARRAY][_FF_KEY_INPUT][_uniqueEvent_]
                    df_y_TrainValTest = dict_fileData[fileName][_FF_KEY_FULL_ARRAY][_FF_KEY_OUTPUT][_uniqueEvent_]

                    df_x_TrainVal = dict_fileData[fileName][_FF_KEY_TRAIN_VAL_ARRAY][_FF_KEY_INPUT][_uniqueEvent_]
                    df_y_TrainVal = dict_fileData[fileName][_FF_KEY_TRAIN_VAL_ARRAY][_FF_KEY_OUTPUT][_uniqueEvent_]

                    df_x_Test = dict_fileData[fileName][_FF_KEY_TEST_ARRAY][_FF_KEY_INPUT][_uniqueEvent_]
                    df_y_Test = dict_fileData[fileName][_FF_KEY_TEST_ARRAY][_FF_KEY_OUTPUT][_uniqueEvent_]

                    # make predictions for all models on TrainValTest
                    dict_Y_TrainValTest = self.mlr_Regression.predict(df_x_TrainValTest,
                                                                      df_y_TrainValTest)
                    # make predictions for all models on TrainVal
                    dict_Y_TrainVal = self.mlr_Regression.predict(df_x_TrainVal,
                                                                  df_y_TrainVal)

                    # make predictions for all models on Test
                    dict_Y_Test = self.mlr_Regression.predict(df_x_Test,
                                                              df_y_Test)

                    # cor_CSV = [headerCorrelation]  # create a CSV list and add the header row
                    for _modelName_ in dict_Y_TrainValTest:  # for each modelName in dict_Y
                        # Create an append row (for workbook) and add the uniqueEvent and modelName
                        tmpAppendRow_TrainValTest = [_uniqueEvent_, _modelName_]
                        tmpAppendRow_TrainVal = [_uniqueEvent_, _modelName_]
                        tmpAppendRow_Test = [_uniqueEvent_, _modelName_]
                        # Create an append row (for cor_CSV) and add the modelName
                        # tmpCorRow_CSV = [_modelName_]

                        # Store the normalized values (real, predicted) to shortened variables
                        df_Y_realNorm_TrainValTest = pd.DataFrame(
                            dict_Y_TrainValTest[_modelName_]['real'],
                            columns=dict_fileData[fileName][
                                _FF_KEY_OUT_COL_HEADER_REAL])
                        df_Y_predNorm_TrainValTest = pd.DataFrame(
                            dict_Y_TrainValTest[_modelName_]['pred'],
                            columns=dict_fileData[fileName][
                                _FF_KEY_OUT_COL_HEADER_PRED])

                        df_Y_realNorm_TrainVal = pd.DataFrame(
                            dict_Y_TrainVal[_modelName_]['real'],
                            columns=dict_fileData[fileName][
                                _FF_KEY_OUT_COL_HEADER_REAL])
                        df_Y_predNorm_TrainVal = pd.DataFrame(
                            dict_Y_TrainVal[_modelName_]['pred'],
                            columns=dict_fileData[fileName][
                                _FF_KEY_OUT_COL_HEADER_PRED])

                        df_Y_realNorm_Test = pd.DataFrame(
                            dict_Y_Test[_modelName_]['real'],
                            columns=dict_fileData[fileName][
                                _FF_KEY_OUT_COL_HEADER_REAL])
                        df_Y_predNorm_Test = pd.DataFrame(
                            dict_Y_Test[_modelName_]['pred'],
                            columns=dict_fileData[fileName][
                                _FF_KEY_OUT_COL_HEADER_PRED])

                        # Copy the normalized values to denormalized variables
                        df_Y_realDenorm_TrainValTest = df_Y_realNorm_TrainValTest.copy()
                        df_Y_predDenorm_TrainValTest = df_Y_predNorm_TrainValTest.copy()

                        df_Y_realDenorm_TrainVal = df_Y_realNorm_TrainVal.copy()
                        df_Y_predDenorm_TrainVal = df_Y_predNorm_TrainVal.copy()

                        df_Y_realDenorm_Test = df_Y_realNorm_Test.copy()
                        df_Y_predDenorm_Test = df_Y_predNorm_Test.copy()

                        # Create the paths for exporting data
                        # Main Directory
                        o_dir_Model = os.path.normpath(exportBaseDirPath + '/' + dirExportPlot + '/' +
                                                       _uniqueEvent_ + '/' + _modelName_) + '/'
                        file_manip.checkAndCreateFolders(o_dir_Model)
                        # Directory to Export the Signal Comparison
                        o_dir_SignalCompare = os.path.normpath(o_dir_Model + 'SignalCompare') + '/'
                        file_manip.checkAndCreateFolders(o_dir_SignalCompare)
                        # Directory to Export the Real and Predicted Plots
                        o_dir_RealPredictPlots = os.path.normpath(o_dir_Model + 'RealPredictPlots') + '/'
                        file_manip.checkAndCreateFolders(o_dir_RealPredictPlots)
                        # Directory to Export the Real and Predicted CSV normalized/denormalized values
                        o_dir_RealPredictCSV = os.path.normpath(o_dir_Model + 'RealPredictCSV') + '/'
                        file_manip.checkAndCreateFolders(o_dir_RealPredictCSV)

                        for _index_ in range(0, dict_fileData[fileName][_FF_KEY_OUT_COL_HEADER_REAL].__len__()):
                            mul_ind = 1.0  # set multiply index to 1
                            # get current row
                            currColumn_real = dict_fileData[fileName][_FF_KEY_OUT_COL_HEADER_REAL][_index_]
                            currColumn_pred = dict_fileData[fileName][_FF_KEY_OUT_COL_HEADER_PRED][_index_]
                            currColumn = dict_fileData[fileName][_FF_KEY_OUTPUT_COLUMNS_FOR_ML][_index_]

                            for _oColumn_ in dict_fileData[fileName][_FF_KEY_OUT_COL_DENORM_VAL]:
                                if currColumn_real.__contains__(_oColumn_):
                                    mul_ind = dict_fileData[fileName][_FF_KEY_OUT_COL_DENORM_VAL][_oColumn_]

                            df_Y_realDenorm_TrainValTest[currColumn_real] *= mul_ind
                            df_Y_predDenorm_TrainValTest[currColumn_pred] *= mul_ind

                            df_Y_realDenorm_TrainVal[currColumn_real] *= mul_ind
                            df_Y_predDenorm_TrainVal[currColumn_pred] *= mul_ind

                            df_Y_realDenorm_Test[currColumn_real] *= mul_ind
                            df_Y_predDenorm_Test[currColumn_pred] *= mul_ind

                            print(file_manip.getCurrentDatetimeForConsole() +
                                  "::Plot normalized figure for column: " + currColumn)
                            y_max = df_Y_realNorm_TrainValTest[currColumn_real].max()
                            if df_Y_predNorm_TrainValTest[currColumn_pred].max() > y_max:
                                y_max = df_Y_predNorm_TrainValTest[currColumn_pred].max()
                            exportFigPath = o_dir_RealPredictPlots + 'Normalized_' + _uniqueEvent_ + '_' + currColumn + '.png'
                            figTitle = _uniqueEvent_ + ': ' + currColumn
                            self.BE_saveRealPredictedOutputFigure(y_Real=df_Y_realNorm_TrainValTest[currColumn_real],
                                                                  y_Pred=df_Y_predNorm_TrainValTest[currColumn_pred],
                                                                  exportPath=exportFigPath,
                                                                  # y_max=y_max,
                                                                  trainTestSplit=trainTestSplitIndex,
                                                                  title=figTitle,
                                                                  yLabel=currColumn + ' (x' + str(mul_ind) + ')',
                                                                  xLabel='')

                            print(file_manip.getCurrentDatetimeForConsole() +
                                  "::Plot normalized figure for column: " + currColumn)
                            exportFigPath = o_dir_RealPredictPlots + 'Denormalized_' + _uniqueEvent_ + '_' + currColumn + '.png'
                            figTitle = _uniqueEvent_ + ': ' + currColumn
                            y_max *= mul_ind
                            self.BE_saveRealPredictedOutputFigure(y_Real=df_Y_realDenorm_TrainValTest[currColumn_real],
                                                                  y_Pred=df_Y_predDenorm_TrainValTest[currColumn_pred],
                                                                  exportPath=exportFigPath,
                                                                  # y_max=y_max,
                                                                  trainTestSplit=trainTestSplitIndex,
                                                                  title=figTitle,
                                                                  yLabel=currColumn,
                                                                  xLabel='')

                            print(file_manip.getCurrentDatetimeForConsole() + "::Calculate the Errors")
                            # Calculate the Errors - TrainValTest
                            normList_yReal = df_Y_realNorm_TrainValTest[currColumn_real].to_numpy()
                            normList_yPred = df_Y_predNorm_TrainValTest[currColumn_pred].to_numpy()
                            denormList_yReal = df_Y_realDenorm_TrainValTest[currColumn_real].to_numpy()
                            denormList_yPred = df_Y_predDenorm_TrainValTest[currColumn_pred].to_numpy()

                            exportFigPath = o_dir_RealPredictPlots + 'Denormalized_TrainValTest_AbsoluteErrors' + \
                                            _uniqueEvent_ + '_' + currColumn + '.png'
                            figTitle = _uniqueEvent_ + ': ' + currColumn + ' Absolute Errors Full Set'
                            tmpErrorArr = self.BE_calculateErrors(
                                normList_yReal, normList_yPred,
                                denormList_yReal, denormList_yPred,
                                path_toSavePlot=exportFigPath,
                                trainTestSplit=trainTestSplitIndex,
                                title=figTitle
                            )

                            exportFigPath = o_dir_RealPredictPlots + 'Denormalized_TrainValTest_QQPlot' + \
                                            _uniqueEvent_ + '_' + currColumn + '.png'
                            figTitle = _uniqueEvent_ + ': ' + currColumn + ' QQ-Plot'
                            self.BE_saveQQPlotFigure(
                                y_Real=denormList_yReal,
                                y_Pred=denormList_yPred,
                                exportPath=exportFigPath,
                                title=figTitle,
                                yLabel='Predicted Values', xLabel='Real Values')

                            for _error_ in tmpErrorArr:
                                tmpAppendRow_TrainValTest.append(_error_)

                            # Calculate the Errors - TrainVal
                            normList_yReal = df_Y_realNorm_TrainVal[currColumn_real].to_numpy()
                            normList_yPred = df_Y_predNorm_TrainVal[currColumn_pred].to_numpy()
                            denormList_yReal = df_Y_realDenorm_TrainVal[currColumn_real].to_numpy()
                            denormList_yPred = df_Y_predDenorm_TrainVal[currColumn_pred].to_numpy()

                            exportFigPath = o_dir_RealPredictPlots + 'Denormalized_TrainVal_AbsoluteErrors' + \
                                            _uniqueEvent_ + '_' + currColumn + '.png'
                            figTitle = _uniqueEvent_ + ': ' + currColumn + ' Absolute Errors Train-Validation Set'
                            tmpErrorArr = self.BE_calculateErrors(
                                normList_yReal, normList_yPred,
                                denormList_yReal, denormList_yPred,
                                path_toSavePlot=exportFigPath,
                                trainTestSplit=None,
                                title=figTitle
                            )

                            for _error_ in tmpErrorArr:
                                tmpAppendRow_TrainVal.append(_error_)

                            # Calculate the Errors - Test
                            normList_yReal = df_Y_realNorm_Test[currColumn_real].to_numpy()
                            normList_yPred = df_Y_predNorm_Test[currColumn_pred].to_numpy()
                            denormList_yReal = df_Y_realDenorm_Test[currColumn_real].to_numpy()
                            denormList_yPred = df_Y_predDenorm_Test[currColumn_pred].to_numpy()

                            exportFigPath = o_dir_RealPredictPlots + 'Denormalized_Test_AbsoluteErrors' + \
                                            _uniqueEvent_ + '_' + currColumn + '.png'
                            figTitle = _uniqueEvent_ + ': ' + currColumn + ' Absolute Errors Test Set'
                            tmpErrorArr = self.BE_calculateErrors(
                                normList_yReal, normList_yPred,
                                denormList_yReal, denormList_yPred,
                                path_toSavePlot=exportFigPath,
                                trainTestSplit=None,
                                title=figTitle
                            )

                            for _error_ in tmpErrorArr:
                                tmpAppendRow_Test.append(_error_)

                            corrFileName = _uniqueEvent_ + '_' + currColumn
                            self.signComp_Methods.signComp_exec_(
                                arrData1=df_Y_realNorm_TrainValTest[currColumn_real].to_numpy(),
                                arrData2=df_Y_predNorm_TrainValTest[currColumn_pred].to_numpy(),
                                exportFigDirPath=o_dir_SignalCompare,
                                exportFigFileName=corrFileName,
                            )

                        print(file_manip.getCurrentDatetimeForConsole() + "::Export Files")
                        ws_ErrorsForTrainValTest.append(tmpAppendRow_TrainValTest)
                        wb_ErrorsForTrainValTest.save(workbookPath_ErrorsForTrainValTest)
                        ws_ErrorsForTrainVal.append(tmpAppendRow_TrainVal)
                        wb_ErrorsForTrainVal.save(workbookPath_ErrorsForTrainVal)
                        ws_ErrorsForTest.append(tmpAppendRow_Test)
                        wb_ErrorsForTest.save(workbookPath_ErrorsForTest)
                        df_Y_realNorm_TrainValTest.to_csv(
                            o_dir_RealPredictCSV + '/' + _uniqueEvent_ + '_' + _modelName_ + '_OutputReal_Normalized.csv')
                        df_Y_predNorm_TrainValTest.to_csv(
                            o_dir_RealPredictCSV + '/' + _uniqueEvent_ + '_' + _modelName_ + '_OutputPred_Normalized.csv')
                        df_Y_realDenorm_TrainValTest.to_csv(
                            o_dir_RealPredictCSV + '/' + _uniqueEvent_ + '_' + _modelName_ + '_OutputReal_Denormalized.csv')
                        df_Y_predDenorm_TrainValTest.to_csv(
                            o_dir_RealPredictCSV + '/' + _uniqueEvent_ + '_' + _modelName_ + '_OutputPred_Denormalized.csv')

                    # o_file = os.path.normpath(dir_path + "/../") + '/Correlation_R2.csv'
                    # my_cal_v2.write_csv(o_file, cor_CSV)
        print(file_manip.getCurrentDatetimeForConsole() + "::Execution Finished Successfully!!!")

    def actionFileListRowChanged_event(self):
        self.listWidget_ColumnList.clear()  # Clear Column Widget
        if self.listWidget_FileList.currentItem() is not None:  # If current item is not None
            self.fileName = self.listWidget_FileList.currentItem().text()  # get current item name
            for column in self.dict_tableFilesPaths[self.fileName][self.dkeyColumnsList()]:  # for each column
                # Add columns as ITEMS to widget
                self.listWidget_ColumnList.addItem(QListWidgetItem(column))

            if self.listWidget_ColumnList.currentItem() is None:  # If COLUMN widget is not None
                self.listWidget_ColumnList.setCurrentRow(0)  # Set first row selected
        else:
            self.fileName = None

    # *********************************************************** #
    # ******************** Helping Functions ******************** #
    # *********************************************************** #
    # *                                                         * #

    def updateInputList(self):
        self.widgetTabInputOutput.listWidget_InputColumns.clear()  # Clear Event Widget
        for key in self.dict_tableFilesPaths.keys():  # For each key (fileName)
            if self.dict_tableFilesPaths[key][self.dkeyInputList()] is not []:  # if event key is not []
                for event in self.dict_tableFilesPaths[key][self.dkeyInputList()]:  # for each EVENT
                    # Add ITEM to INPUT widget
                    self.widgetTabInputOutput.listWidget_InputColumns.addItem(
                        QListWidgetItem(key + " -> " + event))

    def removeFromInputColumn(self, fileName, column):
        # Remove the specified COLUMN from INPUT_COLUMN_LIST for the specified FILE
        if column in self.dict_tableFilesPaths[fileName][self.dkeyInputList()]:
            self.dict_tableFilesPaths[fileName][self.dkeyInputList()].remove(column)

    def updateOutputList(self):
        self.widgetTabInputOutput.listWidget_OutputColumns.clear()  # Clear Event Widget
        for key in self.dict_tableFilesPaths.keys():  # For each key (fileName)
            if self.dict_tableFilesPaths[key][self.dkeyOutputList()] is not []:  # if event key is not []
                for event in self.dict_tableFilesPaths[key][self.dkeyOutputList()]:  # for each EVENT
                    # Add ITEM to OUTPUT widget
                    self.widgetTabInputOutput.listWidget_OutputColumns.addItem(
                        QListWidgetItem(key + " -> " + event))

    def removeFromOutputColumn(self, fileName, column):
        # Remove the specified COLUMN from OUTPUT_COLUMN_LIST for the specified FILE
        if column in self.dict_tableFilesPaths[fileName][self.dkeyOutputList()]:
            self.dict_tableFilesPaths[fileName][self.dkeyOutputList()].remove(column)

    def updatePrimaryEventList(self):
        self.widgetTabInputOutput.listWidget_PrimaryEvent.clear()  # Clear Primary Event Widget
        for key in self.dict_tableFilesPaths.keys():  # For each key (fileName)
            # if primary event key is not None
            if self.dict_tableFilesPaths[key][self.dkeyPrimaryEventColumn()] is not None:
                # Add ITEM to PRIMARY_EVENT widget
                self.widgetTabInputOutput.listWidget_PrimaryEvent.addItem(
                    QListWidgetItem(key + " -> " + self.dict_tableFilesPaths[key][self.dkeyPrimaryEventColumn()]))

    def resetPrimEventColumn(self, fileName):
        # Set PRIMARY_COLUMN for the specified FILE to None
        self.dict_tableFilesPaths[fileName][self.dkeyPrimaryEventColumn()] = None

    # *                                                         * #
    # *********************************************************** #

    def actionButtonInput(self):
        # If some file is selected and some columns are selected
        if self.listWidget_FileList.currentItem() is not None and \
                self.listWidget_ColumnList.currentItem() is not None:
            # get current columns selected
            currentSelectedItems = self.listWidget_ColumnList.selectedItems()
            for currentColumnSelected in currentSelectedItems:  # for each item selected
                # if this column is not in the INPUT List
                if currentColumnSelected.text() not in self.dict_tableFilesPaths[self.fileName][self.dkeyInputList()]:
                    # Add it to list
                    self.dict_tableFilesPaths[self.fileName][self.dkeyInputList()].append(currentColumnSelected.text())
                    # If this column exist in the private event list
                    if currentColumnSelected.text() == self.dict_tableFilesPaths[self.fileName][
                        self.dkeyPrimaryEventColumn()]:
                        self.resetPrimEventColumn(self.fileName)  # remove it from the list
                        self.updatePrimaryEventList()  # update Input widget
                # print(currentFileName, " -> ", currentColumnSelected.text())
            self.updateInputList()  # update Event widget

    def actionButtonRemInput(self):
        # If some file is selected and some columns are selected
        if self.widgetTabInputOutput.listWidget_InputColumns.currentItem() is not None:
            # get selected items
            selectedItems = self.widgetTabInputOutput.listWidget_InputColumns.selectedItems()
            for item in selectedItems:  # for each item
                tmp_str = item.text()  # get text
                fileName = tmp_str.split(' -> ')[0]  # get fileName
                columnName = tmp_str.split(' -> ')[1]  # get columnName
                self.removeFromInputColumn(fileName, columnName)  # remove event from the list
            self.updateInputList()  # update EVENT widget

    def actionButtonOutput(self):
        # If some file is selected and some columns are selected
        if self.listWidget_FileList.currentItem() is not None and \
                self.listWidget_ColumnList.currentItem() is not None:
            # get current columns selected
            currentSelectedItems = self.listWidget_ColumnList.selectedItems()
            for currentColumnSelected in currentSelectedItems:  # for each item selected
                # if this column is not in the INPUT List
                if currentColumnSelected.text() not in self.dict_tableFilesPaths[self.fileName][self.dkeyOutputList()]:
                    # Add it to list
                    self.dict_tableFilesPaths[self.fileName][self.dkeyOutputList()].append(currentColumnSelected.text())
                    # If this column exist in the private event list
                    if currentColumnSelected.text() == self.dict_tableFilesPaths[self.fileName][
                        self.dkeyPrimaryEventColumn()]:
                        self.resetPrimEventColumn(self.fileName)  # remove it from the list
                        self.updatePrimaryEventList()  # update Input widget
                # print(currentFileName, " -> ", currentColumnSelected.text())
            self.updateOutputList()  # update Event widget

    def actionButtonRemOutput(self):
        # If some file is selected and some columns are selected
        if self.widgetTabInputOutput.listWidget_OutputColumns.currentItem() is not None:
            # get selected items
            selectedItems = self.widgetTabInputOutput.listWidget_OutputColumns.selectedItems()
            for item in selectedItems:  # for each item
                tmp_str = item.text()  # get text
                fileName = tmp_str.split(' -> ')[0]  # get fileName
                columnName = tmp_str.split(' -> ')[1]  # get columnName
                self.removeFromOutputColumn(fileName, columnName)  # remove event from the list
            self.updateOutputList()  # update EVENT widget

    def actionButtonPrimaryEvent(self):
        # If some file is selected and some column is selected
        if self.listWidget_FileList.currentItem() is not None and \
                self.listWidget_ColumnList.currentItem() is not None:
            # get current column name
            currentColumnSelected = self.listWidget_ColumnList.currentItem().text()
            # If this column exist in the input list
            if currentColumnSelected in self.dict_tableFilesPaths[self.fileName][self.dkeyInputList()]:
                self.removeFromInputColumn(self.fileName, currentColumnSelected)  # remove it from the list
                self.updateInputList()  # update Input widget
            # If this column exist in the output list
            if currentColumnSelected in self.dict_tableFilesPaths[self.fileName][self.dkeyOutputList()]:
                self.removeFromOutputColumn(self.fileName, currentColumnSelected)  # remove it from the list
                self.updateOutputList()  # update Input widget

            # print(currentFileName, " -> ", currentColumnSelected)

            # Add it to the PRIMARY_EVENT
            self.dict_tableFilesPaths[self.fileName][self.dkeyPrimaryEventColumn()] = currentColumnSelected
            self.updatePrimaryEventList()  # update Primary Event widget

    def actionButtonRemPrimaryEvent(self):
        # If some file is selected and some columns are selected
        if self.widgetTabInputOutput.listWidget_PrimaryEvent.isActiveWindow() and \
                self.widgetTabInputOutput.listWidget_PrimaryEvent.currentItem() is not None:
            # get selected item
            selectedItems = self.widgetTabInputOutput.listWidget_PrimaryEvent.selectedItems()
            for item in selectedItems:  # for each item
                tmp_str = item.text()  # get text
                fileName = tmp_str.split(' -> ')[0]  # get fileName
                self.resetPrimEventColumn(fileName)  # remove PRIMARY_EVENT from the list
            self.updatePrimaryEventList()  # update PRIMARY_EVENT widget

    # ***** SET SETTINGS GENERAL EVENTS ACTIONS ***** #
    def actionExperimentResultChange(self):
        self.dict_machineLearningParameters[self.dkey_mlpExperimentNumber()] = \
            self.widgetTabMachineLearningSettings.tabGeneral.spinBox_ExperimentNumber.value()
        if self.debugMessageFlag:
            print(self.dict_machineLearningParameters[self.dkey_mlpExperimentNumber()])

    def actionTestPercentageChange(self):
        self.dict_machineLearningParameters[self.dkey_mlpTestPercentage()] = \
            self.widgetTabMachineLearningSettings.tabGeneral.doubleSpinBox_TestPercentage.value()
        if self.debugMessageFlag:
            print(self.dict_machineLearningParameters[self.dkey_mlpTestPercentage()])

    def actionHoldoutPercentageChange(self):
        self.dict_machineLearningParameters[self.dkey_mlpHoldoutPercentage()] = \
            self.widgetTabMachineLearningSettings.tabGeneral.doubleSpinBox_HoldoutPercentage.value()
        if self.debugMessageFlag:
            print(self.dict_machineLearningParameters[self.dkey_mlpHoldoutPercentage()])

    def actionLineEditChange(self):
        self.dict_machineLearningParameters[self.dkey_mlpExportFolder()] = \
            self.widgetTabMachineLearningSettings.tabGeneral.lineEdit_SetOutPath.text()
        if self.debugMessageFlag:
            print(self.dict_machineLearningParameters[self.dkey_mlpExportFolder()])

    def actionButtonSetOutPathClicked(self):
        success, dialog = coFunc.openDirectoryDialog(
            classRef=self,
            dialogName='Choose a Directory',
            dialogOpenAt=self.str_pathToTheProject,
            dialogMultipleSelection=False)

        if success:
            self.widgetTabMachineLearningSettings.tabGeneral.lineEdit_SetOutPath.setText(dialog)

    def actionMachineLearningMethodChange(self):
        self.dict_machineLearningParameters[self.dkey_mlpMethod()] = \
            self.widgetTabMachineLearningSettings.tabGeneral.comboBox_MachineLearningMethods.currentText()
        if self.debugMessageFlag:
            print(self.dict_machineLearningParameters[self.dkey_mlpMethod()])

    def actionMachineLearningMethodIndexChange(self):
        self.dict_machineLearningParameters[self.dkey_mlpMethodIndex()] = \
            self.widgetTabMachineLearningSettings.tabGeneral.spinBox_MachineLearningMethodsIndex.value()
        self.mlr_Regression.set3rdDimensionSizeToDeepLearningMethods(
            self.dict_machineLearningParameters[self.dkey_mlpMethodIndex()])
        if self.debugMessageFlag:
            print(self.dict_machineLearningParameters[self.dkey_mlpMethodIndex()])

    def actionMultifileTrainingProcessingChange(self):
        self.dict_machineLearningParameters[self.dkey_multifileTrainingProcessing()] = \
            self.widgetTabMachineLearningSettings.tabGeneral.comboBox_MultifileTrainingProcessing.currentText()
        if self.debugMessageFlag:
            print(self.dict_machineLearningParameters[self.dkey_multifileTrainingProcessing()])

    def actionTestPercentageDistributionChange(self):
        self.dict_machineLearningParameters[self.dkey_mlpTestPercentageDistribution()] = \
            self.widgetTabMachineLearningSettings.tabGeneral.comboBox_TestPercentageDistribution.currentText()
        if self.debugMessageFlag:
            print(self.dict_machineLearningParameters[self.dkey_mlpTestPercentageDistribution()])

    def actionHoldoutPercentageDistributionChange(self):
        self.dict_machineLearningParameters[self.dkey_mlpHoldoutPercentageDistribution()] = \
            self.widgetTabMachineLearningSettings.tabGeneral.comboBox_HoldoutPercentageDistribution.currentText()
        if self.debugMessageFlag:
            print(self.dict_machineLearningParameters[self.dkey_mlpHoldoutPercentageDistribution()])

    # ***** SET SETTINGS MACHINE LEARNING REGRESSION METHODS *** #
    # _____ CHECK BOX STATE CHANGE EVENT _____ *
    def actionStateChangeLinearRegression(self):
        state = self.widgetTabMachineLearningSettings.tabRegressionMethods.getCheckState_LinearRegression()
        # print(state)
        self.mlr_Regression.setLinearRegression_sate(state)

    def actionStateChangeRidge(self):
        state = self.widgetTabMachineLearningSettings.tabRegressionMethods.getCheckState_Ridge()
        # print(state)
        self.mlr_Regression.setRidge_state(state)

    def actionStateChangeBayesianRidge(self):
        state = self.widgetTabMachineLearningSettings.tabRegressionMethods.getCheckState_BayesianRidge()
        # print(state)
        self.mlr_Regression.setBayesianRidge_state(state)

    def actionStateChangeLasso(self):
        state = self.widgetTabMachineLearningSettings.tabRegressionMethods.getCheckState_Lasso()
        # print(state)
        self.mlr_Regression.setLasso_state(state)

    def actionStateChangeLassoLars(self):
        state = self.widgetTabMachineLearningSettings.tabRegressionMethods.getCheckState_LassoLars()
        # print(state)
        self.mlr_Regression.setLassoLars_state(state)

    def actionStateChangeTweedieRegressor(self):
        state = self.widgetTabMachineLearningSettings.tabRegressionMethods.getCheckState_TweedieRegressor()
        # print(state)
        self.mlr_Regression.setTweedieRegressor_state(state)

    def actionStateChangeSGDRegressor(self):
        state = self.widgetTabMachineLearningSettings.tabRegressionMethods.getCheckState_SGDRegressor()
        # print(state)
        self.mlr_Regression.setSGDRegressor_state(state)

    def actionStateChangeSVR(self):
        state = self.widgetTabMachineLearningSettings.tabRegressionMethods.getCheckState_SVR()
        print(state)
        self.mlr_Regression.setSVR_state(state)

    def actionStateChangeLinearSVR(self):
        state = self.widgetTabMachineLearningSettings.tabRegressionMethods.getCheckState_LinearSVR()
        # print(state)
        self.mlr_Regression.setLinearSVR_state(state)

    def actionStateChangeNearestNeighbor(self):
        state = self.widgetTabMachineLearningSettings.tabRegressionMethods.getCheckState_NearestNeighbor()
        # print(state)
        self.mlr_Regression.setNearestNeighbor_state(state)

    def actionStateChangeKNeighborsRegressor(self):
        state = self.widgetTabMachineLearningSettings.tabRegressionMethods.getCheckState_KNeighborsRegressor()
        # print(state)
        self.mlr_Regression.setKNeighborsRegressor_state(state)

    def actionStateChangeDecisionTreeRegressor(self):
        state = self.widgetTabMachineLearningSettings.tabRegressionMethods.getCheckState_DecisionTreeRegressor()
        # print(state)
        self.mlr_Regression.setDecisionTreeRegressor_state(state)

    def actionStateChangeRandomForestRegressor(self):
        state = self.widgetTabMachineLearningSettings.tabRegressionMethods.getCheckState_RandomForestRegressor()
        # print(state)
        self.mlr_Regression.setRandomForestRegressor_state(state)

    def actionStateChangeAdaBoostRegressor(self):
        state = self.widgetTabMachineLearningSettings.tabRegressionMethods.getCheckState_AdaBoostRegressor()
        # print(state)
        self.mlr_Regression.setAdaBoostRegressor_state(state)

    def actionStateChangeGradientBoostingRegressor(self):
        state = self.widgetTabMachineLearningSettings.tabRegressionMethods.getCheckState_GradientBoostingRegressor()
        # print(state)
        self.mlr_Regression.setGradientBoostingRegressor_state(state)

    # _____ BUTTON CLICKED EVENT _____ #
    def actionButtonClickedRidge(self):
        self.widgetOptions_Ridge.show()

    def actionButtonClickedBayesianRidge(self):
        pass

    def actionButtonClickedLasso(self):
        pass

    def actionButtonClickedLassoLars(self):
        pass

    def actionButtonClickedTweedieRegressor(self):
        pass

    def actionButtonClickedSGDRegressor(self):
        pass

    def actionButtonClickedSVR(self):
        self.widgetOptions_SVR.show()

    def actionButtonClickedLinearSVR(self):
        pass

    def actionButtonClickedNearestNeighbor(self):
        pass

    def actionButtonClickedKNeighborsRegressor(self):
        pass

    def actionButtonClickedDecisionTreeRegressor(self):
        pass

    def actionButtonClickedRandomForestRegressor(self):
        pass

    def actionButtonClickedAdaBoostRegressor(self):
        pass

    def actionButtonClickedGradientBoostingRegressor(self):
        pass

    # ***** SET SETTINGS MACHINE DEEP LEARNING REGRESSION METHODS *** #
    # _____ CHECK BOX STATE CHANGE EVENT _____ *
    def actionStateChange_Covid_DeepNeuralNetwork(self):
        state = self.widgetTabMachineLearningSettings.tabDeepRegressionMethods.getCheckState_Covid_Convolutional_1D_LongShortTermMemory()
        # print(state)
        self.mlr_Regression.setCovid_CNN1D_LSTM_reg_state(state)

    def actionStateChange_Covid_LongShortTermMemoryNeuralNetwork(self):
        state = self.widgetTabMachineLearningSettings.tabDeepRegressionMethods.getCheckState_Covid_LongShortTermMemoryNeuralNetwork()
        # print(state)
        self.mlr_Regression.setCovid_LSTM_reg_state(state)

    def actionStateChange_Covid_LongShortTermMemoryNeuralNetwork_Simple(self):
        state = self.widgetTabMachineLearningSettings.tabDeepRegressionMethods.getCheckState_Covid_LongShortTermMemoryNeuralNetwork_Simple()
        # print(state)
        self.mlr_Regression.setCovid_LSTM_Simple_reg_state(state)

    def actionStateChange_Covid_SimpleRecurrentNeuralNetwork(self):
        state = self.widgetTabMachineLearningSettings.tabDeepRegressionMethods.getCheckState_Covid_SimpleRecurrentNeuralNetwork()
        # print(state)
        self.mlr_Regression.setCovid_SimpleRNN_reg_state(state)

    # ***** SET SETTINGS SIGNAL COMPARE METHODS *** #
    # _____ CHECK BOX STATE CHANGE EVENT _____ *
    def actionStateChangePearsonCorr(self):
        state = self.widgetTabMachineLearningSettings.tabSignalCompare.getCheckState_PearsonCorr()
        self.signComp_Methods.setPearsonCorr_state(state)

    def actionStateChangeTimeLagCrossCorrelation(self):
        state = self.widgetTabMachineLearningSettings.tabSignalCompare.getCheckState_TimeLaggedCrossCorrelation()
        self.signComp_Methods.setTimeLaggedCrossCorrelation_state(state)

    def actionStateChangeTimeLagCrossCorrelationNoSplits(self):
        state = self.widgetTabMachineLearningSettings.tabSignalCompare.getCheckState_TimeLaggedCrossCorrelationNoSplits()
        self.signComp_Methods.setTimeLaggedCrossCorrelationNoSplits_state(state)

    def actionStateChangeRollingWindowTimeLagCrossCorrelation(self):
        state = self.widgetTabMachineLearningSettings.tabSignalCompare.getCheckState_RollingWindowTimeLaggedCrossCorrelation()
        self.signComp_Methods.setRollingWindowTimeLaggedCrossCorrelation_state(state)

    def actionStateChangeDynamicTimeWarping(self):
        state = self.widgetTabMachineLearningSettings.tabSignalCompare.getCheckState_DynamicTimeWarping()
        self.signComp_Methods.setDynamicTimeWarping_state(state)

    # ***** MACHINE LEARNING WIDGET OPTIONS EVENTS *** #
    # _____ RIDGE _____ #
    def actionButtonClicked_Ridge_TolAdd(self):
        _items_ = self.widgetOptions_Ridge.getListOptionSelectedItems_Tol()
        for _it_ in _items_:
            stateExist = self.widgetOptions_Ridge.checkIfItemAlreadyInList_Tol(_it_)
            if not stateExist:
                self.widgetOptions_Ridge.addToSelectedList_Tol(_it_)

    def actionButtonClicked_Ridge_SolverAdd(self):
        _items_ = self.widgetOptions_Ridge.getListOptionSelectedItems_Solver()
        for _it_ in _items_:
            stateExist = self.widgetOptions_Ridge.checkIfItemAlreadyInList_Solver(_it_)
            if not stateExist:
                self.widgetOptions_Ridge.addToSelectedList_Solver(_it_)

    def actionChange_Ridge_AlphaMin(self):
        value = round(self.widgetOptions_Ridge.getAlphaMin(), 3)
        # print(value)
        self.mlr_Regression.setRidge_alphaMin(value)

    def actionChange_Ridge_AlphaMax(self):
        value = round(self.widgetOptions_Ridge.getAlphaMax(), 3)
        # print(value)
        self.mlr_Regression.setRidge_alphaMax(value)

    def actionChange_Ridge_AlphaStep(self):
        value = round(self.widgetOptions_Ridge.getAlphaStep(), 3)
        # print(value)
        self.mlr_Regression.setRidge_alphaStep(value)

    def actionChange_Ridge_SelectedTol(self):
        value = self.widgetOptions_Ridge.getSelectedList_Tol()
        # print(value)
        self.mlr_Regression.setRidge_Tol(value)

    def actionChange_Ridge_SelectedSolver(self):
        value = self.widgetOptions_Ridge.getSelectedList_Solver()
        # print(value)
        self.mlr_Regression.setRidge_Solver(value)

    # _____ SVR _____ #
    def actionButtonClicked_SVR_KernelAdd(self):
        _items_ = self.widgetOptions_SVR.getListOptionSelectedItems_Kernel()
        for _it_ in _items_:
            stateExist = self.widgetOptions_SVR.checkIfItemAlreadyInList_Kernel(_it_)
            if not stateExist:
                self.widgetOptions_SVR.addToSelectedList_Kernel(_it_)

    def actionButtonClicked_SVR_GammaAdd(self):
        _items_ = self.widgetOptions_SVR.getListOptionSelectedItems_Gamma()
        for _it_ in _items_:
            stateExist = self.widgetOptions_SVR.checkIfItemAlreadyInList_Gamma(_it_)
            if not stateExist:
                self.widgetOptions_SVR.addToSelectedList_Gamma(_it_)

    def actionButtonClicked_SVR_TolAdd(self):
        _items_ = self.widgetOptions_SVR.getListOptionSelectedItems_Tol()
        for _it_ in _items_:
            stateExist = self.widgetOptions_SVR.checkIfItemAlreadyInList_Tol(_it_)
            if not stateExist:
                self.widgetOptions_SVR.addToSelectedList_Tol(_it_)

    def actionChange_SVR_SelectedKernel(self):
        value = self.widgetOptions_SVR.getSelectedList_Kernel()
        # print(value)
        self.mlr_Regression.setSVR_Kernel(value)

    def actionChange_SVR_SelectedGamma(self):
        value = self.widgetOptions_SVR.getSelectedList_Gamma()
        # print(value)
        self.mlr_Regression.setSVR_Gamma(value)

    def actionChange_SVR_SelectedTol(self):
        value = self.widgetOptions_SVR.getSelectedList_Tol()
        # print(value)
        self.mlr_Regression.setSVR_Tol(value)


# *********************************** #
# *********** Tab Widgets *********** #
# *********************************** #
# *                                 * #

# *********** Machine Learning I/O *********** #
class WidgetTabInputOutput(QWidget):
    def __init__(self):
        super().__init__()

        self.setStyleSheet(setStyle_())  # set the tab style

        # ---------------------- #
        # ----- Set Window ----- #
        # ---------------------- #
        self.vbox_main_layout = QVBoxLayout(self)  # Create the main vbox

        # -------------------------- #
        # ----- Set PushButton ----- #
        # -------------------------- #
        self.buttonInputColumn = QPushButton("Add Input Column (X)")
        self.buttonInputColumn.setMinimumWidth(INT_BUTTON_MIN_WIDTH)  # Set Minimum Width
        self.buttonInputColumn.setMinimumHeight(INT_BUTTON_MIN_WIDTH / 2)  # Set Minimum Height
        self.buttonInputColumn.setShortcut("I")  # Set Shortcut
        self.buttonInputColumn.setToolTip('Set selected column as Input Column.')  # Add Description

        self.buttonRemInputColumn = QPushButton("Remove")
        self.buttonRemInputColumn.setMinimumWidth(INT_BUTTON_MIN_WIDTH)  # Set Minimum Width
        self.buttonRemInputColumn.setMinimumHeight(INT_BUTTON_MIN_WIDTH / 2)  # Set Minimum Height
        self.buttonRemInputColumn.setToolTip('Remove selected column from Input List.')  # Add Description

        self.buttonOutputColumn = QPushButton("Add Output Column (Y)")
        self.buttonOutputColumn.setMinimumWidth(INT_BUTTON_MIN_WIDTH)  # Set Minimum Width
        self.buttonOutputColumn.setMinimumHeight(INT_BUTTON_MIN_WIDTH / 2)  # Set Minimum Height
        self.buttonOutputColumn.setShortcut("O")  # Set Shortcut
        self.buttonOutputColumn.setToolTip('Set selected column as Output Column.')  # Add Description

        self.buttonRemOutputColumn = QPushButton("Remove")
        self.buttonRemOutputColumn.setMinimumWidth(INT_BUTTON_MIN_WIDTH)  # Set Minimum Width
        self.buttonRemOutputColumn.setMinimumHeight(INT_BUTTON_MIN_WIDTH / 2)  # Set Minimum Height
        self.buttonRemOutputColumn.setToolTip('Remove selected column from Output List.')  # Add Description

        self.buttonPrimaryEvent = QPushButton("Primary Event")
        self.buttonPrimaryEvent.setMinimumWidth(INT_BUTTON_MIN_WIDTH)  # Set Minimum Width
        self.buttonPrimaryEvent.setMinimumHeight(INT_BUTTON_MIN_WIDTH / 2)  # Set Minimum Height
        self.buttonPrimaryEvent.setShortcut("P")  # Set Shortcut
        self.buttonPrimaryEvent.setToolTip('Set selected column as Primary Event.')  # Add Description

        self.buttonRemPrimaryEvent = QPushButton("Remove")
        self.buttonRemPrimaryEvent.setMinimumWidth(INT_BUTTON_MIN_WIDTH)  # Set Minimum Width
        self.buttonRemPrimaryEvent.setMinimumHeight(INT_BUTTON_MIN_WIDTH / 2)  # Set Minimum Height
        self.buttonRemPrimaryEvent.setToolTip('Remove selected column from Primary Event.')  # Add Description

        # -------------------------------- #
        # ----- Set QListWidgetItems ----- #
        # -------------------------------- #
        self.listWidget_InputColumns = QListWidget()
        self.listWidget_InputColumns.setSelectionMode(QListWidget.ExtendedSelection)
        self.listWidget_OutputColumns = QListWidget()
        self.listWidget_OutputColumns.setSelectionMode(QListWidget.ExtendedSelection)
        self.listWidget_PrimaryEvent = QListWidget()
        self.listWidget_PrimaryEvent.setSelectionMode(QListWidget.ExtendedSelection)

    # --------------------------- #
    # ----- Reuse Functions ----- #
    # --------------------------- #
    def setWidget(self):
        # Set column/remove buttons in vbox
        hbox_listInputButtons = QHBoxLayout()  # Create a Horizontal Box Layout
        hbox_listInputButtons.addWidget(self.buttonInputColumn)  # Add buttonDate
        hbox_listInputButtons.addWidget(self.buttonRemInputColumn)  # Add buttonRemove

        hbox_listOutputButtons = QHBoxLayout()  # Create a Horizontal Box Layout
        hbox_listOutputButtons.addWidget(self.buttonOutputColumn)  # Add buttonTime
        hbox_listOutputButtons.addWidget(self.buttonRemOutputColumn)  # Add buttonRemove

        hbox_listPrimaryButtons = QHBoxLayout()  # Create a Horizontal Box Layout
        hbox_listPrimaryButtons.addWidget(self.buttonPrimaryEvent)  # Add buttonTime
        hbox_listPrimaryButtons.addWidget(self.buttonRemPrimaryEvent)  # Add buttonRemove

        # Set Input hbox
        labelInputList = QLabel("Input Columns:\n" +
                                "(the columns that will be used as training/test/validation " +
                                "input\nfor the machine learning)")
        vbox_listInputColumns = QVBoxLayout()  # Create a Horizontal Box Layout
        vbox_listInputColumns.addWidget(labelInputList)  # Add Label
        vbox_listInputColumns.addWidget(self.listWidget_InputColumns)  # Add Column List
        vbox_listInputColumns.addLayout(hbox_listInputButtons)  # Add Layout

        # Set Output hbox
        labelOutputList = QLabel("Output Columns:\n" +
                                 "(the columns that will be used as training/test/validation " +
                                 "output\nfor the machine learning)")
        vbox_listOutputColumns = QVBoxLayout()  # Create a Horizontal Box Layout
        vbox_listOutputColumns.addWidget(labelOutputList)  # Add Label
        vbox_listOutputColumns.addWidget(self.listWidget_OutputColumns)  # Add Column List
        vbox_listOutputColumns.addLayout(hbox_listOutputButtons)  # Add Layout

        # Set Primary Event hbox
        labelPrimaryEventList = QLabel("Primary/Common Event Column (Optional):")
        vbox_listPrimaryEvent = QVBoxLayout()  # Create a Horizontal Box Layout
        vbox_listPrimaryEvent.addWidget(labelPrimaryEventList)  # Add Label
        vbox_listPrimaryEvent.addWidget(self.listWidget_PrimaryEvent)  # Add Column List
        vbox_listPrimaryEvent.addLayout(hbox_listPrimaryButtons)  # Add Layout

        # Set ListWidget in hbox
        hbox_listWidget = QHBoxLayout()  # Create Horizontal Layout
        hbox_listWidget.addLayout(vbox_listInputColumns)  # Add vbox_Combine_1 layout
        hbox_listWidget.addLayout(vbox_listOutputColumns)  # Add vbox_Combine_2 layout

        # Set a vbox
        vbox_finalListWidget = QVBoxLayout()
        vbox_finalListWidget.addLayout(hbox_listWidget)
        vbox_finalListWidget.addLayout(vbox_listPrimaryEvent)

        self.vbox_main_layout.addLayout(vbox_finalListWidget)


# *********** Machine Learning Settings *********** #
class WidgetTabMachineLearningSettings(QWidget):
    def __init__(self):
        super().__init__()

        self.setStyleSheet(setStyle_())

        # ---------------------- #
        # ----- Set Window ----- #
        # ---------------------- #
        self.vbox_main_layout = QVBoxLayout(self)  # Create the main vbox

        # ------------------- #
        # ----- Set Tab ----- #
        # ------------------- #
        self.mainTabWidget = QTabWidget()  # Create a main widget tab

        self.tabGeneral = WidgetTabMachineLearningSettingsGeneral()  # create a tab for General (info)
        self.tabRegressionMethods = WidgetTabMachineLearningSettingsRegressionMethods()
        self.tabDeepRegressionMethods = WidgetTabMachineLearningSettingsDeepRegressionMethods()
        self.tabSignalCompare = WidgetTabMachineLearningSettingsSignalCompareMethods()

    def setWidget(self):
        """
            A function to create the widget components into the main QWidget
            :return: Nothing
        """
        self.tabGeneral.setWidget()  # set tab General (info)
        self.tabRegressionMethods.setWidget()  # set tab Regression methods
        self.tabDeepRegressionMethods.setWidget()  # set tab Deep Regression methods
        self.tabSignalCompare.setWidget()  # set tab Signal Compare methods
        self.mainTabWidget.addTab(self.tabGeneral, "General")  # add tab to mainTabWidget
        self.mainTabWidget.addTab(self.tabRegressionMethods, "Regression Methods")  # add tab to mainTabWidget
        self.mainTabWidget.addTab(self.tabDeepRegressionMethods, "Deep Regression Methods")
        self.mainTabWidget.addTab(self.tabSignalCompare, "Signal Compare Methods")  # add tab to mainTabWidget
        self.vbox_main_layout.addWidget(self.mainTabWidget)  # add tabWidget to main layout


# *********** Machine Learning Settings --> General *********** #
class WidgetTabMachineLearningSettingsGeneral(QWidget):
    def __init__(self):
        super().__init__()

        self.setStyleSheet(setStyle_())  # set the tab style

        # ---------------------- #
        # ----- Set Window ----- #
        # ---------------------- #
        self.vbox_main_layout = QVBoxLayout(self)  # Create the main vbox

        # ----------------------- #
        # ----- QPushButton ----- #
        # ----------------------- #
        self.buttonRestoreDefault = QPushButton("Restore Default")
        self.buttonRestoreDefault.setMinimumWidth(150)  # Set Minimum Width
        self.buttonRestoreDefault.setMinimumHeight(30)  # Set Minimum Height
        self.buttonRestoreDefault.setToolTip('Restore the default values.')  # Add Description

        self.buttonSetOutPath = QPushButton("...")
        self.buttonSetOutPath.setMinimumWidth(INT_ADD_REMOVE_BUTTON_SIZE)  # Set Minimum Width
        self.buttonSetOutPath.setMaximumWidth(INT_ADD_REMOVE_BUTTON_SIZE)  # Set Maximum Width
        self.buttonSetOutPath.setMinimumHeight(30)  # Set Minimum Height
        self.buttonSetOutPath.setToolTip('Select the output path')  # Add Description

        # -------------------- #
        # ----- QSpinBox ----- #
        # -------------------- #
        # create a spinBox for the ExperimentNumber
        self.spinBox_ExperimentNumber = QSpinBox()
        # set the minimum value to 1 (at least an experiments must be performed)
        self.spinBox_ExperimentNumber.setMinimum(1)

        # create a spinBox for the MachineLearningMethodIndex
        self.spinBox_MachineLearningMethodsIndex = QSpinBox()
        # set the minimum value to 1 (at least a row needs to be used as input/output)
        self.spinBox_MachineLearningMethodsIndex.setMinimum(1)

        # -------------------------- #
        # ----- QDoubleSpinBox ----- #
        # -------------------------- #
        # create a doubleSpinBox for the test percentage
        self.doubleSpinBox_TestPercentage = QDoubleSpinBox()
        # set minimum value
        self.doubleSpinBox_TestPercentage.setMinimum(MLF_TEST_PERCENTAGE_MIN)
        # set maximum value
        self.doubleSpinBox_TestPercentage.setMaximum(MLF_TEST_PERCENTAGE_MAX)
        # set step
        self.doubleSpinBox_TestPercentage.setSingleStep(MLF_TEST_STEP_WIDGET)
        # create a doubleSpinBox fot the holdout percentage
        self.doubleSpinBox_HoldoutPercentage = QDoubleSpinBox()
        # set minimum value
        self.doubleSpinBox_HoldoutPercentage.setMinimum(MLF_HOLDOUT_PERCENTAGE_MIN)
        # set maximum value
        self.doubleSpinBox_HoldoutPercentage.setMaximum(MLF_HOLDOUT_PERCENTAGE_MAX)
        # set step
        self.doubleSpinBox_HoldoutPercentage.setSingleStep(MLF_HOLDOUT_STEP_WIDGET)

        # --------------------- #
        # ----- QLineEdit ----- #
        # --------------------- #
        self.lineEdit_SetOutPath = QLineEdit()
        self.lineEdit_SetOutPath.setEnabled(False)

        # --------------------- #
        # ----- QComboBox ----- #
        # --------------------- #
        # MachineLearningMethods
        self.comboBox_MachineLearningMethods = QComboBox()
        self.comboBox_MachineLearningMethods.setMinimumWidth(250)
        self.comboBox_MachineLearningMethods.addItems(MLPF_METHOD_LIST_REGRESSION)

        # MultifileTrainingProcessing
        self.comboBox_MultifileTrainingProcessing = QComboBox()
        self.comboBox_MultifileTrainingProcessing.setMinimumWidth(200)
        self.comboBox_MultifileTrainingProcessing.addItems(MLPF_MULTIFILE_TRAINING_PROCESSING_LIST)

        # TestPercentageDistribution
        self.comboBox_TestPercentageDistribution = QComboBox()
        self.comboBox_TestPercentageDistribution.setMinimumWidth(200)
        self.comboBox_TestPercentageDistribution.addItems(MLPF_PERCENTAGE_DISTRIBUTION_LIST)

        # HoldoutPercentageDistribution
        self.comboBox_HoldoutPercentageDistribution = QComboBox()
        self.comboBox_HoldoutPercentageDistribution.setMinimumWidth(200)
        self.comboBox_HoldoutPercentageDistribution.addItems(MLPF_PERCENTAGE_DISTRIBUTION_LIST)

        # ------------------------------ #
        # ----- Set Default Values ----- #
        # ------------------------------ #
        self._experimentNumberDefaultValue = MLF_DEFAULT_EXPERIMENT_VALUE
        self._testPercentageDefaultValue = MLF_DEFAULT_TEST_PERCENTAGE
        self._holdoutPercentageDefaultValue = MLF_DEFAULT_HOLDOUT_PERCENTAGE
        self._exportPathDefaultValue = MLF_DEFAULT_EXPORT_FOLDER_PATH
        self._mlMethodDefaultValue = MLF_DEFAULT_METHOD
        self._mlMethodIndexDefaultValue = MLF_DEFAULT_METHOD_INDEX
        self._mlMultifileTrainingProcessingDefaultValue = MLF_DEFAULT_MULTIFILE_TRAINING_PROCESSING
        self._testPercentageDistributionDefaultValue = MLF_DEFAULT_TEST_PERCENTAGE_DISTRIBUTION
        self._holdoutPercentageDistributionDefaultValue = MLF_DEFAULT_HOLDOUT_PERCENTAGE_DISTRIBUTION

    # --------------------------- #
    # ----- Reuse Functions ----- #
    # --------------------------- #
    def setWidget(self):
        """
            A function to create the widget components into the main QWidget
            :return: Nothing
        """
        self.restoreDefaultValues()
        self.setEvents_()

        # Set Label
        label_ExperimentalNumber = QLabel("Number of Experiments:")
        # label_ExperimentalNumber.setMinimumWidth(100)
        label_TestPercentage = QLabel("Test Percentage (0.00-0.75):")
        # label_TestPercentage.setMinimumWidth(100)
        label_TestPercentageDistribution = QLabel("Test Percentage Distribution:")
        # label_TestPercentageDistribution.setMinimumWidth(100)
        label_HoldOutPercentage = QLabel("Hold-Out Percentage (0.00-0.75):")
        # label_HoldOutPercentage.setMinimumWidth(100)
        label_HoldoutPercentageDistribution = QLabel("Hold-Out Percentage Distribution:")
        # label_HoldoutPercentageDistribution.setMinimumWidth(100)
        label_SetOutPath = QLabel("Export Folder:")
        # label_SetOutPath.setMinimumWidth(100)
        label_MachineLearningMethod = QLabel("Machine Learning Method:")
        # label_MachineLearningMethod.setMinimumWidth(100)
        label_MachineLearningMethodIndex = QLabel("Method Usage Index:")
        # label_MachineLearningMethodIndex.setMinimumWidth(100)
        label_MultifileTrainingProcessing = QLabel("Multifile Training Processing:")
        # label_MultifileTrainingProcessing.setMinimumWidth(100)

        # Set SpinBoxes
        hbox_ExperimentalNumber = QHBoxLayout()
        hbox_ExperimentalNumber.addWidget(label_ExperimentalNumber)
        hbox_ExperimentalNumber.addWidget(self.spinBox_ExperimentNumber)
        hbox_ExperimentalNumber.addSpacerItem(QSpacerItem(INT_MAX_STRETCH, 0))

        hbox_TestPercentage = QHBoxLayout()
        hbox_TestPercentage.addWidget(label_TestPercentage)
        hbox_TestPercentage.addWidget(self.doubleSpinBox_TestPercentage)
        hbox_TestPercentage.addSpacerItem(QSpacerItem(INT_MAX_STRETCH, 0))

        hbox_TestPercentageDistribution = QHBoxLayout()
        hbox_TestPercentageDistribution.addWidget(label_TestPercentageDistribution)
        hbox_TestPercentageDistribution.addWidget(self.comboBox_TestPercentageDistribution)
        hbox_TestPercentageDistribution.addSpacerItem(QSpacerItem(INT_MAX_STRETCH, 0))

        hbox_HoldOutPercentage = QHBoxLayout()
        hbox_HoldOutPercentage.addWidget(label_HoldOutPercentage)
        hbox_HoldOutPercentage.addWidget(self.doubleSpinBox_HoldoutPercentage)
        hbox_HoldOutPercentage.addSpacerItem(QSpacerItem(INT_MAX_STRETCH, 0))

        hbox_HoldOutPercentageDistribution = QHBoxLayout()
        hbox_HoldOutPercentageDistribution.addWidget(label_HoldoutPercentageDistribution)
        hbox_HoldOutPercentageDistribution.addWidget(self.comboBox_HoldoutPercentageDistribution)
        hbox_HoldOutPercentageDistribution.addSpacerItem(QSpacerItem(INT_MAX_STRETCH, 0))

        vbox_finalSpinBoxes = QVBoxLayout()
        vbox_finalSpinBoxes.addLayout(hbox_ExperimentalNumber)
        vbox_finalSpinBoxes.addLayout(hbox_TestPercentage)
        vbox_finalSpinBoxes.addLayout(hbox_TestPercentageDistribution)
        vbox_finalSpinBoxes.addLayout(hbox_HoldOutPercentage)
        vbox_finalSpinBoxes.addLayout(hbox_HoldOutPercentageDistribution)

        # SetOutPath Layout
        hbox_SetOutPath = QHBoxLayout()
        hbox_SetOutPath.addWidget(label_SetOutPath)
        hbox_SetOutPath.addWidget(self.lineEdit_SetOutPath)
        hbox_SetOutPath.addWidget(self.buttonSetOutPath)
        # hbox_SetOutPath.addSpacerItem(QSpacerItem(INT_MAX_STRETCH, 0))

        # Button Layout
        hbox_buttons = QHBoxLayout()
        hbox_buttons.addSpacerItem(QSpacerItem(INT_MAX_STRETCH, 0))
        hbox_buttons.addWidget(self.buttonRestoreDefault)

        # MachineLearningMethods
        hbox_MachineLearningMethods = QHBoxLayout()
        hbox_MachineLearningMethods.addWidget(label_MachineLearningMethod)
        hbox_MachineLearningMethods.addWidget(self.comboBox_MachineLearningMethods)
        hbox_MachineLearningMethods.addSpacerItem(QSpacerItem(INT_MAX_STRETCH, 0))

        hbox_MachineLearningMethodsIndex = QHBoxLayout()
        hbox_MachineLearningMethodsIndex.addWidget(label_MachineLearningMethodIndex)
        hbox_MachineLearningMethodsIndex.addWidget(self.spinBox_MachineLearningMethodsIndex)
        hbox_MachineLearningMethodsIndex.addSpacerItem(QSpacerItem(INT_MAX_STRETCH, 0))

        vbox_finalMachineLearningMethods = QVBoxLayout()
        vbox_finalMachineLearningMethods.addLayout(hbox_MachineLearningMethods)
        vbox_finalMachineLearningMethods.addLayout(hbox_MachineLearningMethodsIndex)

        # MultifileTrainingProcessing
        hbox_MultifileTrainingProcessing = QHBoxLayout()
        hbox_MultifileTrainingProcessing.addWidget(label_MultifileTrainingProcessing)
        hbox_MultifileTrainingProcessing.addWidget(self.comboBox_MultifileTrainingProcessing)
        hbox_MultifileTrainingProcessing.addSpacerItem(QSpacerItem(INT_MAX_STRETCH, 0))

        self.vbox_main_layout.addLayout(vbox_finalSpinBoxes)
        self.vbox_main_layout.addLayout(hbox_SetOutPath)
        self.vbox_main_layout.addLayout(vbox_finalMachineLearningMethods)
        self.vbox_main_layout.addLayout(hbox_MultifileTrainingProcessing)
        # self.vbox_main_layout.addSpacerItem(QSpacerItem(0, INT_MAX_STRETCH))
        self.vbox_main_layout.addLayout(hbox_buttons)

    def restoreDefaultValues(self):
        # set default value
        self.spinBox_ExperimentNumber.setValue(self._experimentNumberDefaultValue)
        self.doubleSpinBox_TestPercentage.setValue(self._testPercentageDefaultValue)
        self.doubleSpinBox_HoldoutPercentage.setValue(self._holdoutPercentageDefaultValue)
        self.lineEdit_SetOutPath.setText(os.path.normpath(self._exportPathDefaultValue))
        self.comboBox_MachineLearningMethods.setCurrentText(self._mlMethodDefaultValue)
        self.spinBox_MachineLearningMethodsIndex.setValue(self._mlMethodIndexDefaultValue)
        self.comboBox_MultifileTrainingProcessing.setCurrentText(self._mlMultifileTrainingProcessingDefaultValue)
        self.comboBox_TestPercentageDistribution.setCurrentText(self._testPercentageDistributionDefaultValue)
        self.comboBox_HoldoutPercentageDistribution.setCurrentText(self._holdoutPercentageDistributionDefaultValue)

    def setEvents_(self):
        self.buttonRestoreDefault.clicked.connect(self.restoreDefaultValues)

    # ------------------------------ #
    # ----- GET DEFAULT VALUES ----- #
    # ------------------------------ #
    def getDefaultExperimentNumber(self):
        return self._experimentNumberDefaultValue

    def getDefaultTestPercentage(self):
        return self._testPercentageDefaultValue

    def getDefaultTestPercentageDistribution(self):
        return self._testPercentageDistributionDefaultValue

    def getDefaultHoldoutPercentage(self):
        return self._holdoutPercentageDefaultValue

    def getDefaultHoldoutPercentageDistribution(self):
        return self._holdoutPercentageDistributionDefaultValue

    def getDefaultExportPath(self):
        return self._exportPathDefaultValue

    def getDefaultMethod(self):
        return self._mlMethodDefaultValue

    def getDefaultMethodIndex(self):
        return self._mlMethodIndexDefaultValue

    def getDefaultMultifileTrainingProcessing(self):
        return self._mlMultifileTrainingProcessingDefaultValue


# *********** Machine Learning Settings --> Regression Methods *********** #
class WidgetTabMachineLearningSettingsRegressionMethods(QWidget):
    def __init__(self):
        super().__init__()

        self.setStyleSheet(setStyle_())  # set the tab style

        # ---------------------- #
        # ----- Set Window ----- #
        # ---------------------- #
        self.vbox_main_layout = QVBoxLayout(self)  # Create the main vbox

        # ----------------------- #
        # ----- PushButtons ----- #
        # ----------------------- #
        _icon = QIcon(QPixmap(ICON_OPTION_SETTINGS))
        self.button_Ridge = QPushButton()
        self.button_BayesianRidge = QPushButton()
        self.button_Lasso = QPushButton()
        self.button_LassoLars = QPushButton()
        self.button_TweedieRegressor = QPushButton()
        self.button_SGDRegressor = QPushButton()
        self.button_SVR = QPushButton()
        self.button_LinearSVR = QPushButton()
        self.button_NearestNeighbor = QPushButton()
        self.button_KNeighborsRegressor = QPushButton()
        self.button_DecisionTreeRegressor = QPushButton()
        self.button_RandomForestRegressor = QPushButton()
        self.button_AdaBoostRegressor = QPushButton()
        self.button_GradientBoostingRegressor = QPushButton()

        self.button_Ridge.setIcon(_icon)
        self.button_BayesianRidge.setIcon(_icon)
        self.button_Lasso.setIcon(_icon)
        self.button_LassoLars.setIcon(_icon)
        self.button_TweedieRegressor.setIcon(_icon)
        self.button_SGDRegressor.setIcon(_icon)
        self.button_SVR.setIcon(_icon)
        self.button_LinearSVR.setIcon(_icon)
        self.button_NearestNeighbor.setIcon(_icon)
        self.button_KNeighborsRegressor.setIcon(_icon)
        self.button_DecisionTreeRegressor.setIcon(_icon)
        self.button_RandomForestRegressor.setIcon(_icon)
        self.button_AdaBoostRegressor.setIcon(_icon)
        self.button_GradientBoostingRegressor.setIcon(_icon)

        self.button_BayesianRidge.setEnabled(False)
        self.button_Lasso.setEnabled(False)
        self.button_LassoLars.setEnabled(False)
        self.button_TweedieRegressor.setEnabled(False)
        self.button_SGDRegressor.setEnabled(False)
        self.button_LinearSVR.setEnabled(False)
        self.button_NearestNeighbor.setEnabled(False)
        self.button_KNeighborsRegressor.setEnabled(False)
        self.button_DecisionTreeRegressor.setEnabled(False)
        self.button_RandomForestRegressor.setEnabled(False)
        self.button_AdaBoostRegressor.setEnabled(False)
        self.button_GradientBoostingRegressor.setEnabled(False)

        # ---------------------- #
        # ----- CheckBoxes ----- #
        # ---------------------- #
        self.checkbox_LinearRegression = QCheckBox()
        self.checkbox_Ridge = QCheckBox()
        self.checkbox_BayesianRidge = QCheckBox()
        self.checkbox_Lasso = QCheckBox()
        self.checkbox_LassoLars = QCheckBox()
        self.checkbox_TweedieRegressor = QCheckBox()
        self.checkbox_SGDRegressor = QCheckBox()
        self.checkbox_SVR = QCheckBox()
        self.checkbox_LinearSVR = QCheckBox()
        self.checkbox_NearestNeighbor = QCheckBox()
        self.checkbox_KNeighborsRegressor = QCheckBox()
        self.checkbox_DecisionTreeRegressor = QCheckBox()
        self.checkbox_RandomForestRegressor = QCheckBox()
        self.checkbox_AdaBoostRegressor = QCheckBox()
        self.checkbox_GradientBoostingRegressor = QCheckBox()

        self.checkbox_BayesianRidge.setEnabled(False)
        self.checkbox_Lasso.setEnabled(False)
        self.checkbox_LassoLars.setEnabled(False)
        self.checkbox_TweedieRegressor.setEnabled(False)
        self.checkbox_SGDRegressor.setEnabled(False)
        self.checkbox_LinearSVR.setEnabled(False)
        self.checkbox_NearestNeighbor.setEnabled(False)
        self.checkbox_KNeighborsRegressor.setEnabled(False)
        self.checkbox_DecisionTreeRegressor.setEnabled(False)
        self.checkbox_RandomForestRegressor.setEnabled(False)
        self.checkbox_AdaBoostRegressor.setEnabled(False)
        self.checkbox_GradientBoostingRegressor.setEnabled(False)

        # ---------------------- #
        # ----- ScrollArea ----- #
        # ---------------------- #
        self.scrollArea_regMethods = QScrollArea()

    # --------------------------- #
    # ----- Reuse Functions ----- #
    # --------------------------- #
    def setWidget(self):
        """
            A function to create the widget components into the main QWidget
            :return: Nothing
        """
        self.scrollArea_regMethods.setWidgetResizable(True)
        self.scrollArea_regMethods.setWidget(self._setGridLayout())
        self.vbox_main_layout.addWidget(self.scrollArea_regMethods)

    def _setGridLayout(self):
        # label_min_width = 200

        # Set Label
        label_Method = QLabel('<b><u>Method<\\u><\\b>')
        # label_Method.setMaximumHeight(30)

        label_State = QLabel('<b><u>State<\\u><\\b>')
        # label_State.setMaximumHeight(30)

        label_Options = QLabel('<b><u>Options<\\u><\\b>')
        # label_Options.setMaximumHeight(30)

        label_LinearRegression = QLabel(mlr.MLR_REG_LINEAR_REGRESSION)
        label_Ridge = QLabel(mlr.MLR_REG_RIDGE)
        label_BayesianRidge = QLabel(mlr.MLR_REG_BAYESIAN_RIDGE)
        label_Lasso = QLabel(mlr.MLR_REG_LASSO)
        label_LassoLars = QLabel(mlr.MLR_REG_LASSO_LARS)
        label_TweedieRegressor = QLabel(mlr.MLR_REG_TWEEDIE_REGRESSOR)
        label_SGDRegressor = QLabel(mlr.MLR_REG_SGD_REGRESSOR)
        label_SVR = QLabel(mlr.MLR_REG_SVR)
        label_LinearSVR = QLabel(mlr.MLR_REG_LINEAR_SVR)
        label_NearestNeighbor = QLabel(mlr.MLR_REG_NEAREST_NEIGHBORS)
        label_KNeighborsRegressor = QLabel(mlr.MLR_REG_K_NEIGHBORS_REGRESSOR)
        label_DecisionTreeRegressor = QLabel(mlr.MLR_REG_DECISION_TREE_REGRESSOR)
        label_RandomForestRegressor = QLabel(mlr.MLR_REG_RANDOM_FOREST_REGRESSOR)
        label_AdaBoostRegressor = QLabel(mlr.MLR_REG_ADA_BOOST_REGRESSOR)
        label_GradientBoostingRegressor = QLabel(mlr.MLR_REG_GRADIENT_BOOSTING_REGRESSOR)

        # Set layout
        scrollAreaWidget = QWidget()
        scrollAreaWidget.setMaximumWidth(840)
        scrollAreaWidget.setMaximumHeight(512)
        gridBox_Methods = QGridLayout(scrollAreaWidget)

        gridBox_Methods.addWidget(label_Method, 0, 0, alignment=Qt.AlignLeft)
        gridBox_Methods.addWidget(label_State, 0, 1, alignment=Qt.AlignCenter)
        gridBox_Methods.addWidget(label_Options, 0, 2, alignment=Qt.AlignCenter)

        gridBox_Methods.addWidget(label_LinearRegression, 1, 0, alignment=Qt.AlignLeft)
        gridBox_Methods.addWidget(self.checkbox_LinearRegression, 1, 1, alignment=Qt.AlignHCenter)

        gridBox_Methods.addWidget(label_Ridge, 2, 0, alignment=Qt.AlignLeft)
        gridBox_Methods.addWidget(self.checkbox_Ridge, 2, 1, alignment=Qt.AlignHCenter)
        gridBox_Methods.addWidget(self.button_Ridge, 2, 2, alignment=Qt.AlignHCenter)

        gridBox_Methods.addWidget(label_BayesianRidge, 3, 0, alignment=Qt.AlignLeft)
        gridBox_Methods.addWidget(self.checkbox_BayesianRidge, 3, 1, alignment=Qt.AlignHCenter)
        gridBox_Methods.addWidget(self.button_BayesianRidge, 3, 2, alignment=Qt.AlignHCenter)

        gridBox_Methods.addWidget(label_Lasso, 4, 0, alignment=Qt.AlignLeft)
        gridBox_Methods.addWidget(self.checkbox_Lasso, 4, 1, alignment=Qt.AlignHCenter)
        gridBox_Methods.addWidget(self.button_Lasso, 4, 2, alignment=Qt.AlignHCenter)

        gridBox_Methods.addWidget(label_LassoLars, 5, 0, alignment=Qt.AlignLeft)
        gridBox_Methods.addWidget(self.checkbox_LassoLars, 5, 1, alignment=Qt.AlignHCenter)
        gridBox_Methods.addWidget(self.button_LassoLars, 5, 2, alignment=Qt.AlignHCenter)

        gridBox_Methods.addWidget(label_TweedieRegressor, 6, 0, alignment=Qt.AlignLeft)
        gridBox_Methods.addWidget(self.checkbox_TweedieRegressor, 6, 1, alignment=Qt.AlignHCenter)
        gridBox_Methods.addWidget(self.button_TweedieRegressor, 6, 2, alignment=Qt.AlignHCenter)

        gridBox_Methods.addWidget(label_SGDRegressor, 7, 0, alignment=Qt.AlignLeft)
        gridBox_Methods.addWidget(self.checkbox_SGDRegressor, 7, 1, alignment=Qt.AlignHCenter)
        gridBox_Methods.addWidget(self.button_SGDRegressor, 7, 2, alignment=Qt.AlignHCenter)

        gridBox_Methods.addWidget(label_SVR, 8, 0, alignment=Qt.AlignLeft)
        gridBox_Methods.addWidget(self.checkbox_SVR, 8, 1, alignment=Qt.AlignHCenter)
        gridBox_Methods.addWidget(self.button_SVR, 8, 2, alignment=Qt.AlignHCenter)

        gridBox_Methods.addWidget(label_LinearSVR, 9, 0, alignment=Qt.AlignLeft)
        gridBox_Methods.addWidget(self.checkbox_LinearSVR, 9, 1, alignment=Qt.AlignHCenter)
        gridBox_Methods.addWidget(self.button_LinearSVR, 9, 2, alignment=Qt.AlignHCenter)

        gridBox_Methods.addWidget(label_NearestNeighbor, 10, 0, alignment=Qt.AlignLeft)
        gridBox_Methods.addWidget(self.checkbox_NearestNeighbor, 10, 1, alignment=Qt.AlignHCenter)
        gridBox_Methods.addWidget(self.button_NearestNeighbor, 10, 2, alignment=Qt.AlignHCenter)

        gridBox_Methods.addWidget(label_KNeighborsRegressor, 11, 0, alignment=Qt.AlignLeft)
        gridBox_Methods.addWidget(self.checkbox_KNeighborsRegressor, 11, 1, alignment=Qt.AlignHCenter)
        gridBox_Methods.addWidget(self.button_KNeighborsRegressor, 11, 2, alignment=Qt.AlignHCenter)

        gridBox_Methods.addWidget(label_DecisionTreeRegressor, 12, 0, alignment=Qt.AlignLeft)
        gridBox_Methods.addWidget(self.checkbox_DecisionTreeRegressor, 12, 1, alignment=Qt.AlignHCenter)
        gridBox_Methods.addWidget(self.button_DecisionTreeRegressor, 12, 2, alignment=Qt.AlignHCenter)

        gridBox_Methods.addWidget(label_RandomForestRegressor, 13, 0, alignment=Qt.AlignLeft)
        gridBox_Methods.addWidget(self.checkbox_RandomForestRegressor, 13, 1, alignment=Qt.AlignHCenter)
        gridBox_Methods.addWidget(self.button_RandomForestRegressor, 13, 2, alignment=Qt.AlignHCenter)

        gridBox_Methods.addWidget(label_AdaBoostRegressor, 14, 0, alignment=Qt.AlignLeft)
        gridBox_Methods.addWidget(self.checkbox_AdaBoostRegressor, 14, 1, alignment=Qt.AlignHCenter)
        gridBox_Methods.addWidget(self.button_AdaBoostRegressor, 14, 2, alignment=Qt.AlignHCenter)

        gridBox_Methods.addWidget(label_GradientBoostingRegressor, 15, 0, alignment=Qt.AlignLeft)
        gridBox_Methods.addWidget(self.checkbox_GradientBoostingRegressor, 15, 1, alignment=Qt.AlignHCenter)
        gridBox_Methods.addWidget(self.button_GradientBoostingRegressor, 15, 2, alignment=Qt.AlignHCenter)

        return scrollAreaWidget

    def _setHorLayout(self):
        label_min_width = 200

        # Set Label
        label_Method = QLabel('Method')
        label_Method.setMinimumWidth(label_min_width)

        label_LinearRegression = QLabel(mlr.MLR_REG_LINEAR_REGRESSION)
        label_LinearRegression.setMinimumWidth(label_min_width)
        label_Ridge = QLabel(mlr.MLR_REG_RIDGE)
        label_Ridge.setMinimumWidth(label_min_width)
        label_BayesianRidge = QLabel(mlr.MLR_REG_BAYESIAN_RIDGE)
        label_BayesianRidge.setMinimumWidth(label_min_width)
        label_Lasso = QLabel(mlr.MLR_REG_LASSO)
        label_Lasso.setMinimumWidth(label_min_width)
        label_LassoLars = QLabel(mlr.MLR_REG_LASSO_LARS)
        label_LassoLars.setMinimumWidth(label_min_width)
        label_TweedieRegressor = QLabel(mlr.MLR_REG_TWEEDIE_REGRESSOR)
        label_TweedieRegressor.setMinimumWidth(label_min_width)
        label_SGDRegressor = QLabel(mlr.MLR_REG_SGD_REGRESSOR)
        label_SGDRegressor.setMinimumWidth(label_min_width)
        label_SVR = QLabel(mlr.MLR_REG_SVR)
        label_SVR.setMinimumWidth(label_min_width)
        label_LinearSVR = QLabel(mlr.MLR_REG_LINEAR_SVR)
        label_LinearSVR.setMinimumWidth(label_min_width)
        label_NearestNeighbor = QLabel(mlr.MLR_REG_NEAREST_NEIGHBORS)
        label_NearestNeighbor.setMinimumWidth(label_min_width)
        label_KNeighborsRegressor = QLabel(mlr.MLR_REG_K_NEIGHBORS_REGRESSOR)
        label_KNeighborsRegressor.setMinimumWidth(label_min_width)
        label_DecisionTreeRegressor = QLabel(mlr.MLR_REG_DECISION_TREE_REGRESSOR)
        label_DecisionTreeRegressor.setMinimumWidth(label_min_width)
        label_RandomForestRegressor = QLabel(mlr.MLR_REG_RANDOM_FOREST_REGRESSOR)
        label_RandomForestRegressor.setMinimumWidth(label_min_width)
        label_AdaBoostRegressor = QLabel(mlr.MLR_REG_ADA_BOOST_REGRESSOR)
        label_AdaBoostRegressor.setMinimumWidth(label_min_width)
        label_GradientBoostingRegressor = QLabel(mlr.MLR_REG_GRADIENT_BOOSTING_REGRESSOR)
        label_GradientBoostingRegressor.setMinimumWidth(label_min_width)

        # Set hboxes
        hbox_header = QHBoxLayout()
        hbox_LinearRegression = QHBoxLayout()
        hbox_Ridge = QHBoxLayout()
        hbox_BayesianRidge = QHBoxLayout()
        hbox_Lasso = QHBoxLayout()
        hbox_LassoLars = QHBoxLayout()
        hbox_TweedieRegressor = QHBoxLayout()
        hbox_SGDRegressor = QHBoxLayout()
        hbox_SVR = QHBoxLayout()
        hbox_LinearSVR = QHBoxLayout()
        hbox_NearestNeighbor = QHBoxLayout()
        hbox_KNeighborsRegressor = QHBoxLayout()
        hbox_DecisionTreeRegressor = QHBoxLayout()
        hbox_RandomForestRegressor = QHBoxLayout()
        hbox_AdaBoostRegressor = QHBoxLayout()
        hbox_GradientBoostingRegressor = QHBoxLayout()

        # Add to hboxes
        hbox_header.addWidget(label_Method)
        hbox_header.addWidget(QLabel("State"))
        hbox_header.addWidget(QLabel("Options"))

        hbox_LinearRegression.addWidget(label_LinearRegression)
        hbox_LinearRegression.addWidget(self.checkbox_LinearRegression)
        hbox_LinearRegression.addSpacerItem(QSpacerItem(INT_MAX_STRETCH, 0))

        hbox_Ridge.addWidget(label_Ridge)
        hbox_Ridge.addWidget(self.checkbox_Ridge)
        hbox_Ridge.addWidget(self.button_Ridge)
        hbox_Ridge.addSpacerItem(QSpacerItem(INT_MAX_STRETCH, 0))

        hbox_BayesianRidge.addWidget(label_BayesianRidge)
        hbox_BayesianRidge.addWidget(self.checkbox_BayesianRidge)
        hbox_BayesianRidge.addWidget(self.button_BayesianRidge)
        hbox_BayesianRidge.addSpacerItem(QSpacerItem(INT_MAX_STRETCH, 0))

        hbox_Lasso.addWidget(label_Lasso)
        hbox_Lasso.addWidget(self.checkbox_Lasso)
        hbox_Lasso.addWidget(self.button_Lasso)
        hbox_Lasso.addSpacerItem(QSpacerItem(INT_MAX_STRETCH, 0))

        hbox_LassoLars.addWidget(label_LassoLars)
        hbox_LassoLars.addWidget(self.checkbox_LassoLars)
        hbox_LassoLars.addWidget(self.button_LassoLars)
        hbox_LassoLars.addSpacerItem(QSpacerItem(INT_MAX_STRETCH, 0))

        hbox_TweedieRegressor.addWidget(label_TweedieRegressor)
        hbox_TweedieRegressor.addWidget(self.checkbox_TweedieRegressor)
        hbox_TweedieRegressor.addWidget(self.button_TweedieRegressor)
        hbox_TweedieRegressor.addSpacerItem(QSpacerItem(INT_MAX_STRETCH, 0))

        hbox_SGDRegressor.addWidget(label_SGDRegressor)
        hbox_SGDRegressor.addWidget(self.checkbox_SGDRegressor)
        hbox_SGDRegressor.addWidget(self.button_SGDRegressor)
        hbox_SGDRegressor.addSpacerItem(QSpacerItem(INT_MAX_STRETCH, 0))

        hbox_SVR.addWidget(label_SVR)
        hbox_SVR.addWidget(self.checkbox_SVR)
        hbox_SVR.addWidget(self.button_SVR)
        hbox_SVR.addSpacerItem(QSpacerItem(INT_MAX_STRETCH, 0))

        hbox_LinearSVR.addWidget(label_LinearSVR)
        hbox_LinearSVR.addWidget(self.checkbox_LinearSVR)
        hbox_LinearSVR.addWidget(self.button_LinearSVR)
        hbox_LinearSVR.addSpacerItem(QSpacerItem(INT_MAX_STRETCH, 0))

        hbox_NearestNeighbor.addWidget(label_NearestNeighbor)
        hbox_NearestNeighbor.addWidget(self.checkbox_NearestNeighbor)
        hbox_NearestNeighbor.addWidget(self.button_NearestNeighbor)
        hbox_NearestNeighbor.addSpacerItem(QSpacerItem(INT_MAX_STRETCH, 0))

        hbox_KNeighborsRegressor.addWidget(label_KNeighborsRegressor)
        hbox_KNeighborsRegressor.addWidget(self.checkbox_KNeighborsRegressor)
        hbox_KNeighborsRegressor.addWidget(self.button_KNeighborsRegressor)
        hbox_KNeighborsRegressor.addSpacerItem(QSpacerItem(INT_MAX_STRETCH, 0))

        hbox_DecisionTreeRegressor.addWidget(label_DecisionTreeRegressor)
        hbox_DecisionTreeRegressor.addWidget(self.checkbox_DecisionTreeRegressor)
        hbox_DecisionTreeRegressor.addWidget(self.button_DecisionTreeRegressor)
        hbox_DecisionTreeRegressor.addSpacerItem(QSpacerItem(INT_MAX_STRETCH, 0))

        hbox_RandomForestRegressor.addWidget(label_RandomForestRegressor)
        hbox_RandomForestRegressor.addWidget(self.checkbox_RandomForestRegressor)
        hbox_RandomForestRegressor.addWidget(self.button_RandomForestRegressor)
        hbox_RandomForestRegressor.addSpacerItem(QSpacerItem(INT_MAX_STRETCH, 0))

        hbox_AdaBoostRegressor.addWidget(label_AdaBoostRegressor)
        hbox_AdaBoostRegressor.addWidget(self.checkbox_AdaBoostRegressor)
        hbox_AdaBoostRegressor.addWidget(self.button_AdaBoostRegressor)
        hbox_AdaBoostRegressor.addSpacerItem(QSpacerItem(INT_MAX_STRETCH, 0))

        hbox_GradientBoostingRegressor.addWidget(label_GradientBoostingRegressor)
        hbox_GradientBoostingRegressor.addWidget(self.checkbox_GradientBoostingRegressor)
        hbox_GradientBoostingRegressor.addWidget(self.button_GradientBoostingRegressor)
        hbox_GradientBoostingRegressor.addSpacerItem(QSpacerItem(INT_MAX_STRETCH, 0))

        vbox_Methods = QVBoxLayout()
        vbox_Methods.addLayout(hbox_header)
        vbox_Methods.addLayout(hbox_LinearRegression)
        vbox_Methods.addLayout(hbox_Ridge)
        vbox_Methods.addLayout(hbox_BayesianRidge)
        vbox_Methods.addLayout(hbox_Lasso)
        vbox_Methods.addLayout(hbox_LassoLars)
        vbox_Methods.addLayout(hbox_TweedieRegressor)
        vbox_Methods.addLayout(hbox_SGDRegressor)
        vbox_Methods.addLayout(hbox_SVR)
        vbox_Methods.addLayout(hbox_LinearSVR)
        vbox_Methods.addLayout(hbox_NearestNeighbor)
        vbox_Methods.addLayout(hbox_KNeighborsRegressor)
        vbox_Methods.addLayout(hbox_DecisionTreeRegressor)
        vbox_Methods.addLayout(hbox_RandomForestRegressor)
        vbox_Methods.addLayout(hbox_AdaBoostRegressor)
        vbox_Methods.addLayout(hbox_GradientBoostingRegressor)

        return vbox_Methods

    # ----------------------------- #
    # ----- Setters / Getters ----- #
    # ----------------------------- #
    def setCheckState_LinearRegression(self, state: bool):
        self.checkbox_LinearRegression.setCheckState(state)

    def getCheckState_LinearRegression(self):
        return self.checkbox_LinearRegression.isChecked()

    def setCheckState_Ridge(self, state: bool):
        self.checkbox_Ridge.setCheckState(state)

    def getCheckState_Ridge(self):
        return self.checkbox_Ridge.isChecked()

    def setCheckState_BayesianRidge(self, state: bool):
        self.checkbox_BayesianRidge.setCheckState(state)

    def getCheckState_BayesianRidge(self):
        return self.checkbox_BayesianRidge.isChecked()

    def setCheckState_Lasso(self, state: bool):
        self.checkbox_Lasso.setCheckState(state)

    def getCheckState_Lasso(self):
        return self.checkbox_Lasso.isChecked()

    def setCheckState_LassoLars(self, state: bool):
        self.checkbox_LassoLars.setCheckState(state)

    def getCheckState_LassoLars(self):
        return self.checkbox_LassoLars.isChecked()

    def setCheckState_TweedieRegressor(self, state: bool):
        self.checkbox_TweedieRegressor.setCheckState(state)

    def getCheckState_TweedieRegressor(self):
        return self.checkbox_TweedieRegressor.isChecked()

    def setCheckState_SGDRegressor(self, state: bool):
        self.checkbox_SGDRegressor.setCheckState(state)

    def getCheckState_SGDRegressor(self):
        return self.checkbox_SGDRegressor.isChecked()

    def setCheckState_SVR(self, state: bool):
        self.checkbox_SVR.setCheckState(state)

    def getCheckState_SVR(self):
        return self.checkbox_SVR.isChecked()

    def setCheckState_LinearSVR(self, state: bool):
        self.checkbox_LinearSVR.setCheckState(state)

    def getCheckState_LinearSVR(self):
        return self.checkbox_LinearSVR.isChecked()

    def setCheckState_NearestNeighbor(self, state: bool):
        self.checkbox_NearestNeighbor.setCheckState(state)

    def getCheckState_NearestNeighbor(self):
        return self.checkbox_NearestNeighbor.isChecked()

    def setCheckState_KNeighborsRegressor(self, state: bool):
        self.checkbox_KNeighborsRegressor.setCheckState(state)

    def getCheckState_KNeighborsRegressor(self):
        return self.checkbox_KNeighborsRegressor.isChecked()

    def setCheckState_DecisionTreeRegressor(self, state: bool):
        self.checkbox_DecisionTreeRegressor.setCheckState(state)

    def getCheckState_DecisionTreeRegressor(self):
        return self.checkbox_DecisionTreeRegressor.isChecked()

    def setCheckState_RandomForestRegressor(self, state: bool):
        self.checkbox_RandomForestRegressor.setCheckState(state)

    def getCheckState_RandomForestRegressor(self):
        return self.checkbox_RandomForestRegressor.isChecked()

    def setCheckState_AdaBoostRegressor(self, state: bool):
        self.checkbox_RandomForestRegressor.setCheckState(state)

    def getCheckState_AdaBoostRegressor(self):
        return self.checkbox_RandomForestRegressor.isChecked()

    def setCheckState_GradientBoostingRegressor(self, state: bool):
        self.checkbox_GradientBoostingRegressor.setCheckState(state)

    def getCheckState_GradientBoostingRegressor(self):
        return self.checkbox_GradientBoostingRegressor.isChecked()


# *********** Machine Learning Settings --> Deep Regression Methods *********** #
class WidgetTabMachineLearningSettingsDeepRegressionMethods(QWidget):
    def __init__(self):
        super().__init__()

        self.setStyleSheet(setStyle_())  # set the tab style

        # ---------------------- #
        # ----- Set Window ----- #
        # ---------------------- #
        self.vbox_main_layout = QVBoxLayout(self)  # Create the main vbox

        # ----------------------- #
        # ----- PushButtons ----- #
        # ----------------------- #
        _icon = QIcon(QPixmap(ICON_OPTION_SETTINGS))

        self.button_Covid_Convolutional_1D_LongShortTermMemory = QPushButton()
        self.button_Covid_Convolutional_1D_LongShortTermMemory.setIcon(_icon)

        self.button_Covid_LongShortTermMemoryNeuralNetwork = QPushButton()
        self.button_Covid_LongShortTermMemoryNeuralNetwork.setIcon(_icon)

        self.button_Covid_LongShortTermMemoryNeuralNetwork_Simple = QPushButton()
        self.button_Covid_LongShortTermMemoryNeuralNetwork_Simple.setIcon(_icon)

        self.button_Covid_SimpleRecurrentNeuralNetwork = QPushButton()
        self.button_Covid_SimpleRecurrentNeuralNetwork.setIcon(_icon)

        self.button_Covid_Convolutional_1D_LongShortTermMemory.setEnabled(False)
        self.button_Covid_LongShortTermMemoryNeuralNetwork.setEnabled(False)
        self.button_Covid_LongShortTermMemoryNeuralNetwork_Simple.setEnabled(False)
        self.button_Covid_SimpleRecurrentNeuralNetwork.setEnabled(False)

        # ---------------------- #
        # ----- CheckBoxes ----- #
        # ---------------------- #
        self.checkbox_Covid_Convolutional_1D_LongShortTermMemory = QCheckBox()
        self.checkbox_Covid_LongShortTermMemoryNeuralNetwork = QCheckBox()
        self.checkbox_Covid_LongShortTermMemoryNeuralNetwork_Simple = QCheckBox()
        self.checkbox_Covid_SimpleRecurrentNeuralNetwork = QCheckBox()

        # ---------------------- #
        # ----- ScrollArea ----- #
        # ---------------------- #
        self.scrollArea_regMethods = QScrollArea()

    # --------------------------- #
    # ----- Reuse Functions ----- #
    # --------------------------- #
    def setWidget(self):
        """
            A function to create the widget components into the main QWidget
            :return: Nothing
        """
        self.scrollArea_regMethods.setWidgetResizable(True)
        self.scrollArea_regMethods.setWidget(self._setGridLayout())
        self.vbox_main_layout.addWidget(self.scrollArea_regMethods)

    def _setGridLayout(self):
        # label_min_width = 200

        # Set Label
        label_Method = QLabel('<b><u>Method<\\u><\\b>')
        # label_Method.setMaximumHeight(30)

        label_State = QLabel('<b><u>State<\\u><\\b>')
        # label_State.setMaximumHeight(30)

        label_Options = QLabel('<b><u>Options<\\u><\\b>')
        # label_Options.setMaximumHeight(30)

        label_Covid_Convolutional_1D_LongShortTermMemory = QLabel(mlr.MLR_REG_COVID_CONV1D_LSTM)
        label_Covid_LongShortTermMemoryNeuralNetwork = QLabel(mlr.MLR_REG_COVID_LSTM)
        label_Covid_LongShortTermMemoryNeuralNetwork_Simple = QLabel(mlr.MLR_REG_COVID_GRU)
        label_Covid_SimpleRecurrentNeuralNetwork = QLabel(mlr.MLR_REG_COVID_SIMPLE_RNN)

        # Set layout
        scrollAreaWidget = QWidget()
        scrollAreaWidget.setMaximumWidth(840)
        scrollAreaWidget.setMaximumHeight(512)
        gridBox_Methods = QGridLayout(scrollAreaWidget)

        tmpRowIndex = 0  # Create this index for dynamically change the rows when code changed
        gridBox_Methods.addWidget(label_Method, tmpRowIndex, 0,
                                  alignment=Qt.AlignLeft)
        gridBox_Methods.addWidget(label_State, tmpRowIndex, 1,
                                  alignment=Qt.AlignCenter)
        gridBox_Methods.addWidget(label_Options, tmpRowIndex, 2,
                                  alignment=Qt.AlignCenter)

        tmpRowIndex += 1
        gridBox_Methods.addWidget(label_Covid_Convolutional_1D_LongShortTermMemory, tmpRowIndex, 0,
                                  alignment=Qt.AlignLeft)
        gridBox_Methods.addWidget(self.checkbox_Covid_Convolutional_1D_LongShortTermMemory, tmpRowIndex, 1,
                                  alignment=Qt.AlignHCenter)
        gridBox_Methods.addWidget(self.button_Covid_Convolutional_1D_LongShortTermMemory, tmpRowIndex, 2,
                                  alignment=Qt.AlignHCenter)

        tmpRowIndex += 1
        gridBox_Methods.addWidget(label_Covid_LongShortTermMemoryNeuralNetwork, tmpRowIndex, 0,
                                  alignment=Qt.AlignLeft)
        gridBox_Methods.addWidget(self.checkbox_Covid_LongShortTermMemoryNeuralNetwork, tmpRowIndex, 1,
                                  alignment=Qt.AlignHCenter)
        gridBox_Methods.addWidget(self.button_Covid_LongShortTermMemoryNeuralNetwork, tmpRowIndex, 2,
                                  alignment=Qt.AlignHCenter)

        tmpRowIndex += 1
        gridBox_Methods.addWidget(label_Covid_LongShortTermMemoryNeuralNetwork_Simple, tmpRowIndex, 0,
                                  alignment=Qt.AlignLeft)
        gridBox_Methods.addWidget(self.checkbox_Covid_LongShortTermMemoryNeuralNetwork_Simple, tmpRowIndex, 1,
                                  alignment=Qt.AlignHCenter)
        gridBox_Methods.addWidget(self.button_Covid_LongShortTermMemoryNeuralNetwork_Simple, tmpRowIndex, 2,
                                  alignment=Qt.AlignHCenter)

        tmpRowIndex += 1
        gridBox_Methods.addWidget(label_Covid_SimpleRecurrentNeuralNetwork, tmpRowIndex, 0,
                                  alignment=Qt.AlignLeft)
        gridBox_Methods.addWidget(self.checkbox_Covid_SimpleRecurrentNeuralNetwork, tmpRowIndex, 1,
                                  alignment=Qt.AlignHCenter)
        gridBox_Methods.addWidget(self.button_Covid_SimpleRecurrentNeuralNetwork, tmpRowIndex, 2,
                                  alignment=Qt.AlignHCenter)

        tmpRowIndex += 1
        gridBox_Methods.addItem(QSpacerItem(0, 10), tmpRowIndex, 0, tmpRowIndex, 2)

        return scrollAreaWidget

    # ----------------------------- #
    # ----- Setters / Getters ----- #
    # ----------------------------- #
    def setCheckState_Covid_Convolutional_1D_LongShortTermMemory(self, state: bool):
        self.checkbox_Covid_Convolutional_1D_LongShortTermMemory.setCheckState(state)

    def getCheckState_Covid_Convolutional_1D_LongShortTermMemory(self):
        return self.checkbox_Covid_Convolutional_1D_LongShortTermMemory.isChecked()

    def setCheckState_Covid_LongShortTermMemoryNeuralNetwork(self, state: bool):
        self.checkbox_Covid_LongShortTermMemoryNeuralNetwork.setCheckState(state)

    def getCheckState_Covid_LongShortTermMemoryNeuralNetwork(self):
        return self.checkbox_Covid_LongShortTermMemoryNeuralNetwork.isChecked()

    def setCheckState_Covid_LongShortTermMemoryNeuralNetwork_Simple(self, state: bool):
        self.checkbox_Covid_LongShortTermMemoryNeuralNetwork_Simple.setCheckState(state)

    def getCheckState_Covid_LongShortTermMemoryNeuralNetwork_Simple(self):
        return self.checkbox_Covid_LongShortTermMemoryNeuralNetwork_Simple.isChecked()

    def setCheckState_Covid_SimpleRecurrentNeuralNetwork(self, state: bool):
        self.checkbox_Covid_SimpleRecurrentNeuralNetwork.setCheckState(state)

    def getCheckState_Covid_SimpleRecurrentNeuralNetwork(self):
        return self.checkbox_Covid_SimpleRecurrentNeuralNetwork.isChecked()


# *********** Machine Learning Settings --> SignalCompare Methods *********** #
class WidgetTabMachineLearningSettingsSignalCompareMethods(QWidget):
    def __init__(self):
        super().__init__()

        # ---------------------- #
        # ----- Set Window ----- #
        # ---------------------- #
        self.vbox_main_layout = QVBoxLayout(self)  # Create the main vbox

        # ---------------------- #
        # ----- CheckBoxes ----- #
        # ---------------------- #
        self.checkbox_PearsonCorr = QCheckBox()
        self.checkbox_TimeLaggedCrossCorrelation = QCheckBox()
        self.checkbox_TimeLaggedCrossCorrelationNoSplits = QCheckBox()
        self.checkbox_RollingWindowTimeLaggedCrossCorrelation = QCheckBox()
        self.checkbox_DynamicTimeWarping = QCheckBox()

        self.checkbox_TimeLaggedCrossCorrelationNoSplits.setEnabled(False)
        self.checkbox_RollingWindowTimeLaggedCrossCorrelation.setEnabled(False)

        # ---------------------- #
        # ----- ScrollArea ----- #
        # ---------------------- #
        self.scrollArea_compMethods = QScrollArea()

    # --------------------------- #
    # ----- Reuse Functions ----- #
    # --------------------------- #
    def setWidget(self):
        """
            A function to create the widget components into the main QWidget
            :return: Nothing
        """
        self.scrollArea_compMethods.setWidgetResizable(True)
        self.scrollArea_compMethods.setWidget(self._setGridLayout())
        self.vbox_main_layout.addWidget(self.scrollArea_compMethods)

    def _setGridLayout(self):
        # label_min_width = 200

        # Set Label
        label_Method = QLabel('<b><u>Method<\\u><\\b>')
        # label_Method.setMaximumHeight(30)

        label_State = QLabel('<b><u>State<\\u><\\b>')
        # label_State.setMaximumHeight(30)

        label_PearsonCorr = QLabel(signComp.SC_PEARSON_CORRELATION)
        label_TimeLaggedCrossCorrelation = QLabel(signComp.SC_TIME_LAGGED_CROSS_CORRELATION)
        label_TimeLaggedCrossCorrelationNoSplits = QLabel(signComp.SC_TIME_LAGGED_CROSS_CORRELATION_NO_SPLITS)
        label_RollingWindowTimeLaggedCrossCorrelation = QLabel(signComp.SC_ROLLING_WINDOW_TIME_LAGGED_CROSS_CORRELATION)
        label_DynamicTimeWarping = QLabel(signComp.SC_DYNAMIC_TIME_WARPING)

        # Set layout
        scrollAreaWidget = QWidget()
        scrollAreaWidget.setMaximumWidth(480)
        scrollAreaWidget.setMaximumHeight(200)
        gridBox_Methods = QGridLayout(scrollAreaWidget)

        gridBox_Methods.addWidget(label_Method, 0, 0, alignment=Qt.AlignLeft)
        gridBox_Methods.addWidget(label_State, 0, 1, alignment=Qt.AlignCenter)

        gridBox_Methods.addWidget(label_PearsonCorr, 1, 0, alignment=Qt.AlignLeft)
        gridBox_Methods.addWidget(self.checkbox_PearsonCorr, 1, 1, alignment=Qt.AlignHCenter)

        gridBox_Methods.addWidget(label_TimeLaggedCrossCorrelation, 2, 0, alignment=Qt.AlignLeft)
        gridBox_Methods.addWidget(self.checkbox_TimeLaggedCrossCorrelation, 2, 1, alignment=Qt.AlignHCenter)

        gridBox_Methods.addWidget(label_TimeLaggedCrossCorrelationNoSplits, 3, 0, alignment=Qt.AlignLeft)
        gridBox_Methods.addWidget(self.checkbox_TimeLaggedCrossCorrelationNoSplits, 3, 1, alignment=Qt.AlignHCenter)

        gridBox_Methods.addWidget(label_RollingWindowTimeLaggedCrossCorrelation, 4, 0, alignment=Qt.AlignLeft)
        gridBox_Methods.addWidget(self.checkbox_RollingWindowTimeLaggedCrossCorrelation, 4, 1,
                                  alignment=Qt.AlignHCenter)

        gridBox_Methods.addWidget(label_DynamicTimeWarping, 5, 0, alignment=Qt.AlignLeft)
        gridBox_Methods.addWidget(self.checkbox_DynamicTimeWarping, 5, 1, alignment=Qt.AlignHCenter)

        return scrollAreaWidget

    # ------------------------------ #
    # ----- Change Check State ----- #
    # ------------------------------ #
    def setCheckState_PearsonCorr(self, state: bool):
        self.checkbox_PearsonCorr.setCheckState(state)

    def setCheckState_TimeLaggedCrossCorrelation(self, state: bool):
        self.checkbox_TimeLaggedCrossCorrelation.setCheckState(state)

    def setCheckState_TimeLaggedCrossCorrelationNoSplits(self, state: bool):
        self.checkbox_TimeLaggedCrossCorrelationNoSplits.setCheckState(state)

    def setCheckState_RollingWindowTimeLaggedCrossCorrelation(self, state: bool):
        self.checkbox_RollingWindowTimeLaggedCrossCorrelation.setCheckState(state)

    def setCheckState_DynamicTimeWarping(self, state: bool):
        self.checkbox_DynamicTimeWarping.setCheckState(state)

    def getCheckState_PearsonCorr(self):
        return self.checkbox_PearsonCorr.checkState()

    def getCheckState_TimeLaggedCrossCorrelation(self):
        return self.checkbox_TimeLaggedCrossCorrelation.checkState()

    def getCheckState_TimeLaggedCrossCorrelationNoSplits(self):
        return self.checkbox_TimeLaggedCrossCorrelationNoSplits.checkState()

    def getCheckState_RollingWindowTimeLaggedCrossCorrelation(self):
        return self.checkbox_RollingWindowTimeLaggedCrossCorrelation.checkState()

    def getCheckState_DynamicTimeWarping(self):
        return self.checkbox_DynamicTimeWarping.checkState()


# *                                 * #
# *********************************** #

# ************************************************ #
# *********** Machine Learning Methods *********** #
# ************************************************ #
# *                                              * #


class WidgetRidgeML(QWidget):
    def __init__(self, w=512, h=512, minW=256, minH=256, maxW=None, maxH=None,
                 winTitle='My Window', iconPath=None):
        super().__init__()

        self.setStyleSheet(setStyle_())  # set the tab style

        self.iconPath = iconPath
        # ---------------------- #
        # ----- Set Window ----- #
        # ---------------------- #
        self.setWindowTitle(winTitle)  # Set Window Title
        self.setWindowIcon(QIcon(self.iconPath))  # Set Window Icon
        self.setGeometry(INT_SCREEN_WIDTH / 4, INT_SCREEN_HEIGHT / 4, w, h)  # Set Window Geometry
        self.setMinimumWidth(minW)  # Set Window Minimum Width
        self.setMinimumHeight(minH)  # Set Window Minimum Height
        if maxW is not None:
            self.setMaximumWidth(maxW)  # Set Window Maximum Width
        if maxH is not None:
            self.setMaximumHeight(maxH)  # Set Window Maximum Width

        self.vbox_main_layout = QVBoxLayout(self)  # Create the main vbox

        # -------------------------- #
        # ----- QDoubleSpinBox ----- #
        # -------------------------- #
        self.doubleSpinBox_AlphaMin = QDoubleSpinBox()
        self.doubleSpinBox_AlphaMax = QDoubleSpinBox()
        self.doubleSpinBox_AlphaStep = QDoubleSpinBox()

        self.doubleSpinBox_AlphaMin.setMinimum(0.01)
        self.doubleSpinBox_AlphaMax.setMinimum(0.01)
        self.doubleSpinBox_AlphaStep.setMinimum(0.01)

        _stepValue = 0.05
        self.doubleSpinBox_AlphaMin.setSingleStep(_stepValue)
        self.doubleSpinBox_AlphaMax.setSingleStep(_stepValue)
        self.doubleSpinBox_AlphaStep.setSingleStep(_stepValue)

        # ----------------------- #
        # ----- QListWidget ----- #
        # ----------------------- #
        self.listWidget_TolOptionsList = QListWidget()
        self.listWidget_TolOptionsList.setSelectionMode(QListWidget.ExtendedSelection)  # Set Extended Selection
        self.listWidget_TolSelectedList = QListWidget()

        self.listWidget_SolverOptionsList = QListWidget()
        self.listWidget_SolverOptionsList.setSelectionMode(QListWidget.ExtendedSelection)  # Set Extended Selection
        self.listWidget_SolverSelectedList = QListWidget()

        # ----------------------- #
        # ----- QPushButton ----- #
        # ----------------------- #
        _iconAdd = QIcon(QPixmap(ICON_ADD_RIGHT_LIST))
        _iconDel = QIcon(QPixmap(ICON_DELETE_FROM_LIST))

        self.button_TolAdd = QPushButton()
        self.button_TolAdd.setIcon(_iconAdd)
        self.button_TolRemove = QPushButton()
        self.button_TolRemove.setIcon(_iconDel)

        self.button_SolverAdd = QPushButton()
        self.button_SolverAdd.setIcon(_iconAdd)
        self.button_SolverRemove = QPushButton()
        self.button_SolverRemove.setIcon(_iconDel)

        self.button_RestoreDefault = QPushButton("Restore Default")
        self.button_RestoreDefault.setMinimumWidth(150)  # Set Minimum Width
        self.button_RestoreDefault.setMinimumHeight(30)  # Set Minimum Height

    def setWidget(self):
        """
            A function to create the widget components into the main QWidget
            :return: Nothing
        """
        self.vbox_main_layout.addLayout(self._setHorLayout())

    def _setHorLayout(self):
        label_AlphaMin = QLabel("Alpha (min):")
        label_AlphaMax = QLabel("Alpha (max):")
        label_AlphaStep = QLabel("Alpha (step):")

        labelTolOptionList = QLabel("Tolerance Options:")
        labelTolSelectedList = QLabel("Tolerance Selected:")

        labelSolverOptionList = QLabel("Solver Options:")
        labelSolverSelectedList = QLabel("Solver Selected:")

        # Set hbox
        hbox_AlphaMin = QHBoxLayout()
        hbox_AlphaMax = QHBoxLayout()
        hbox_AlphaStep = QHBoxLayout()

        hbox_AlphaMin.addWidget(label_AlphaMin)
        hbox_AlphaMin.addWidget(self.doubleSpinBox_AlphaMin)
        hbox_AlphaMin.addSpacerItem(QSpacerItem(INT_MAX_STRETCH, 0))
        hbox_AlphaMax.addWidget(label_AlphaMax)
        hbox_AlphaMax.addWidget(self.doubleSpinBox_AlphaMax)
        hbox_AlphaMax.addSpacerItem(QSpacerItem(INT_MAX_STRETCH, 0))
        hbox_AlphaStep.addWidget(label_AlphaStep)
        hbox_AlphaStep.addWidget(self.doubleSpinBox_AlphaStep)
        hbox_AlphaStep.addSpacerItem(QSpacerItem(INT_MAX_STRETCH, 0))

        hbox_Restore = QHBoxLayout()
        hbox_Restore.addSpacerItem(QSpacerItem(INT_MAX_STRETCH, 0))
        hbox_Restore.addWidget(self.button_RestoreDefault)

        # Set vbox
        vbox_TolOptions = QVBoxLayout()
        vbox_TolSelected = QVBoxLayout()

        vbox_SolverOptions = QVBoxLayout()
        vbox_SolverSelected = QVBoxLayout()

        vbox_ButtonsTol = QVBoxLayout()
        vbox_ButtonsSolver = QVBoxLayout()

        vbox_TolOptions.addWidget(labelTolOptionList)
        vbox_TolOptions.addWidget(self.listWidget_TolOptionsList)
        vbox_TolSelected.addWidget(labelTolSelectedList)
        vbox_TolSelected.addWidget(self.listWidget_TolSelectedList)

        vbox_SolverOptions.addWidget(labelSolverOptionList)
        vbox_SolverOptions.addWidget(self.listWidget_SolverOptionsList)
        vbox_SolverSelected.addWidget(labelSolverSelectedList)
        vbox_SolverSelected.addWidget(self.listWidget_SolverSelectedList)

        vbox_ButtonsTol.addWidget(self.button_TolAdd)
        vbox_ButtonsTol.addWidget(self.button_TolRemove)
        vbox_ButtonsSolver.addWidget(self.button_SolverAdd)
        vbox_ButtonsSolver.addWidget(self.button_SolverRemove)

        hbox_Tol = QHBoxLayout()
        hbox_Tol.addLayout(vbox_TolOptions)
        hbox_Tol.addLayout(vbox_ButtonsTol)
        hbox_Tol.addLayout(vbox_TolSelected)

        hbox_Solver = QHBoxLayout()
        hbox_Solver.addLayout(vbox_SolverOptions)
        hbox_Solver.addLayout(vbox_ButtonsSolver)
        hbox_Solver.addLayout(vbox_SolverSelected)

        vbox_final = QVBoxLayout()
        vbox_final.addLayout(hbox_AlphaMin)
        vbox_final.addLayout(hbox_AlphaMax)
        vbox_final.addLayout(hbox_AlphaStep)
        vbox_final.addLayout(hbox_Tol)
        vbox_final.addLayout(hbox_Solver)
        vbox_final.addLayout(hbox_Restore)

        return vbox_final

    def setAlphaMin(self, value: float):
        self.doubleSpinBox_AlphaMin.setValue(value)

    def getAlphaMin(self):
        return self.doubleSpinBox_AlphaMin.value()

    def setAlphaMax(self, value: float):
        self.doubleSpinBox_AlphaMax.setValue(value)

    def getAlphaMax(self):
        return self.doubleSpinBox_AlphaMax.value()

    def setAlphaStep(self, value: float):
        self.doubleSpinBox_AlphaStep.setValue(value)

    def getAlphaStep(self):
        return self.doubleSpinBox_AlphaStep.value()

    def sortOptionList_Tol(self):
        self.listWidget_TolOptionsList.sortItems(Qt.AscendingOrder)

    def setOptionList_Tol(self, listValue: []):
        self.listWidget_TolOptionsList.clear()
        for _item_ in listValue:
            self.listWidget_TolOptionsList.addItem("{:.0e}".format(_item_))
        if self.listWidget_TolOptionsList.item(0):
            self.listWidget_TolOptionsList.setCurrentRow(0)
        self.sortOptionList_Tol()

    def sortOptionList_Solver(self):
        self.listWidget_SolverOptionsList.sortItems(Qt.AscendingOrder)

    def setOptionList_Solver(self, listValue: []):
        self.listWidget_SolverOptionsList.clear()
        for _item_ in listValue:
            self.listWidget_SolverOptionsList.addItem(str(_item_))
        if self.listWidget_SolverOptionsList.item(0):
            self.listWidget_SolverOptionsList.setCurrentRow(0)
        self.sortOptionList_Solver()

    def sortSelectedList_Tol(self):
        self.listWidget_TolSelectedList.sortItems(Qt.AscendingOrder)

    def setSelectedList_Tol(self, listValue: []):
        self.listWidget_TolSelectedList.clear()
        for _item_ in listValue:
            self.listWidget_TolSelectedList.addItem("{:.0e}".format(_item_))
        if self.listWidget_TolSelectedList.item(0):
            self.listWidget_TolSelectedList.setCurrentRow(0)
        self.sortSelectedList_Tol()

    def getSelectedList_Tol(self):
        valueList = []
        for _index_ in range(self.listWidget_TolSelectedList.count()):
            valueList.append(self.listWidget_TolSelectedList.item(_index_).text())
        return valueList

    def sortSelectedList_Solver(self):
        self.listWidget_SolverSelectedList.sortItems(Qt.AscendingOrder)

    def setSelectedList_Solver(self, listValue: []):
        self.listWidget_SolverSelectedList.clear()
        for _item_ in listValue:
            self.listWidget_SolverSelectedList.addItem(str(_item_))
        if self.listWidget_SolverSelectedList.item(0):
            self.listWidget_SolverSelectedList.setCurrentRow(0)
        self.sortSelectedList_Solver()

    def getSelectedList_Solver(self):
        valueList = []
        for _index_ in range(self.listWidget_SolverSelectedList.count()):
            valueList.append(self.listWidget_SolverSelectedList.item(_index_).text())
        return valueList

    def addToOptionList_Tol(self, value):
        self.listWidget_TolOptionsList.addItem(str(value))
        self.sortOptionList_Tol()

    def addToOptionList_Solver(self, value):
        self.listWidget_SolverOptionsList.addItem(str(value))
        self.sortOptionList_Solver()

    def addToSelectedList_Tol(self, value):
        self.listWidget_TolSelectedList.addItem(str(value))
        self.sortSelectedList_Tol()

    def addToSelectedList_Solver(self, value):
        self.listWidget_SolverSelectedList.addItem(str(value))
        self.sortSelectedList_Solver()

    def getListOptionSelectedItems_Tol(self):
        return [_item_.text() for _item_ in self.listWidget_TolOptionsList.selectedItems()]

    def getListOptionSelectedItems_Solver(self):
        return [_item_.text() for _item_ in self.listWidget_SolverOptionsList.selectedItems()]

    def checkIfItemAlreadyInList_Tol(self, item):
        itemFound = False
        for _index_ in range(self.listWidget_TolSelectedList.count()):
            if item == self.listWidget_TolSelectedList.item(_index_).text():
                itemFound = True
                break
        return itemFound

    def checkIfItemAlreadyInList_Solver(self, item):
        itemFound = False
        for _index_ in range(self.listWidget_SolverSelectedList.count()):
            if item == self.listWidget_SolverSelectedList.item(_index_).text():
                itemFound = True
                break
        return itemFound

    def removeItemFromList_TolSelected(self):
        for _item_ in self.listWidget_TolSelectedList.selectedItems():
            if self.checkIfItemAlreadyInList_Tol(_item_.text()) and self.listWidget_TolSelectedList.count() > 1:
                self.listWidget_TolSelectedList.takeItem(self.listWidget_TolSelectedList.row(_item_))

    def removeItemFromList_SolverSelected(self):
        for _item_ in self.listWidget_SolverSelectedList.selectedItems():
            if self.checkIfItemAlreadyInList_Solver(_item_.text()) and self.listWidget_SolverSelectedList.count() > 1:
                self.listWidget_SolverSelectedList.takeItem(self.listWidget_SolverSelectedList.row(_item_))


class WidgetSVRML(QWidget):
    def __init__(self, w=512, h=512, minW=256, minH=256, maxW=None, maxH=None,
                 winTitle='My Window', iconPath=None):
        super().__init__()

        self.setStyleSheet(setStyle_())  # set the tab style

        self.iconPath = iconPath
        # ---------------------- #
        # ----- Set Window ----- #
        # ---------------------- #
        self.setWindowTitle(winTitle)  # Set Window Title
        self.setWindowIcon(QIcon(self.iconPath))  # Set Window Icon
        self.setGeometry(INT_SCREEN_WIDTH / 4, INT_SCREEN_HEIGHT / 4, w, h)  # Set Window Geometry
        self.setMinimumWidth(minW)  # Set Window Minimum Width
        self.setMinimumHeight(minH)  # Set Window Minimum Height
        if maxW is not None:
            self.setMaximumWidth(maxW)  # Set Window Maximum Width
        if maxH is not None:
            self.setMaximumHeight(maxH)  # Set Window Maximum Width

        self.vbox_main_layout = QVBoxLayout(self)  # Create the main vbox

        # ----------------------- #
        # ----- QListWidget ----- #
        # ----------------------- #
        self.listWidget_KernelOptionsList = QListWidget()
        self.listWidget_KernelOptionsList.setSelectionMode(QListWidget.ExtendedSelection)  # Set Extended Selection
        self.listWidget_KernelSelectedList = QListWidget()

        self.listWidget_GammaOptionsList = QListWidget()
        self.listWidget_GammaOptionsList.setSelectionMode(QListWidget.ExtendedSelection)  # Set Extended Selection
        self.listWidget_GammaSelectedList = QListWidget()

        self.listWidget_TolOptionsList = QListWidget()
        self.listWidget_TolOptionsList.setSelectionMode(QListWidget.ExtendedSelection)  # Set Extended Selection
        self.listWidget_TolSelectedList = QListWidget()

        # ----------------------- #
        # ----- QPushButton ----- #
        # ----------------------- #
        _iconAdd = QIcon(QPixmap(ICON_ADD_RIGHT_LIST))
        _iconDel = QIcon(QPixmap(ICON_DELETE_FROM_LIST))

        self.button_KernelAdd = QPushButton()
        self.button_KernelAdd.setIcon(_iconAdd)
        self.button_KernelRemove = QPushButton()
        self.button_KernelRemove.setIcon(_iconDel)

        self.button_GammaAdd = QPushButton()
        self.button_GammaAdd.setIcon(_iconAdd)
        self.button_GammaRemove = QPushButton()
        self.button_GammaRemove.setIcon(_iconDel)

        self.button_TolAdd = QPushButton()
        self.button_TolAdd.setIcon(_iconAdd)
        self.button_TolRemove = QPushButton()
        self.button_TolRemove.setIcon(_iconDel)

        self.button_RestoreDefault = QPushButton("Restore Default")
        self.button_RestoreDefault.setMinimumWidth(150)  # Set Minimum Width
        self.button_RestoreDefault.setMinimumHeight(30)  # Set Minimum Height

    def setWidget(self):
        """
            A function to create the widget components into the main QWidget
            :return: Nothing
        """
        self.vbox_main_layout.addLayout(self._setHorLayout())

    def _setHorLayout(self):
        labelKernelOptionList = QLabel("Kernel Options:")
        labelKernelSelectedList = QLabel("Kernel Selected:")

        labelGammaOptionList = QLabel("Gamma Options:")
        labelGammaSelectedList = QLabel("Gamma Selected:")

        labelTolOptionList = QLabel("Tolerance Options:")
        labelTolSelectedList = QLabel("Tolerance Selected:")

        # Set hbox
        hbox_Restore = QHBoxLayout()
        hbox_Restore.addSpacerItem(QSpacerItem(INT_MAX_STRETCH, 0))
        hbox_Restore.addWidget(self.button_RestoreDefault)

        # Set vbox
        vbox_KernelOptions = QVBoxLayout()  # Create Kernel Options layout - vbox
        vbox_KernelSelected = QVBoxLayout()  # Create Kernel Selected layout - vbox

        vbox_GammaOptions = QVBoxLayout()  # Create Gamma Options layout - vbox
        vbox_GammaSelected = QVBoxLayout()  # Create Gamma Selected layout - vbox

        vbox_TolOptions = QVBoxLayout()  # Create Tol Options layout - vbox
        vbox_TolSelected = QVBoxLayout()  # Create Tol Selected layout - vbox

        vbox_ButtonsKernel = QVBoxLayout()  # Create ButtonsKernel - vbox
        vbox_ButtonsGamma = QVBoxLayout()  # Create ButtonsGamma - vbox
        vbox_ButtonsTol = QVBoxLayout()  # Create ButtonsTol - vbox

        # Add widgets to Kernel (Option, Selected)-layouts
        vbox_KernelOptions.addWidget(labelKernelOptionList)
        vbox_KernelOptions.addWidget(self.listWidget_KernelOptionsList)
        vbox_KernelSelected.addWidget(labelKernelSelectedList)
        vbox_KernelSelected.addWidget(self.listWidget_KernelSelectedList)

        # Add widgets to Gamma (Option, Selected)-layouts
        vbox_GammaOptions.addWidget(labelGammaOptionList)
        vbox_GammaOptions.addWidget(self.listWidget_GammaOptionsList)
        vbox_GammaSelected.addWidget(labelGammaSelectedList)
        vbox_GammaSelected.addWidget(self.listWidget_GammaSelectedList)

        # Add widgets to Tol (Option, Selected)-layouts
        vbox_TolOptions.addWidget(labelTolOptionList)
        vbox_TolOptions.addWidget(self.listWidget_TolOptionsList)
        vbox_TolSelected.addWidget(labelTolSelectedList)
        vbox_TolSelected.addWidget(self.listWidget_TolSelectedList)

        # Add widgets to button (Kernel, Gamma, Tol)-layouts
        vbox_ButtonsKernel.addWidget(self.button_KernelAdd)
        vbox_ButtonsKernel.addWidget(self.button_KernelRemove)
        vbox_ButtonsGamma.addWidget(self.button_GammaAdd)
        vbox_ButtonsGamma.addWidget(self.button_GammaRemove)
        vbox_ButtonsTol.addWidget(self.button_TolAdd)
        vbox_ButtonsTol.addWidget(self.button_TolRemove)

        # Create hbox Kernel layout and add Kernel-layouts
        hbox_Kernel = QHBoxLayout()
        hbox_Kernel.addLayout(vbox_KernelOptions)
        hbox_Kernel.addLayout(vbox_ButtonsKernel)
        hbox_Kernel.addLayout(vbox_KernelSelected)

        # Create hbox Gamma layout and add Gamma-layouts
        hbox_Gamma = QHBoxLayout()
        hbox_Gamma.addLayout(vbox_GammaOptions)
        hbox_Gamma.addLayout(vbox_ButtonsGamma)
        hbox_Gamma.addLayout(vbox_GammaSelected)

        # Create hbox Tol layout and add Tol-layouts
        hbox_Tol = QHBoxLayout()
        hbox_Tol.addLayout(vbox_TolOptions)
        hbox_Tol.addLayout(vbox_ButtonsTol)
        hbox_Tol.addLayout(vbox_TolSelected)

        vbox_final = QVBoxLayout()
        vbox_final.addLayout(hbox_Kernel)
        vbox_final.addLayout(hbox_Gamma)
        vbox_final.addLayout(hbox_Tol)
        vbox_final.addLayout(hbox_Restore)

        return vbox_final

    def sortOptionList_Kernel(self):
        self.listWidget_KernelOptionsList.sortItems(Qt.AscendingOrder)

    def setOptionList_Kernel(self, listValue: []):
        self.listWidget_KernelOptionsList.clear()
        for _item_ in listValue:
            self.listWidget_KernelOptionsList.addItem(str(_item_))
        if self.listWidget_KernelOptionsList.item(0):
            self.listWidget_KernelOptionsList.setCurrentRow(0)
        self.sortOptionList_Kernel()

    def sortOptionList_Gamma(self):
        self.listWidget_GammaOptionsList.sortItems(Qt.AscendingOrder)

    def setOptionList_Gamma(self, listValue: []):
        self.listWidget_GammaOptionsList.clear()
        for _item_ in listValue:
            self.listWidget_GammaOptionsList.addItem(str(_item_))
        if self.listWidget_GammaOptionsList.item(0):
            self.listWidget_GammaOptionsList.setCurrentRow(0)
        self.sortOptionList_Gamma()

    def sortOptionList_Tol(self):
        self.listWidget_TolOptionsList.sortItems(Qt.AscendingOrder)

    def setOptionList_Tol(self, listValue: []):
        self.listWidget_TolOptionsList.clear()
        for _item_ in listValue:
            self.listWidget_TolOptionsList.addItem("{:.0e}".format(_item_))
        if self.listWidget_TolOptionsList.item(0):
            self.listWidget_TolOptionsList.setCurrentRow(0)
        self.sortOptionList_Tol()

    def sortSelectedList_Kernel(self):
        self.listWidget_KernelSelectedList.sortItems(Qt.AscendingOrder)

    def setSelectedList_Kernel(self, listValue: []):
        self.listWidget_KernelSelectedList.clear()
        for _item_ in listValue:
            self.listWidget_KernelSelectedList.addItem(str(_item_))
        if self.listWidget_KernelSelectedList.item(0):
            self.listWidget_KernelSelectedList.setCurrentRow(0)
        self.sortSelectedList_Kernel()

    def getSelectedList_Kernel(self):
        valueList = []
        for _index_ in range(self.listWidget_KernelSelectedList.count()):
            valueList.append(self.listWidget_KernelSelectedList.item(_index_).text())
        return valueList

    def sortSelectedList_Gamma(self):
        self.listWidget_GammaSelectedList.sortItems(Qt.AscendingOrder)

    def setSelectedList_Gamma(self, listValue: []):
        self.listWidget_GammaSelectedList.clear()
        for _item_ in listValue:
            self.listWidget_GammaSelectedList.addItem(str(_item_))
        if self.listWidget_GammaSelectedList.item(0):
            self.listWidget_GammaSelectedList.setCurrentRow(0)
        self.sortSelectedList_Gamma()

    def getSelectedList_Gamma(self):
        valueList = []
        for _index_ in range(self.listWidget_GammaSelectedList.count()):
            valueList.append(self.listWidget_GammaSelectedList.item(_index_).text())
        return valueList

    def sortSelectedList_Tol(self):
        self.listWidget_TolSelectedList.sortItems(Qt.AscendingOrder)

    def setSelectedList_Tol(self, listValue: []):
        self.listWidget_TolSelectedList.clear()
        for _item_ in listValue:
            self.listWidget_TolSelectedList.addItem("{:.0e}".format(_item_))
        if self.listWidget_TolSelectedList.item(0):
            self.listWidget_TolSelectedList.setCurrentRow(0)
        self.sortSelectedList_Tol()

    def getSelectedList_Tol(self):
        valueList = []
        for _index_ in range(self.listWidget_TolSelectedList.count()):
            valueList.append(self.listWidget_TolSelectedList.item(_index_).text())
        return valueList

    def addToOptionList_Kernel(self, value):
        self.listWidget_KernelOptionsList.addItem(str(value))
        self.sortOptionList_Kernel()

    def addToOptionList_Gamma(self, value):
        self.listWidget_GammaOptionsList.addItem(str(value))
        self.sortOptionList_Gamma()

    def addToOptionList_Tol(self, value):
        self.listWidget_TolOptionsList.addItem(str(value))
        self.sortOptionList_Tol()

    def addToSelectedList_Kernel(self, value):
        self.listWidget_KernelSelectedList.addItem(str(value))
        self.sortSelectedList_Kernel()

    def addToSelectedList_Gamma(self, value):
        self.listWidget_GammaSelectedList.addItem(str(value))
        self.sortSelectedList_Gamma()

    def addToSelectedList_Tol(self, value):
        self.listWidget_TolSelectedList.addItem(str(value))
        self.sortSelectedList_Tol()

    def getListOptionSelectedItems_Kernel(self):
        return [_item_.text() for _item_ in self.listWidget_KernelOptionsList.selectedItems()]

    def getListOptionSelectedItems_Gamma(self):
        return [_item_.text() for _item_ in self.listWidget_GammaOptionsList.selectedItems()]

    def getListOptionSelectedItems_Tol(self):
        return [_item_.text() for _item_ in self.listWidget_TolOptionsList.selectedItems()]

    def checkIfItemAlreadyInList_Kernel(self, item):
        itemFound = False
        for _index_ in range(self.listWidget_KernelSelectedList.count()):
            if item == self.listWidget_KernelSelectedList.item(_index_).text():
                itemFound = True
                break
        return itemFound

    def checkIfItemAlreadyInList_Gamma(self, item):
        itemFound = False
        for _index_ in range(self.listWidget_GammaSelectedList.count()):
            if item == self.listWidget_GammaSelectedList.item(_index_).text():
                itemFound = True
                break
        return itemFound

    def checkIfItemAlreadyInList_Tol(self, item):
        itemFound = False
        for _index_ in range(self.listWidget_TolSelectedList.count()):
            if item == self.listWidget_TolSelectedList.item(_index_).text():
                itemFound = True
                break
        return itemFound

    def removeItemFromList_KernelSelected(self):
        for _item_ in self.listWidget_KernelSelectedList.selectedItems():
            if self.checkIfItemAlreadyInList_Kernel(_item_.text()) and self.listWidget_KernelSelectedList.count() > 1:
                self.listWidget_KernelSelectedList.takeItem(self.listWidget_KernelSelectedList.row(_item_))

    def removeItemFromList_GammaSelected(self):
        for _item_ in self.listWidget_GammaSelectedList.selectedItems():
            if self.checkIfItemAlreadyInList_Gamma(_item_.text()) and self.listWidget_GammaSelectedList.count() > 1:
                self.listWidget_GammaSelectedList.takeItem(self.listWidget_GammaSelectedList.row(_item_))

    def removeItemFromList_TolSelected(self):
        for _item_ in self.listWidget_TolSelectedList.selectedItems():
            if self.checkIfItemAlreadyInList_Tol(_item_.text()) and self.listWidget_TolSelectedList.count() > 1:
                self.listWidget_TolSelectedList.takeItem(self.listWidget_TolSelectedList.row(_item_))


# *                                              * #
# ************************************************ #


# ******************************************************* #
# ********************   EXECUTION   ******************** #
# ******************************************************* #


def exec_app(w=512, h=512, minW=256, minH=256, maxW=512, maxH=512, winTitle='My Window', iconPath=''):
    myApp = QApplication(sys.argv)  # Set Up Application
    widgetWin = WidgetMachineLearningRegressionWidget(w=w, h=h, minW=minW, minH=minH, maxW=maxW, maxH=maxH,
                                                      winTitle=winTitle, iconPath=iconPath)  # Create MainWindow
    widgetWin.show()  # Show Window
    myApp.exec_()  # Execute Application
    sys.exit(0)  # Exit Application


if __name__ == "__main__":
    exec_app(w=1024, h=512, minW=512, minH=256, maxW=512, maxH=512,
             winTitle='WidgetTemplate', iconPath=PROJECT_FOLDER + '/icon/crabsMLearning_32x32.png')
