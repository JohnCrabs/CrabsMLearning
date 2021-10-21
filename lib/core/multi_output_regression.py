print('Loading required libraries')
try:
    import pandas as pd
    import numpy as np
    import sys
    import time
    import openpyxl as op
    import tensorflow.keras as keras
    import datetime as dt
    import os
    import gzip

    # load the Regressors
    from sklearn.neighbors import KNeighborsRegressor
    from sklearn.linear_model import Lasso
    from sklearn.linear_model import ElasticNet
    from sklearn.linear_model import LinearRegression
    from sklearn.linear_model import RidgeCV

    # from sklearn.ensemble import ExtraTreesRegressor  # no need for that here
    from sklearn.ensemble import GradientBoostingRegressor
    from sklearn.ensemble import RandomForestRegressor
    from sklearn.ensemble import AdaBoostRegressor

    from sklearn.tree import DecisionTreeRegressor

    from sklearn.multioutput import MultiOutputRegressor

    # to save the models
    import joblib

    # now import the regression related metric scores
    from sklearn.metrics import mean_absolute_error, mean_squared_error, max_error

    # load the mat file loader
    from scipy.io import loadmat, savemat


except IOError:
    print('.. Unable to successfully load the necessary libraries.')
    print('.. Terminating program')
    sys.exit(1)
print('.. Required libraries loading was successful.')


FLAG_LINREG = 'LinearRegression'
FLAG_RIDGE = 'Ridge'
FLAG_LASSO = 'Lasso'
FLAG_DTR = 'DecisionTreeRegressor'
FLAG_RFR = 'RandomForestRegressor'
FLAG_MOP_GBR = 'Multi_OP_GBR'
FLAG_MOP_ADAB = 'Multi_OP_AdaB'
FLAG_KNN = 'KNN'
FLAG_DNN = 'DNN'
FLAG_LSTM = 'SIMPLE_LSTM'
FLAG_S2S_LSTM = 'S2S_LSTM'

LIST_WITH_MODEL_FLAGS = [FLAG_KNN, FLAG_LINREG, FLAG_RIDGE, FLAG_LASSO, FLAG_RFR, FLAG_DTR,
                         FLAG_MOP_GBR, FLAG_MOP_ADAB, FLAG_DNN, FLAG_LSTM, FLAG_S2S_LSTM]


def MachineLearning(input_data_train_val, output_data_train_val, input_data_test, output_data_test,
                    name='', path='', dnn_LactFunc='selu', dnn_OactFunc='sigmoid',
                    dnn_loss='mse', lstm_LactFunc='tanh', lstm_DactFunc='sigmoid',
                    lstm_loss='mean_squared_error', lstm_optimizer='adam', epochs=100, batch_size=100,
                    min_lr=0.001):
    # Remember input_data should have the following form:
    # m_paradigms X n_feature_values_per_paradigm
    # Be careful:
    # a) Data must be Normalized
    # b) They should not have nan
    current_datetime = dt.datetime.now().strftime("%d%m%Y_%H%M%S")  # take the current datetime (for folder creation)
    exportFileName_path = name + '_PerfRes.xlsx'  # Export the Performance Scores of all methods
    str_list_model_paths = []  # Create a list to store the paths of the models and return it later
    dir_path = path + '/' + current_datetime + '/' + 'TrainedModels' + '/'
    dir_path = os.path.normpath(dir_path) + '/'
    if not os.path.exists(dir_path):
        os.makedirs(dir_path)
    workbook_path = dir_path + "../" + exportFileName_path

    # convert from dict to array
    inputDataToUse_train_val = input_data_train_val
    outputDataToUse_train_val = output_data_train_val

    inputDataToUse_test = input_data_test
    outputDataToUse_test = output_data_test

    input_shape_train_val = inputDataToUse_train_val.shape
    output_shape_train_val = outputDataToUse_train_val.shape
    print(input_shape_train_val)
    print(output_shape_train_val)

    input_shape_test = inputDataToUse_test.shape
    output_shape_test = outputDataToUse_test.shape
    print(input_shape_test)
    print(output_shape_test)

    # DNN model here
    inputs = keras.Input(shape=(input_shape_train_val[1],))
    # lr_flatten = keras.layers.Flatten()(inputs)
    lr1 = keras.layers.Dense(input_shape_train_val[1] * 2, activation=dnn_LactFunc)(inputs)
    do1 = keras.layers.Dropout(0.2)(lr1)
    lr2 = keras.layers.Dense(input_shape_train_val[1], activation=dnn_LactFunc)(do1)  # decoder
    lr3 = keras.layers.Dense(input_shape_train_val[1] * 2, activation=dnn_LactFunc)(lr2)
    do2 = keras.layers.Dropout(0.2)(lr3)
    outputs = keras.layers.Dense(output_shape_train_val[1], activation=dnn_OactFunc)(do2)
    # outputs = keras.layers.Reshape([28, 28])(lr4)
    DNN = keras.models.Model(inputs, outputs)
    DNN.compile(loss=dnn_loss, optimizer=keras.optimizers.RMSprop())
    # DNN.compile(loss=dnn_loss, optimizer=keras.optimizers.Adadelta())
    # DNN.compile(loss=dnn_loss, optimizer=keras.optimizers.Adam())
    # DNN.compile(loss=dnn_loss, optimizer=keras.optimizers.Nadam())

    # LSTM model here
    dense_layer_size = int(input_shape_train_val[1] * 0.8)
    if dense_layer_size % 2 == 1:
        dense_layer_size += 1
    LSTM_model = keras.Sequential()
    LSTM_model.add(keras.layers.InputLayer(input_shape=(input_shape_train_val[1], 1)))
    # LSTM_model.add(keras.layers.Bidirectional(keras.layers.LSTM(input_shape[1], return_sequences=True),
    #                                           input_shape=(input_shape[1], 1)))
    LSTM_model.add(keras.layers.Dense(dense_layer_size, activation=lstm_LactFunc))
    LSTM_model.add(keras.layers.Reshape((output_shape_train_val[1], -1)))
    LSTM_model.add(keras.layers.Dense(1, activation=lstm_DactFunc))
    LSTM_model.compile(loss=lstm_loss, optimizer=lstm_optimizer)
    LSTM_model.summary()

    # call back options for the deep learning servers
    callbacksOptions = [
        keras.callbacks.EarlyStopping(patience=15, verbose=1),
        keras.callbacks.ReduceLROnPlateau(factor=0.1, patience=5, min_lr=min_lr, verbose=1)]

    # now initialize all of the regressors
    print(' ... instantiating the regressors.')
    # Prepare a dictionary of estimators after instantiating each one of them
    availableEstimators = {
        FLAG_KNN: KNeighborsRegressor(),  # Accept default parameters
        FLAG_LINREG: LinearRegression(),
        FLAG_RIDGE: RidgeCV(),
        FLAG_LASSO: Lasso(),
        FLAG_RFR: RandomForestRegressor(max_depth=4, random_state=2),
        FLAG_DTR: DecisionTreeRegressor(max_depth=5),
        FLAG_MOP_GBR: MultiOutputRegressor(GradientBoostingRegressor(n_estimators=5)),
        FLAG_MOP_ADAB: MultiOutputRegressor(AdaBoostRegressor(n_estimators=5)),
        FLAG_DNN: DNN,
        FLAG_LSTM: LSTM_model,
    }

    # other parameters to setup before start testing
    deepModelsTrainingEpochs = epochs

    # start a for loop is 15 experiments (540 x 15 = 8100 paradigms each time)
    tmpInputs_train_val = inputDataToUse_train_val
    tmpOutputs_train_val = outputDataToUse_train_val

    tmpInputs_test = inputDataToUse_test
    tmpOutputs_test = outputDataToUse_test
    # create train, validation and test sets indexes
    randomIndexes = \
        np.random.permutation(tmpInputs_train_val.shape[0])
    # keep 75% as training, 5% as validation and 20% as training
    trainIdxs = np.sort(randomIndexes[:int(np.round(0.85 * len(randomIndexes)))])
    valIdxs = np.sort(randomIndexes[int(np.round(0.85 * len(randomIndexes))) + 1:])

    # train and test the models
    for modelName, estimator in availableEstimators.items():
        print(' .. training the', modelName, 'model.')

        # training process here and generating predictions for analysis
        if modelName is FLAG_LSTM:
            estimator.fit(np.expand_dims(tmpInputs_train_val[trainIdxs, :], axis=2),
                          np.expand_dims(tmpOutputs_train_val[trainIdxs, :], axis=2),
                          batch_size=batch_size, shuffle=True, epochs=deepModelsTrainingEpochs, verbose=1,
                          callbacks=callbacksOptions,
                          validation_data=(np.expand_dims(tmpInputs_train_val[valIdxs, :], axis=2),
                                           np.expand_dims(tmpOutputs_train_val[valIdxs, :], axis=2)))
            print(' ... training completed.')
            # calculate models outputs for all entries
            trainSetEstimatorPredictions = np.squeeze(
                estimator.predict(np.expand_dims(tmpInputs_train_val[trainIdxs, :], axis=2)), axis=2)
            testSetEstimatorPredictions = np.squeeze(estimator.predict(np.expand_dims(tmpInputs_test, axis=2)),
                                                     axis=2)

        elif modelName is FLAG_DNN:
            estimator.fit(tmpInputs_train_val[trainIdxs, :], tmpOutputs_train_val[trainIdxs, :],
                          batch_size=batch_size, shuffle=True, epochs=deepModelsTrainingEpochs, verbose=1,
                          callbacks=callbacksOptions,
                          validation_data=(tmpInputs_train_val[valIdxs, :], tmpOutputs_train_val[valIdxs, :]))
            print(' ... training completed.')
            # calculate models outputs for all entries
            trainSetEstimatorPredictions = estimator.predict(tmpInputs_train_val[trainIdxs, :])
            testSetEstimatorPredictions = estimator.predict(tmpInputs_test)

        else:
            estimator.fit(tmpInputs_train_val[trainIdxs, :], tmpOutputs_train_val[trainIdxs, :])
            print(' ... training completed.')
            # calculate models outputs for all entries
            trainSetEstimatorPredictions = estimator.predict(tmpInputs_train_val[trainIdxs, :])
            testSetEstimatorPredictions = estimator.predict(tmpInputs_test)
            if (modelName == FLAG_LASSO or modelName == FLAG_RFR or modelName == FLAG_DTR) and tmpOutputs_train_val.shape[1] == 1:
                trainSetEstimatorPredictions = np.expand_dims(trainSetEstimatorPredictions, axis=1)
                testSetEstimatorPredictions = np.expand_dims(testSetEstimatorPredictions, axis=1)

        # this part is the same for all of the models
        # pre-allocating results array for raw predicted values
        errorPerObservationPointMAETrain = np.zeros(tmpOutputs_train_val.shape[1])
        errorPerObservationPointMSETrain = np.zeros(tmpOutputs_train_val.shape[1])
        errorPerObservationPointMaxErrorTrain = np.zeros(tmpOutputs_train_val.shape[1])

        errorPerObservationPointMAETest = np.zeros(tmpOutputs_test.shape[1])
        errorPerObservationPointMSETest = np.zeros(tmpOutputs_test.shape[1])
        errorPerObservationPointMaxErrorTest = np.zeros(tmpOutputs_test.shape[1])

        # calculate related errors
        for observationPointIdx in range(0, tmpOutputs_train_val.shape[1]):
            # normalized first
            errorPerObservationPointMAETrain[observationPointIdx] = \
                mean_absolute_error(tmpOutputs_train_val[trainIdxs, observationPointIdx],
                                    trainSetEstimatorPredictions[:, observationPointIdx])
            errorPerObservationPointMSETrain[observationPointIdx] = \
                mean_squared_error(tmpOutputs_train_val[trainIdxs, observationPointIdx],
                                   trainSetEstimatorPredictions[:, observationPointIdx])
            errorPerObservationPointMaxErrorTrain[observationPointIdx] = \
                max_error(tmpOutputs_train_val[trainIdxs, observationPointIdx],
                          trainSetEstimatorPredictions[:, observationPointIdx])

            errorPerObservationPointMAETest[observationPointIdx] = \
                mean_absolute_error(tmpOutputs_test,
                                    testSetEstimatorPredictions[:, observationPointIdx])
            errorPerObservationPointMSETest[observationPointIdx] = \
                mean_squared_error(tmpOutputs_test,
                                   testSetEstimatorPredictions[:, observationPointIdx])
            errorPerObservationPointMaxErrorTest[observationPointIdx] = \
                max_error(tmpOutputs_test,
                          testSetEstimatorPredictions[:, observationPointIdx])

        # now pass the results to an excel file

        new_row = [modelName]

        tmpScoresList = [list(map(float, errorPerObservationPointMAETrain)),
                         list(map(float, errorPerObservationPointMSETrain)),
                         list(map(float, errorPerObservationPointMaxErrorTrain)),
                         list(map(float, errorPerObservationPointMAETest)),
                         list(map(float, errorPerObservationPointMSETest)),
                         list(map(float, errorPerObservationPointMaxErrorTest)),
                         ]
        tmpScoresList = [item for sublist in tmpScoresList for item in sublist]  # map sublists to fat list
        new_row = new_row + tmpScoresList

        # Confirm file exists.
        # If not, create it, add headers, then append new data
        dir_path = os.path.normpath(dir_path) + '/'
        if not os.path.exists(dir_path):
            os.makedirs(dir_path)
        try:
            wb = op.load_workbook(workbook_path)
            ws = wb.worksheets[0]  # select first worksheet
        except FileNotFoundError:
            headers_row = ['Technique']
            for i in range(0, output_shape_train_val[1]):
                headers_row.append('MAE-Tr-P' + str(i + 1))
            for i in range(0, output_shape_train_val[1]):
                headers_row.append('MSE-Tr-P' + str(i + 1))
            for i in range(0, output_shape_train_val[1]):
                headers_row.append('maxError-Tr-P' + str(i + 1))
            for i in range(0, output_shape_test[1]):
                headers_row.append('MAE-Te-P' + str(i + 1))
            for i in range(0, output_shape_test[1]):
                headers_row.append('MSE-Te-P' + str(i + 1))
            for i in range(0, output_shape_test[1]):
                headers_row.append('maxError-Te-P' + str(i + 1))

            wb = op.Workbook()
            ws = wb.active
            ws.append(headers_row)

        joblib_sav = '.h5'
        h5_sav = '.h5'
        availableEstimators[FLAG_DNN].save(dir_path + FLAG_DNN + '_' + current_datetime + h5_sav)
        str_list_model_paths.append(dir_path + FLAG_DNN + '_' + current_datetime + h5_sav)
        availableEstimators[FLAG_LSTM].save(dir_path + FLAG_LSTM + '_' + current_datetime + h5_sav)
        str_list_model_paths.append(dir_path + FLAG_LSTM + '_' + current_datetime + h5_sav)
        availableEstimators[FLAG_S2S_LSTM].save(dir_path + FLAG_S2S_LSTM + '_' + current_datetime + h5_sav)
        str_list_model_paths.append(dir_path + FLAG_S2S_LSTM + '_' + current_datetime + h5_sav)
        joblib.dump(availableEstimators[FLAG_KNN],
                    dir_path + FLAG_KNN + '_' + current_datetime + joblib_sav)
        str_list_model_paths.append(dir_path + FLAG_KNN + '_' + current_datetime + joblib_sav)
        joblib.dump(availableEstimators[FLAG_LINREG],
                    dir_path + FLAG_LINREG + '_' + current_datetime + joblib_sav)
        str_list_model_paths.append(dir_path + FLAG_LINREG + '_' + current_datetime + joblib_sav)
        joblib.dump(availableEstimators[FLAG_RIDGE],
                    dir_path + FLAG_RIDGE + '_' + current_datetime + joblib_sav)
        str_list_model_paths.append(dir_path + FLAG_RIDGE + '_' + current_datetime + joblib_sav)
        joblib.dump(availableEstimators[FLAG_LASSO],
                    dir_path + FLAG_LASSO + '_' + current_datetime + joblib_sav)
        str_list_model_paths.append(dir_path + FLAG_LASSO + '_' + current_datetime + joblib_sav)
        joblib.dump(availableEstimators[FLAG_RFR],
                    dir_path + FLAG_RFR + '_' + current_datetime + joblib_sav)
        str_list_model_paths.append(dir_path + FLAG_RFR + '_' + current_datetime + joblib_sav)
        joblib.dump(availableEstimators[FLAG_DTR],
                    dir_path + FLAG_DTR + '_' + current_datetime + joblib_sav)
        str_list_model_paths.append(dir_path + FLAG_DTR + '_' + current_datetime + joblib_sav)
        joblib.dump(availableEstimators[FLAG_MOP_GBR],
                    dir_path + FLAG_MOP_GBR + '_' + current_datetime + joblib_sav)
        str_list_model_paths.append(dir_path + FLAG_MOP_GBR + '_' + current_datetime + joblib_sav)
        joblib.dump(availableEstimators[FLAG_MOP_ADAB],
                    dir_path + FLAG_MOP_ADAB + '_' + current_datetime + joblib_sav)
        str_list_model_paths.append(dir_path + FLAG_MOP_ADAB + '_' + current_datetime + joblib_sav)

        ws.append(new_row)
        wb.save(workbook_path)
        time.sleep(1)

    str_list_model_paths = list(dict.fromkeys(str_list_model_paths))  # remove duplicates
    return str_list_model_paths, dir_path


def MachineLearning_Sequential(input_data_train_val, output_data_train_val, input_data_test, output_data_test,
                               name='', path='', dnn_LactFunc='selu', dnn_OactFunc='sigmoid', dnn_loss='mse',
                               lstm_LactFunc='tanh', lstm_DactFunc='sigmoid', css_LactFunc='sigmoid',
                               lstm_loss='mean_squared_error', lstm_optimizer='adam', seq_div=1, epochs=100,
                               batch_size=100, min_lr=0.001, dropout_percentage=0.1, kernel_size=3):
    # Remember input_data should have the following form:
    # m_paradigms X n_feature_values_per_paradigm
    # Be careful:
    # a) Data must be Normalized
    # b) They should not have nan
    current_datetime = dt.datetime.now().strftime("%d%m%Y_%H%M%S")  # take the current datetime (for folder creation)
    exportFileName_path = name + '_PerfRes.xlsx'  # Export the Performance Scores of all methods
    str_list_model_paths = []  # Create a list to store the paths of the models and return it later
    dir_path = path + '/' + current_datetime + '/' + 'TrainedModels' + '/'
    dir_path = os.path.normpath(dir_path) + '/'
    if not os.path.exists(dir_path):
        os.makedirs(dir_path)
    workbook_dir_path = dir_path + "../"
    workbook_path = dir_path + "../" + exportFileName_path

    # ------------------------------------------------------------------------------- #

    # convert from dict to array
    inputDataToUse_train_val = input_data_train_val
    outputDataToUse_train_val = output_data_train_val

    inputDataToUse_test = input_data_test
    outputDataToUse_test = output_data_test

    input_shape_train_val = inputDataToUse_train_val.shape
    output_shape_train_val = outputDataToUse_train_val.shape
    print(input_shape_train_val)
    print(output_shape_train_val)

    input_shape_test = inputDataToUse_test.shape
    output_shape_test = outputDataToUse_test.shape
    print(input_shape_test)
    print(output_shape_test)

    # ------------------------------------------------------------------------------- #

    # DNN model here
    inputs = keras.Input(shape=(input_shape_train_val[1],))
    # lr_flatten = keras.layers.Flatten()(inputs)
    lr1 = keras.layers.Dense(input_shape_train_val[1] * 2, activation=dnn_LactFunc)(inputs)
    do1 = keras.layers.Dropout(0.2)(lr1)
    lr2 = keras.layers.Dense(input_shape_train_val[1], activation=dnn_LactFunc)(do1)  # decoder
    lr3 = keras.layers.Dense(input_shape_train_val[1] * 2, activation=dnn_LactFunc)(lr2)
    do2 = keras.layers.Dropout(0.2)(lr3)
    outputs = keras.layers.Dense(output_shape_train_val[1], activation=dnn_OactFunc)(do2)
    # outputs = keras.layers.Reshape([28, 28])(lr4)
    DNN = keras.models.Model(inputs, outputs)
    DNN.compile(loss=dnn_loss, optimizer=keras.optimizers.RMSprop())
    # DNN.compile(loss=dnn_loss, optimizer=keras.optimizers.Adadelta())
    # DNN.compile(loss=dnn_loss, optimizer=keras.optimizers.Adam())
    # DNN.compile(loss=dnn_loss, optimizer=keras.optimizers.Nadam())

    # LSTM model here
    dense_layer_size = int(input_shape_train_val[1] * 0.8)
    if dense_layer_size % 2 == 1:
        dense_layer_size += 1
    LSTM_model = keras.Sequential()
    LSTM_model.add(keras.layers.InputLayer(input_shape=(input_shape_train_val[1], 1)))
    # LSTM_model.add(keras.layers.Bidirectional(keras.layers.LSTM(input_shape[1], return_sequences=True),
    #                                           input_shape=(input_shape[1], 1)))
    LSTM_model.add(keras.layers.Dense(dense_layer_size, activation=lstm_LactFunc))
    LSTM_model.add(keras.layers.Reshape((output_shape_train_val[1], -1)))
    LSTM_model.add(keras.layers.Dense(1, activation=lstm_DactFunc))
    LSTM_model.compile(loss=lstm_loss, optimizer=lstm_optimizer)
    LSTM_model.summary()

    # call back options for the deep learning servers
    callbacksOptions = [
        keras.callbacks.EarlyStopping(patience=15, verbose=1),
        keras.callbacks.ReduceLROnPlateau(factor=0.1, patience=5, min_lr=min_lr, verbose=1)]

    # Sequential
    inputSeqData_train_val = input_data_train_val.reshape(input_shape_train_val[0], seq_div,
                                                          int(input_shape_train_val[1] / seq_div))
    outputSeqData_train_val = output_data_train_val.reshape(output_shape_train_val[0], seq_div,
                                                            int(output_shape_train_val[1] / seq_div))

    inputSeqData_test = input_data_test.reshape(input_shape_test[0], seq_div, int(input_shape_test[1] / seq_div))
    # outputSeqData_test = output_data_test.reshape(output_shape_test[0], seq_div, int(output_shape_test[1] / seq_div))

    input_seq_shape = inputSeqData_train_val.shape
    output_seq_shape = outputSeqData_train_val.shape

    print(input_seq_shape)
    print(output_seq_shape)

    # conv_size = int((input_seq_shape[1]*input_seq_shape[2]) / 2)
    conv_size = 60
    hidden_units = 90
    Sequential = keras.models.Sequential()
    Sequential.add(keras.layers.InputLayer(input_shape=(input_seq_shape[1], input_seq_shape[2])))
    Sequential.add(keras.layers.Conv1D(conv_size, kernel_size, activation=css_LactFunc))  # convolutional layer
    Sequential.add((keras.layers.SimpleRNN(hidden_units, return_sequences=True)))
    Sequential.add(keras.layers.LSTM(hidden_units * 2, return_sequences=True))
    # Sequential.add(keras.layers.LSTM(hidden_units, return_sequences=True))
    Sequential.add((keras.layers.SimpleRNN(output_seq_shape[1], return_sequences=True)))
    Sequential.add(keras.layers.Dropout(dropout_percentage))
    Sequential.add(keras.layers.Reshape((output_seq_shape[1], -1)))
    Sequential.add(keras.layers.Dense(output_seq_shape[2], activation=lstm_LactFunc))
    Sequential.compile(loss='mean_absolute_error', optimizer=lstm_optimizer)
    print(Sequential.summary())

    # now initialize all of the regressors
    print(' ... instantiating the regressors.')
    # Prepare a dictionary of estimators after instantiating each one of them
    availableEstimators = {
        FLAG_KNN: KNeighborsRegressor(),  # Accept default parameters
        FLAG_LINREG: LinearRegression(),
        FLAG_RIDGE: RidgeCV(),
        FLAG_LASSO: Lasso(),
        FLAG_RFR: RandomForestRegressor(max_depth=4, random_state=2),
        FLAG_DTR: DecisionTreeRegressor(max_depth=5),
        FLAG_MOP_GBR: MultiOutputRegressor(GradientBoostingRegressor(n_estimators=5)),
        FLAG_MOP_ADAB: MultiOutputRegressor(AdaBoostRegressor(n_estimators=5)),
        FLAG_DNN: DNN,
        FLAG_LSTM: LSTM_model,
        FLAG_S2S_LSTM: Sequential
    }

    # ------------------------------------------------------------------------------- #

    # other parameters to setup before start testing
    deepModelsTrainingEpochs = epochs

    # start a for loop is 15 experiments (540 x 15 = 8100 paradigms each time)
    tmpInputs_train_val = inputDataToUse_train_val
    tmpOutputs_train_val = outputDataToUse_train_val

    tmpInputs_test = inputDataToUse_test
    tmpOutputs_test = outputDataToUse_test
    # create train, validation and test sets indexes
    randomIndexes = \
        np.random.permutation(tmpInputs_train_val.shape[0])
    # keep 75% as training, 5% as validation and 20% as training
    trainIdxs = np.sort(randomIndexes[:int(np.round(0.85 * len(randomIndexes)))])
    valIdxs = np.sort(randomIndexes[int(np.round(0.85 * len(randomIndexes))) + 1:])

    # ------------------------------------------------------------------------------- #

    # train and test the models
    for modelName, estimator in availableEstimators.items():
        print(' .. training the', modelName, 'model.')

        # training process here and generating predictions for analysis
        if modelName is FLAG_LSTM:
            estimator.fit(np.expand_dims(tmpInputs_train_val[trainIdxs, :], axis=2),
                          np.expand_dims(tmpOutputs_train_val[trainIdxs, :], axis=2),
                          batch_size=batch_size,
                          shuffle=True,
                          epochs=deepModelsTrainingEpochs,
                          verbose=1,
                          callbacks=callbacksOptions,
                          validation_data=(np.expand_dims(tmpInputs_train_val[valIdxs, :], axis=2),
                                           np.expand_dims(tmpOutputs_train_val[valIdxs, :], axis=2)))
            print(' ... training completed.')
            # calculate models outputs for all entries
            trainSetEstimatorPredictions = np.squeeze(
                estimator.predict(np.expand_dims(tmpInputs_train_val[trainIdxs, :], axis=2)), axis=2)
            testSetEstimatorPredictions = np.squeeze(estimator.predict(np.expand_dims(tmpInputs_test[:, :], axis=2)),
                                                     axis=2)

        elif modelName is FLAG_DNN:
            estimator.fit(tmpInputs_train_val[trainIdxs, :], tmpOutputs_train_val[trainIdxs, :],
                          batch_size=batch_size,
                          shuffle=True,
                          epochs=deepModelsTrainingEpochs,
                          verbose=1,
                          callbacks=callbacksOptions,
                          validation_data=(tmpInputs_train_val[valIdxs, :], tmpOutputs_train_val[valIdxs, :]))
            print(' ... training completed.')
            # calculate models outputs for all entries
            trainSetEstimatorPredictions = estimator.predict(tmpInputs_train_val[trainIdxs, :])
            testSetEstimatorPredictions = estimator.predict(tmpInputs_test[:, :])

        elif modelName is FLAG_S2S_LSTM:
            estimator.fit(inputSeqData_train_val[trainIdxs, :, :],
                          outputSeqData_train_val[trainIdxs, :, :],
                          batch_size=batch_size,
                          shuffle=True,
                          epochs=deepModelsTrainingEpochs,
                          verbose=1,
                          callbacks=callbacksOptions,
                          validation_data=(inputSeqData_train_val[valIdxs, :, :],
                                           outputSeqData_train_val[valIdxs, :, :]))
            print(' ... training completed.')
            # calculate models outputs for all entries
            trainSetEstimatorPredictions = estimator.predict(inputSeqData_train_val[trainIdxs, :, :])
            trainSetShape = trainSetEstimatorPredictions.shape
            trainSetEstimatorPredictions = trainSetEstimatorPredictions.reshape(trainSetShape[0],
                                                                                trainSetShape[1] * trainSetShape[2])
            testSetEstimatorPredictions = estimator.predict(inputSeqData_test[:, :, :])
            testSetShape = testSetEstimatorPredictions.shape
            testSetEstimatorPredictions = testSetEstimatorPredictions.reshape(testSetShape[0],
                                                                              testSetShape[1] * testSetShape[2])
        else:
            estimator.fit(tmpInputs_train_val[trainIdxs, :], tmpOutputs_train_val[trainIdxs, :])
            print(' ... training completed.')
            # calculate models outputs for all entries
            trainSetEstimatorPredictions = estimator.predict(tmpInputs_train_val[trainIdxs, :])
            testSetEstimatorPredictions = estimator.predict(tmpInputs_test[:, :])
            if (modelName == FLAG_LASSO or modelName == FLAG_RFR or
                modelName == FLAG_DTR) and tmpOutputs_train_val.shape[1] == 1 \
                    and tmpOutputs_test.shape[1] == 1:
                trainSetEstimatorPredictions = np.expand_dims(trainSetEstimatorPredictions, axis=1)
                testSetEstimatorPredictions = np.expand_dims(testSetEstimatorPredictions, axis=1)

        # this part is the same for all of the models
        # pre-allocating results array for raw predicted values
        errorPerObservationPointMAETrain = np.zeros(tmpOutputs_train_val.shape[1])
        errorPerObservationPointMSETrain = np.zeros(tmpOutputs_train_val.shape[1])
        errorPerObservationPointMaxErrorTrain = np.zeros(tmpOutputs_train_val.shape[1])

        errorPerObservationPointMAETest = np.zeros(tmpOutputs_test.shape[1])
        errorPerObservationPointMSETest = np.zeros(tmpOutputs_test.shape[1])
        errorPerObservationPointMaxErrorTest = np.zeros(tmpOutputs_test.shape[1])

        # calculate related errors
        for observationPointIdx in range(0, tmpOutputs_train_val.shape[1]):
            # normalized first
            errorPerObservationPointMAETrain[observationPointIdx] = \
                mean_absolute_error(tmpOutputs_train_val[trainIdxs, observationPointIdx],
                                    trainSetEstimatorPredictions[:, observationPointIdx])
            errorPerObservationPointMSETrain[observationPointIdx] = \
                mean_squared_error(tmpOutputs_train_val[trainIdxs, observationPointIdx],
                                   trainSetEstimatorPredictions[:, observationPointIdx])
            errorPerObservationPointMaxErrorTrain[observationPointIdx] = \
                max_error(tmpOutputs_train_val[trainIdxs, observationPointIdx],
                          trainSetEstimatorPredictions[:, observationPointIdx])

        for observationPointIdx in range(0, tmpOutputs_test.shape[1]):
            errorPerObservationPointMAETest[observationPointIdx] = \
                mean_absolute_error(tmpOutputs_test[:, observationPointIdx],
                                    testSetEstimatorPredictions[:, observationPointIdx])
            errorPerObservationPointMSETest[observationPointIdx] = \
                mean_squared_error(tmpOutputs_test[:, observationPointIdx],
                                   testSetEstimatorPredictions[:, observationPointIdx])
            errorPerObservationPointMaxErrorTest[observationPointIdx] = \
                max_error(tmpOutputs_test[:, observationPointIdx],
                          testSetEstimatorPredictions[:, observationPointIdx])

        # now pass the results to an excel file

        new_row = [modelName]

        tmpScoresList = [list(map(float, errorPerObservationPointMAETrain)),
                         list(map(float, errorPerObservationPointMSETrain)),
                         list(map(float, errorPerObservationPointMaxErrorTrain)),
                         list(map(float, errorPerObservationPointMAETest)),
                         list(map(float, errorPerObservationPointMSETest)),
                         list(map(float, errorPerObservationPointMaxErrorTest)),
                         ]
        tmpScoresList = [item for sublist in tmpScoresList for item in sublist]  # map sublists to fat list
        new_row = new_row + tmpScoresList

        # Confirm file exists.
        # If not, create it, add headers, then append new data
        try:
            wb = op.load_workbook(workbook_path)
            ws = wb.worksheets[0]  # select first worksheet
        except FileNotFoundError:
            headers_row = ['Technique']
            for i in range(0, output_shape_train_val[1]):
                headers_row.append('MAE-Tr-P' + str(i + 1))
            for i in range(0, output_shape_train_val[1]):
                headers_row.append('MSE-Tr-P' + str(i + 1))
            for i in range(0, output_shape_train_val[1]):
                headers_row.append('maxError-Tr-P' + str(i + 1))
            for i in range(0, output_shape_test[1]):
                headers_row.append('MAE-Te-P' + str(i + 1))
            for i in range(0, output_shape_test[1]):
                headers_row.append('MSE-Te-P' + str(i + 1))
            for i in range(0, output_shape_test[1]):
                headers_row.append('maxError-Te-P' + str(i + 1))

            wb = op.Workbook()
            ws = wb.active
            ws.append(headers_row)
            wb.save(workbook_path)

        joblib_sav = '.h5'
        h5_sav = '.h5'
        availableEstimators[FLAG_DNN].save(dir_path + FLAG_DNN + '_' + current_datetime + h5_sav)
        str_list_model_paths.append(dir_path + FLAG_DNN + '_' + current_datetime + h5_sav)
        availableEstimators[FLAG_LSTM].save(dir_path + FLAG_LSTM + '_' + current_datetime + h5_sav)
        str_list_model_paths.append(dir_path + FLAG_LSTM + '_' + current_datetime + h5_sav)
        availableEstimators[FLAG_S2S_LSTM].save(dir_path + FLAG_S2S_LSTM + '_' + current_datetime + h5_sav)
        str_list_model_paths.append(dir_path + FLAG_S2S_LSTM + '_' + current_datetime + h5_sav)
        joblib.dump(availableEstimators[FLAG_KNN],
                    dir_path + FLAG_KNN + '_' + current_datetime + joblib_sav)
        str_list_model_paths.append(dir_path + FLAG_KNN + '_' + current_datetime + joblib_sav)
        joblib.dump(availableEstimators[FLAG_LINREG],
                    dir_path + FLAG_LINREG + '_' + current_datetime + joblib_sav)
        str_list_model_paths.append(dir_path + FLAG_LINREG + '_' + current_datetime + joblib_sav)
        joblib.dump(availableEstimators[FLAG_RIDGE],
                    dir_path + FLAG_RIDGE + '_' + current_datetime + joblib_sav)
        str_list_model_paths.append(dir_path + FLAG_RIDGE + '_' + current_datetime + joblib_sav)
        joblib.dump(availableEstimators[FLAG_LASSO],
                    dir_path + FLAG_LASSO + '_' + current_datetime + joblib_sav)
        str_list_model_paths.append(dir_path + FLAG_LASSO + '_' + current_datetime + joblib_sav)
        joblib.dump(availableEstimators[FLAG_RFR],
                    dir_path + FLAG_RFR + '_' + current_datetime + joblib_sav)
        str_list_model_paths.append(dir_path + FLAG_RFR + '_' + current_datetime + joblib_sav)
        joblib.dump(availableEstimators[FLAG_DTR],
                    dir_path + FLAG_DTR + '_' + current_datetime + joblib_sav)
        str_list_model_paths.append(dir_path + FLAG_DTR + '_' + current_datetime + joblib_sav)
        joblib.dump(availableEstimators[FLAG_MOP_GBR],
                    dir_path + FLAG_MOP_GBR + '_' + current_datetime + joblib_sav)
        str_list_model_paths.append(dir_path + FLAG_MOP_GBR + '_' + current_datetime + joblib_sav)
        joblib.dump(availableEstimators[FLAG_MOP_ADAB],
                    dir_path + FLAG_MOP_ADAB + '_' + current_datetime + joblib_sav)
        str_list_model_paths.append(dir_path + FLAG_MOP_ADAB + '_' + current_datetime + joblib_sav)

        ws.append(new_row)
        wb.save(workbook_path)
        time.sleep(1)

    str_list_model_paths = list(dict.fromkeys(str_list_model_paths))  # remove duplicates
    return str_list_model_paths, dir_path, workbook_dir_path


def loadModel(m_path: str):
    if m_path.__contains__(FLAG_DNN) or m_path.__contains__(FLAG_LSTM) or m_path.__contains__(FLAG_S2S_LSTM):
        model = keras.models.load_model(m_path)
    else:
        model = joblib.load(m_path)
    return model


def predModel(model, dx):
    return model.predict(dx)
